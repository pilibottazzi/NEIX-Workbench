from __future__ import annotations
import logging
import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Tuple
from datetime import datetime

import pandas as pd
import streamlit as st
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak, KeepTogether
)

# ── Colores ───────────────────────────────────────────────────────────────────
INK = "#0F172A"; SUB = "#64748B"; BDR = "#E2E8F0"
BG  = "#F8FAFC"; CRD = "#FFFFFF"; R = "#C8102E"
GRN = "#059669"; BLU = "#2563EB"

RL_INK = colors.HexColor("#0F172A"); RL_SUB = colors.HexColor("#64748B")
RL_BDR = colors.HexColor("#E2E8F0"); RL_BG  = colors.HexColor("#F8FAFC")
RL_RED = colors.HexColor("#C8102E"); RL_GRN = colors.HexColor("#059669")
RL_BLU = colors.HexColor("#2563EB"); RL_WHT = colors.white

LOGO = '<svg viewBox="0 0 60 20" xmlns="http://www.w3.org/2000/svg" style="height:18px;display:inline-block;vertical-align:middle;"><rect width="60" height="20" rx="3" fill="#C8102E"/><text x="5" y="14" font-family="Arial Black,Arial" font-weight="900" font-size="12" letter-spacing="-0.5" fill="white">NEIX</text></svg>'

# ── Config ────────────────────────────────────────────────────────────────────
def _write_config():
    d = Path.home() / ".streamlit"; d.mkdir(exist_ok=True)
    (d / "config.toml").write_text(
        '[theme]\nbase="light"\nbackgroundColor="#F8FAFC"\n'
        'secondaryBackgroundColor="#FFFFFF"\ntextColor="#0F172A"\n'
        'primaryColor="#C8102E"\n[server]\nheadless=true\n'
    )

# ── CSS ───────────────────────────────────────────────────────────────────────
CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');
html,body,.stApp,[data-testid="stAppViewContainer"],[data-testid="stMain"],.main{
  background:#F8FAFC!important;color:#0F172A!important;font-family:'Inter',sans-serif!important;}
[data-testid="stMarkdownContainer"],[data-testid="stMarkdownContainer"]>div,
[data-testid="stVerticalBlock"]{background:transparent!important;color:#0F172A!important;}
[data-testid="stMarkdownContainer"] p,span,div{color:inherit!important;}
[data-testid="stHeader"]{background:transparent!important;}
.block-container{max-width:1380px!important;padding-top:1.8rem!important;padding-bottom:3rem!important;}

/* topbar */
.topbar{display:flex;align-items:center;justify-content:space-between;
  padding-bottom:1.2rem;border-bottom:2px solid #0F172A;margin-bottom:1.8rem;}
.t-left{display:flex;align-items:center;gap:1rem;}
.t-name{font-size:0.92rem;font-weight:700;color:#0F172A;}
.t-sub{font-size:0.74rem;color:#64748B;margin-top:0.05rem;}
.t-right{font-size:0.72rem;color:#94A3B8;text-align:right;line-height:1.8;}

/* upload */
.u-label{font-size:0.64rem;font-weight:700;text-transform:uppercase;
  letter-spacing:0.12em;color:#64748B;margin-bottom:0.45rem;}
.u-hint{font-size:0.71rem;color:#94A3B8;margin-top:0.28rem;}
div[data-testid="stFileUploader"]{background:#fff!important;
  border:1px solid #E2E8F0!important;border-radius:8px!important;}

/* empty */
.empty{text-align:center;padding:4rem 2rem;}
.empty-t{font-size:0.9rem;font-weight:700;color:#0F172A;margin-bottom:0.3rem;}
.empty-s{font-size:0.8rem;color:#64748B;}

/* meta */
.meta{display:flex;gap:0.45rem;flex-wrap:wrap;margin-bottom:1.6rem;}
.chip{font-size:0.69rem;color:#64748B;font-weight:500;padding:0.18rem 0.6rem;
  background:#fff;border:1px solid #E2E8F0;border-radius:5px;}

/* bridge — 6 celdas */
.bridge{display:grid;grid-template-columns:repeat(6,1fr);
  border:1px solid #E2E8F0;border-radius:10px;overflow:hidden;
  background:#fff;margin-bottom:1.6rem;}
.bc{padding:1rem 0.9rem;border-right:1px solid #E2E8F0;position:relative;}
.bc:last-child{border-right:none;}
.bc::before{content:'';position:absolute;top:0;left:0;right:0;height:2.5px;}
.bc.ink::before{background:#0F172A;} .bc.red::before{background:#C8102E;}
.bc.grn::before{background:#059669;} .bc.blu::before{background:#2563EB;}
.bc-l{font-size:0.59rem;font-weight:700;text-transform:uppercase;
  letter-spacing:0.1em;color:#64748B;margin-bottom:0.45rem;}
.bc-v{font-size:0.95rem;font-weight:800;letter-spacing:-0.03em;
  font-family:'JetBrains Mono',monospace;line-height:1.1;}
.bc-s{font-size:0.64rem;color:#64748B;margin-top:0.22rem;}
.c-ink{color:#0F172A;} .c-red{color:#C8102E;} .c-grn{color:#059669;} .c-blu{color:#2563EB;}

/* tabla ejecutiva */
.etbl{width:100%;border-collapse:collapse;font-size:0.83rem;}
.etbl th{font-size:0.59rem;font-weight:700;text-transform:uppercase;letter-spacing:0.09em;
  color:#64748B;padding:0.38rem 0.65rem;text-align:left;border-bottom:1px solid #E2E8F0;white-space:nowrap;}
.etbl th.r{text-align:right;}
.etbl td{padding:0.42rem 0.65rem;border-bottom:1px solid #F1F5F9;color:#0F172A;vertical-align:middle;}
.etbl td.r{text-align:right;font-family:'JetBrains Mono',monospace;font-size:0.77rem;white-space:nowrap;}
.etbl td.sm{font-size:0.77rem;color:#64748B;}
.etbl tr:last-child td{border-bottom:none;}
.etbl tr.tot td{font-weight:700;border-top:1px solid #0F172A;background:#F8FAFC;}
.etbl tr:hover td{background:#F8FAFC;}

/* badges */
.bg{display:inline-block;padding:0.11rem 0.45rem;border-radius:4px;
  font-size:0.68rem;font-weight:700;font-family:'JetBrains Mono',monospace;}
.bg-p{background:#ECFDF5;color:#059669;} .bg-n{background:#FEF2F2;color:#C8102E;}
.bg-0{background:#F1F5F9;color:#64748B;}

/* resumen table (2 col) */
.rtbl{width:100%;border-collapse:collapse;font-size:0.84rem;}
.rtbl td{padding:0.42rem 0;border-bottom:1px solid #F1F5F9;color:#64748B;}
.rtbl td.v{text-align:right;font-family:'JetBrains Mono',monospace;font-size:0.82rem;}
.rtbl tr:last-child td{border-bottom:none;}
.rtbl tr.sep td{border-top:1px solid #E2E8F0;padding-top:0.6rem;margin-top:0.4rem;}
.rtbl tr.bold td{font-weight:700;color:#0F172A;}

/* section label */
.slbl{font-size:0.6rem;font-weight:700;text-transform:uppercase;
  letter-spacing:0.12em;color:#64748B;margin:0 0 0.65rem;}

/* card */
.card{background:#fff;border:1px solid #E2E8F0;border-radius:8px;
  padding:1.2rem 1.4rem;margin-bottom:0.8rem;}

/* tabs */
div[data-testid="stTabs"] button{background:#fff!important;
  border:1px solid #E2E8F0!important;border-radius:6px!important;
  color:#64748B!important;font-weight:600!important;
  font-size:0.78rem!important;padding:0.28rem 0.85rem!important;}
div[data-testid="stTabs"] button[aria-selected="true"]{
  background:#0F172A!important;border-color:#0F172A!important;color:#fff!important;}

/* download */
div[data-testid="stDownloadButton"]>button{
  border-radius:7px!important;background:#0F172A!important;
  color:#fff!important;border:none!important;font-weight:700!important;
  font-family:'Inter',sans-serif!important;font-size:0.79rem!important;
  letter-spacing:0.01em!important;padding:0.5rem 1.2rem!important;}
</style>
"""

# ── Helpers numéricos ─────────────────────────────────────────────────────────
def _n(s) -> float:
    if s is None: return 0.0
    s = re.sub(r"\.(?=\d{3})", "", str(s).strip()).replace(",", ".")
    try: return float(s)
    except: return 0.0

def _ars(v: float, dec=0) -> str:
    s = ("{:,."+str(dec)+"f}").format(abs(v)).replace(",","X").replace(".",",").replace("X",".")
    return ("-$ " if v < 0 else "$ ") + s

def _pct(v: float) -> str: return "{:+.1f}%".format(v)
def _col(v: float) -> str: return GRN if v > 0 else (R if v < 0 else SUB)
def _cls(v: float) -> str: return "c-grn" if v > 0 else ("c-red" if v < 0 else "c-ink")

def _bg(v: float, fn=None) -> str:
    fn = fn or _ars
    cls = "bg-p" if v > 0 else ("bg-n" if v < 0 else "bg-0")
    return '<span class="bg ' + cls + '">' + fn(v) + '</span>'

CATS = {
    "Acciones": {"EDN","GGAL","YPFD","SUPV"},
    "CEDEARs":  {"AXP","GOOGL","MELI","META","NVDA","ADBE","TSLA","GLOB",
                 "BABA","VIST","UNH","SPY","COIN","LAC","NU","SPCE","MSFT"},
    "Bonos":    {"AL30","AL35","GD35"},
    "Fondos":   {"CAUCION","MEGA PES A","FIMA PRE A"},
    "LECAPs":   {"S29G5","S16A5","S12S5","S30S5","S15G5","S31O5"},
}
def _cat(t): 
    for c,ts in CATS.items():
        if t.upper() in ts: return c
    return "Otros"

# ── Parsers ───────────────────────────────────────────────────────────────────
def parse_historico(b: bytes):
    soup = BeautifulSoup(b.decode("utf-8-sig"), "html.parser")
    meta = {"usuario":"","comitente":"","fecha_desde":"","fecha_hasta":""}
    enc = soup.find("div", id="encabezadoExcel")
    if enc:
        rows = enc.find_all("tr")
        if len(rows) >= 2:
            ths = [t.get_text(strip=True) for t in rows[0].find_all("th")]
            tds = [t.get_text(strip=True) for t in rows[1].find_all("td")]
            m = dict(zip(ths,tds))
            meta = {"usuario":m.get("Usuario","").strip(),"comitente":m.get("Comitente","").strip(),
                    "fecha_desde":m.get("Fecha Desde","").strip(),"fecha_hasta":m.get("Fecha Hasta","").strip()}
    summ, movs = [], []
    for box in soup.find_all("div", class_="box box-default"):
        h3 = box.find("h3"); full = h3.get_text(strip=True) if h3 else "?"
        ticker = full.split()[0].upper(); nombre = " ".join(full.split()[1:])
        table = box.find("table");
        if not table: continue
        ta = tu = 0.0; bm = []
        for row in table.find_all("tr")[1:]:
            cells = [td.get_text(strip=True) for td in row.find_all(["th","td"])]
            if not cells: continue
            cpbt = cells[0].upper()
            if cpbt == "TOTAL":
                ta = _n(cells[6]) if len(cells)>6 else 0.0
                tu = _n(cells[8]) if len(cells)>8 else 0.0
            elif len(cells) >= 9:
                mv = {"Ticker":ticker,"Especie":nombre,"Comprobante":cpbt,
                      "Numero":cells[1],"Fecha Concertacion":cells[2],"Fecha Liquidacion":cells[3],
                      "Cantidad":_n(cells[4]),"Precio":_n(cells[5]),
                      "Neto ARS":_n(cells[6]),"Moneda":cells[7],"Neto USD":_n(cells[8])}
                bm.append(mv); movs.append(mv)
        # Compras: COMPRA + SUSCRIPCION (fondos) + COL CAUCION + TOMADOR CAUC
        CPBT_COMPRA = {"COMPRA","SUSCRIPCION","COL CAUCION","TOMADOR CAUC"}
        # Ventas/cobros: VENTA + RESCATE (fondos) + RETIRO (letras al vencimiento) + CIERRE CAUC
        CPBT_VENTA  = {"VENTA","RESCATE","RETIRO","CIERRE CAUC"}
        # Solo DIVIDENDOS reales (acciones/CEDEARs) — RTA/AMORT es amortización de letra, va en ventas
        CPBT_DIV    = {"DIVIDENDOS","CREDITO RTA"}
        # RTA/AMORT: para LECAPs/bonos es cobro de capital+renta → va en cobros, no dividendos
        CPBT_RTA    = {"RTA/AMORT"}

        comp = sum(abs(m["Neto ARS"]) for m in bm if m["Comprobante"] in CPBT_COMPRA and m["Neto ARS"]<0)
        vent = sum(m["Neto ARS"] for m in bm if m["Comprobante"] in CPBT_VENTA  and m["Neto ARS"]>0)
        rta  = sum(m["Neto ARS"] for m in bm if m["Comprobante"] in CPBT_RTA    and m["Neto ARS"]>0)
        divs = sum(m["Neto ARS"] for m in bm if m["Comprobante"] in CPBT_DIV    and m["Neto ARS"]>0)
        qc   = sum(m["Cantidad"] for m in bm if m["Comprobante"] in CPBT_COMPRA and m["Cantidad"]>0)
        qv   = sum(abs(m["Cantidad"]) for m in bm if m["Comprobante"] in CPBT_VENTA  and m["Cantidad"]<0)
        # Para RTA/AMORT sin cantidad (letras): no afecta qty, es solo flujo de caja
        summ.append({"Ticker":ticker,"Especie":nombre,"Categoria":_cat(ticker),
                     "Qty Comprada":qc,"Qty Vendida":qv,"Saldo":qc-qv,
                     "Compras ARS":comp,"Ventas ARS":vent,
                     "Rentas/Amort ARS":rta,"Dividendos ARS":divs,
                     "PnL ARS":ta,"PnL USD":tu,"N Movs":len(bm)})
    return pd.DataFrame(summ), pd.DataFrame(movs), meta

def parse_resultados(b: bytes):
    soup = BeautifulSoup(b.decode("utf-8-sig"), "html.parser")
    table = soup.find("table", class_="table-consultas")
    if table is None: return pd.DataFrame()
    records = []
    for row in table.find("tbody").find_all("tr"):
        cells = [td.get_text(strip=True) for td in row.find_all(["td","th"])]
        if not cells or len(cells)<7 or not cells[1].strip(): continue
        raw = cells[0].strip()
        if raw.lower() in {"perdida","ganancia","total",""}: continue
        parts = raw.split(); ticker = parts[0].upper(); nombre = " ".join(parts[1:])
        inv=_n(cells[3]); val=_n(cells[5]); dif=_n(cells[6])
        records.append({"Ticker":ticker,"Especie":nombre,"Categoria":_cat(ticker),
                        "Cantidad":_n(cells[1]),"PPP":_n(cells[2]),"Inversion":inv,
                        "Precio Actual":_n(cells[4]),"Valuacion":val,"Diferencia":dif,
                        "Rend %":(dif/inv*100 if inv else 0.0)})
    return pd.DataFrame(records)

# ── PDF ───────────────────────────────────────────────────────────────────────
def build_pdf(df_h, df_movs, df_r, meta):
    buf = BytesIO()
    W_PAGE = A4[0] - 4*cm
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2.2*cm, bottomMargin=2*cm)

    def sty(**kw):
        d = dict(fontName="Helvetica", fontSize=9, textColor=RL_INK, leading=13)
        d.update(kw); return ParagraphStyle("s_"+str(id(kw)), **d)

    S  = sty(); SR = sty(alignment=TA_RIGHT)
    SB = sty(fontName="Helvetica-Bold")
    SBR= sty(fontName="Helvetica-Bold", alignment=TA_RIGHT)
    SS = sty(fontSize=7.5, textColor=RL_SUB)
    SC = sty(fontName="Courier", fontSize=8.5, alignment=TA_RIGHT)
    SCG= sty(fontName="Courier-Bold", fontSize=8.5, textColor=RL_GRN, alignment=TA_RIGHT)
    SCR= sty(fontName="Courier-Bold", fontSize=8.5, textColor=RL_RED, alignment=TA_RIGHT)
    SCB= sty(fontName="Courier-Bold", fontSize=8.5, textColor=RL_BLU, alignment=TA_RIGHT)
    SH = sty(fontName="Helvetica-Bold", fontSize=7, textColor=RL_SUB)
    SHR= sty(fontName="Helvetica-Bold", fontSize=7, textColor=RL_SUB, alignment=TA_RIGHT)

    def p(txt, s=None): return Paragraph(str(txt), s or S)
    def sp(n=6): return Spacer(1, n)
    def hr(): return HRFlowable(width="100%", thickness=0.5, color=RL_BDR, spaceAfter=6, spaceBefore=6)

    def num(v, bold=False, usd=False):
        s_str = _ars(v, 2 if usd else 0)
        if v > 0: return p(s_str, SCG if bold else sty(fontName="Courier",fontSize=8.5,textColor=RL_GRN,alignment=TA_RIGHT))
        if v < 0: return p(s_str, SCR if bold else sty(fontName="Courier",fontSize=8.5,textColor=RL_RED,alignment=TA_RIGHT))
        return p(s_str, SC)

    def tbl_style(total=True):
        cmds = [
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),7),
            ("TEXTCOLOR",(0,0),(-1,0),RL_SUB),("LINEBELOW",(0,0),(-1,0),0.5,RL_BDR),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[RL_WHT,RL_BG]),
            ("FONTSIZE",(0,1),(-1,-1),8),("TEXTCOLOR",(0,1),(-1,-1),RL_INK),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("GRID",(0,0),(-1,-1),0.25,RL_BDR),
        ]
        if total:
            cmds += [("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
                     ("LINEABOVE",(0,-1),(-1,-1),0.75,RL_INK),
                     ("BACKGROUND",(0,-1),(-1,-1),RL_BG)]
        return TableStyle(cmds)

    story = []

    # Header
    logo_row = Table([[p("NEIX", sty(fontName="Helvetica-Bold",fontSize=18,textColor=RL_RED)),
                       p("Informe de Portafolio", sty(fontSize=8,textColor=RL_SUB,alignment=TA_RIGHT))]],
                     colWidths=[W_PAGE*0.5, W_PAGE*0.5])
    logo_row.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                                   ("LINEBELOW",(0,0),(-1,0),2,RL_INK),
                                   ("BOTTOMPADDING",(0,0),(-1,0),8)]))
    story += [logo_row, sp(14)]
    story.append(p("INFORME DE RENDIMIENTO DE PORTAFOLIO",
                   sty(fontName="Helvetica-Bold",fontSize=17,textColor=RL_INK,leading=21,spaceAfter=6)))

    info = [["Comitente",meta.get("comitente","—"),"Período",meta.get("fecha_desde","—")+" → "+meta.get("fecha_hasta","—")],
            ["Usuario",meta.get("usuario","—"),"Generado",datetime.now().strftime("%d/%m/%Y")]]
    it = Table(info, colWidths=[W_PAGE*0.15,W_PAGE*0.35,W_PAGE*0.15,W_PAGE*0.35])
    it.setStyle(TableStyle([
        ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8),("TEXTCOLOR",(0,0),(0,-1),RL_SUB),
        ("TEXTCOLOR",(2,0),(2,-1),RL_SUB),("TEXTCOLOR",(1,0),(-1,-1),RL_INK),
        ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
        ("LINEBELOW",(0,-1),(-1,-1),0.5,RL_BDR)]))
    story += [it, sp(20)]

    # 1. Resultados consolidados
    pnl_r  = df_h["PnL ARS"].sum(); cap = df_h["Compras ARS"].sum()
    divs   = df_h["Dividendos ARS"].sum(); vent = df_h["Ventas ARS"].sum()
    rp     = pnl_r/cap*100 if cap else 0.0
    pnl_nr = df_r["Diferencia"].sum() if not df_r.empty else 0.0
    val_h  = df_r["Valuacion"].sum()  if not df_r.empty else 0.0
    inv_ab = df_r["Inversion"].sum()  if not df_r.empty else 0.0
    rp_ab  = pnl_nr/inv_ab*100 if inv_ab else 0.0
    pnl_t  = pnl_r+pnl_nr; rp_t = pnl_t/cap*100 if cap else 0.0

    story.append(p("1. RESULTADOS CONSOLIDADOS",
                   sty(fontName="Helvetica-Bold",fontSize=11,textColor=RL_INK,spaceBefore=0,spaceAfter=6)))
    story.append(hr())

    krows = [
        [p("Capital total invertido",SB), num(cap,True),
         p("Total ventas realizadas",SB), num(vent,True)],
        [p("P&L realizado",SB), num(pnl_r,True),
         p("Dividendos / rentas cobrados",SB), num(divs,True)],
        [p("Rendimiento s/ capital",SB), p(_pct(rp), sty(fontName="Courier-Bold",fontSize=9,
           textColor=RL_GRN if rp>=0 else RL_RED,alignment=TA_RIGHT)),
         p("Especies ganadoras / perdedoras",SB),
         p("{}/{}".format(int((df_h["PnL ARS"]>0).sum()),int((df_h["PnL ARS"]<0).sum())),SC)],
    ]
    if not df_r.empty:
        krows += [
            [p("Inversión posiciones abiertas",SB), num(inv_ab,True),
             p("Valuación a mercado (hoy)",SB), num(val_h,True)],
            [p("P&L no realizado (papel)",SB), num(pnl_nr,True),
             p("Rendimiento s/ abierto",SB), p(_pct(rp_ab),
               sty(fontName="Courier-Bold",fontSize=9,textColor=RL_GRN if rp_ab>=0 else RL_RED,alignment=TA_RIGHT))],
        ]
    krows.append([p("P&L TOTAL COMBINADO",sty(fontName="Helvetica-Bold",fontSize=9.5,textColor=RL_INK)),
                  num(pnl_t,True),
                  p("Rendimiento total s/ capital",SB),
                  p(_pct(rp_t),sty(fontName="Courier-Bold",fontSize=9.5,
                    textColor=RL_GRN if rp_t>=0 else RL_RED,alignment=TA_RIGHT))])
    kt = Table(krows, colWidths=[W_PAGE*0.35,W_PAGE*0.15,W_PAGE*0.35,W_PAGE*0.15])
    kt.setStyle(TableStyle([
        ("ROWBACKGROUNDS",(0,0),(-1,-2),[RL_WHT,RL_BG]),
        ("LINEABOVE",(0,-1),(-1,-1),0.75,RL_INK),
        ("BACKGROUND",(0,-1),(-1,-1),RL_BG),
        ("GRID",(0,0),(-1,-1),0.25,RL_BDR),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
    ]))
    story += [kt, PageBreak()]

    # 2. G/P por especie
    story.append(p("2. GANANCIAS Y PÉRDIDAS REALIZADAS POR ESPECIE",
                   sty(fontName="Helvetica-Bold",fontSize=11,textColor=RL_INK,spaceAfter=4)))
    story.append(hr())
    story.append(p("Resultado neto de todas las operaciones cerradas en el período, agrupadas por categoría.", SS))
    story.append(sp(8))

    for cat in ["Acciones","CEDEARs","Bonos","LECAPs","Fondos","Otros"]:
        dfc = df_h[df_h["Categoria"]==cat]
        if dfc.empty: continue
        story.append(p(cat.upper(), sty(fontName="Helvetica-Bold",fontSize=8,textColor=RL_SUB,
                                        spaceBefore=10,spaceAfter=4)))
        hdr = [[p("Ticker",SH),p("Especie",SH),p("Compradas",SHR),p("Vendidas",SHR),
                p("Saldo",SHR),p("Compras ARS",SHR),p("Ventas ARS",SHR),
                p("Dividendos",SHR),p("P&L ARS",SHR)]]
        rows_gp = []
        for _,r in dfc.sort_values("PnL ARS").iterrows():
            vc = RL_GRN if r["PnL ARS"]>0 else RL_RED
            rows_gp.append([
                p(r["Ticker"],SB), p(str(r["Especie"])[:32],SS),
                p("{:,.0f}".format(r["Qty Comprada"]),SC), p("{:,.0f}".format(r["Qty Vendida"]),SC),
                p("{:,.0f}".format(r["Saldo"]),SC), num(r["Compras ARS"]),
                num(r["Ventas ARS"]), num(r["Dividendos ARS"]),
                p(_ars(r["PnL ARS"]),sty(fontName="Courier-Bold",fontSize=8.5,textColor=vc,alignment=TA_RIGHT)),
            ])
        rows_gp.append([
            p("SUBTOTAL",SB), p(" ",SS),
            p(" ",SC), p(" ",SC), p(" ",SC),
            num(dfc["Compras ARS"].sum(),True), num(dfc["Ventas ARS"].sum(),True),
            num(dfc["Dividendos ARS"].sum(),True),
            p(_ars(dfc["PnL ARS"].sum()), sty(fontName="Courier-Bold",fontSize=8.5,
              textColor=RL_GRN if dfc["PnL ARS"].sum()>=0 else RL_RED,alignment=TA_RIGHT)),
        ])
        cw = [W_PAGE*0.07,W_PAGE*0.22,W_PAGE*0.09,W_PAGE*0.09,W_PAGE*0.07,
              W_PAGE*0.15,W_PAGE*0.13,W_PAGE*0.10,W_PAGE*0.08]
        t = Table(hdr+rows_gp, colWidths=cw, repeatRows=1)
        t.setStyle(tbl_style(True))
        story.append(KeepTogether(t))

    story += [sp(8), hr(),
              Table([[p("TOTAL G/P REALIZADAS",sty(fontName="Helvetica-Bold",fontSize=10,textColor=RL_INK)),
                      p(_ars(pnl_r),sty(fontName="Courier-Bold",fontSize=12,
                        textColor=RL_GRN if pnl_r>=0 else RL_RED,alignment=TA_RIGHT))]],
                    colWidths=[W_PAGE*0.6,W_PAGE*0.4]),
              PageBreak()]

    # 3. Tenencia actual
    if not df_r.empty:
        story.append(p("3. TENENCIA ACTUAL — POSICIÓN ABIERTA",
                       sty(fontName="Helvetica-Bold",fontSize=11,textColor=RL_INK,spaceAfter=4)))
        story.append(hr())
        story.append(p("Posiciones vigentes valuadas a precio de mercado a la fecha del reporte.", SS))
        story.append(sp(8))
        hdr = [[p("Ticker",SH),p("Especie",SH),p("Cat.",SH),p("Cantidad",SHR),
                p("PPP",SHR),p("Inversión",SHR),p("Precio Act.",SHR),
                p("Valuación",SHR),p("Diferencia",SHR),p("Rend %",SHR)]]
        rows_t = []
        for _,r in df_r.sort_values("Diferencia").iterrows():
            vc = RL_GRN if r["Diferencia"]>=0 else RL_RED
            rows_t.append([
                p(r["Ticker"],SB), p(str(r["Especie"])[:28],SS), p(str(r["Categoria"]),SS),
                p("{:,.0f}".format(r["Cantidad"]),SC), p(_ars(r["PPP"],2),SC),
                num(r["Inversion"]), num(r["Precio Actual"]),
                p(_ars(r["Valuacion"]),SCB),
                p(_ars(r["Diferencia"]),sty(fontName="Courier-Bold",fontSize=8.5,textColor=vc,alignment=TA_RIGHT)),
                p(_pct(r["Rend %"]),sty(fontName="Courier-Bold",fontSize=8.5,textColor=vc,alignment=TA_RIGHT)),
            ])
        rows_t.append([
            p("TOTAL",SB), p(" ",SS), p(" ",SS), p(" ",SC), p(" ",SC),
            num(inv_ab,True), p(" ",SC), p(_ars(val_h),SCB),
            p(_ars(pnl_nr),sty(fontName="Courier-Bold",fontSize=8.5,
              textColor=RL_GRN if pnl_nr>=0 else RL_RED,alignment=TA_RIGHT)),
            p(_pct(rp_ab),sty(fontName="Courier-Bold",fontSize=8.5,
              textColor=RL_GRN if rp_ab>=0 else RL_RED,alignment=TA_RIGHT)),
        ])
        cw = [W_PAGE*0.07,W_PAGE*0.20,W_PAGE*0.08,W_PAGE*0.08,W_PAGE*0.10,
              W_PAGE*0.12,W_PAGE*0.10,W_PAGE*0.11,W_PAGE*0.10,W_PAGE*0.04]
        t = Table(hdr+rows_t, colWidths=cw, repeatRows=1)
        t.setStyle(tbl_style(True))
        story += [t, PageBreak()]

    # 4. Movimientos
    story.append(p("4. DETALLE DE MOVIMIENTOS",
                   sty(fontName="Helvetica-Bold",fontSize=11,textColor=RL_INK,spaceAfter=4)))
    story.append(hr())
    story.append(p("Registro completo de operaciones realizadas en el período.", SS))
    story.append(sp(8))
    for ticker in sorted(df_movs["Ticker"].unique()):
        dfm = df_movs[df_movs["Ticker"]==ticker]
        story.append(p(ticker+" — "+str(dfm.iloc[0]["Especie"])[:50],
                       sty(fontName="Helvetica-Bold",fontSize=8,textColor=RL_INK,spaceBefore=10,spaceAfter=3)))
        hdr = [[p("Cpbt.",SH),p("Número",SH),p("F.Concertación",SH),p("F.Liquidación",SH),
                p("Cantidad",SHR),p("Precio",SHR),p("Neto ARS",SHR),p("Moneda",SH),p("Neto USD",SHR)]]
        rows_m = []
        for _,r in dfm.iterrows():
            na=r["Neto ARS"]; nu=r["Neto USD"]
            rows_m.append([
                p(str(r["Comprobante"]),SB), p(str(r["Numero"]),SS),
                p(str(r["Fecha Concertacion"]),SS), p(str(r["Fecha Liquidacion"]),SS),
                p("{:,.2f}".format(r["Cantidad"]) if r["Cantidad"]!=0 else "—",SC),
                p("{:,.3f}".format(r["Precio"]) if r["Precio"]!=0 else "—",SC),
                p(_ars(na),sty(fontName="Courier",fontSize=8,
                  textColor=RL_GRN if na>0 else(RL_RED if na<0 else RL_SUB),alignment=TA_RIGHT)),
                p(str(r["Moneda"]),SS),
                p(_ars(nu,2) if nu!=0 else "—",sty(fontName="Courier",fontSize=8,
                  textColor=RL_GRN if nu>0 else(RL_RED if nu<0 else RL_SUB),alignment=TA_RIGHT)),
            ])
        cw=[W_PAGE*0.10,W_PAGE*0.09,W_PAGE*0.11,W_PAGE*0.11,W_PAGE*0.09,
            W_PAGE*0.11,W_PAGE*0.16,W_PAGE*0.10,W_PAGE*0.13]
        t=Table(hdr+rows_m,colWidths=cw,repeatRows=1)
        t.setStyle(tbl_style(False))
        story.append(KeepTogether(t))

    story += [sp(14), hr(),
              p("NEIX — Informe generado el "+datetime.now().strftime("%d/%m/%Y %H:%M")+
                " — Comitente "+meta.get("comitente","—"),
                sty(fontSize=7.5,textColor=RL_SUB))]

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica",7); canvas.setFillColor(RL_SUB)
        canvas.drawString(2*cm,1.2*cm,"NEIX · Comitente "+meta.get("comitente","—"))
        canvas.drawRightString(A4[0]-2*cm,1.2*cm,"Página "+str(doc.page))
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0); return buf.getvalue()

# ── Excel ─────────────────────────────────────────────────────────────────────
def build_excel(df_h, df_movs, df_r, meta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    CK={"INK":"0F172A","RED":"C8102E","GRN":"059669","SUB":"64748B",
        "BG":"F8FAFC","BDR":"E2E8F0","BLU":"2563EB","WHT":"FFFFFF"}
    thin=Side(style="thin",color=CK["BDR"]); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    med=Side(style="medium",color=CK["INK"]); bdr_t=Border(left=thin,right=thin,top=med,bottom=med)

    def hrow(ws,r,cols,bg=CK["INK"],fg=CK["WHT"]):
        for c,(v,w) in enumerate(cols,1):
            cl=ws.cell(row=r,column=c,value=v)
            cl.font=Font(name="Calibri",bold=True,color=fg,size=9)
            cl.fill=PatternFill("solid",fgColor=bg)
            cl.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cl.border=bdr
            ws.column_dimensions[get_column_letter(c)].width=w

    def vcell(ws,r,c,v,fmt="#,##0",bold=False,bg=CK["WHT"]):
        cl=ws.cell(row=r,column=c,value=v)
        col=CK["GRN"] if isinstance(v,(int,float)) and v>0 else(CK["RED"] if isinstance(v,(int,float)) and v<0 else CK["INK"])
        cl.font=Font(name="Calibri",size=9,bold=bold,color=col)
        cl.number_format=fmt; cl.alignment=Alignment(horizontal="right")
        cl.fill=PatternFill("solid",fgColor=bg); cl.border=bdr

    def tcell(ws,r,c,v,bold=False,color=None,bg=CK["WHT"],align="left"):
        cl=ws.cell(row=r,column=c,value=v)
        cl.font=Font(name="Calibri",size=9,bold=bold,color=color or CK["INK"])
        cl.alignment=Alignment(horizontal=align,vertical="center")
        cl.fill=PatternFill("solid",fgColor=bg); cl.border=bdr

    # Sheet 1: Resumen
    ws1=wb.active; ws1.title="Resumen"; ws1.row_dimensions[1].height=36
    ws1.merge_cells("A1:D1")
    cl=ws1["A1"]; cl.value="NEIX — INFORME DE PORTAFOLIO"
    cl.font=Font(name="Calibri",bold=True,size=14,color=CK["WHT"])
    cl.fill=PatternFill("solid",fgColor=CK["INK"])
    cl.alignment=Alignment(horizontal="left",vertical="center")
    for i,(k,v) in enumerate([("Comitente",meta.get("comitente","—")),("Usuario",meta.get("usuario","—")),
                               ("Período",meta.get("fecha_desde","—")+" → "+meta.get("fecha_hasta","—")),
                               ("Generado",datetime.now().strftime("%d/%m/%Y"))],1):
        ws1.cell(row=2,column=i,value=k+": "+str(v)).font=Font(name="Calibri",size=8,color=CK["SUB"],italic=True)
        ws1.column_dimensions[get_column_letter(i)].width=30
    pnl_r=df_h["PnL ARS"].sum(); cap=df_h["Compras ARS"].sum()
    divs=df_h["Dividendos ARS"].sum(); vent=df_h["Ventas ARS"].sum()
    rp=pnl_r/cap*100 if cap else 0.0
    pnl_nr=df_r["Diferencia"].sum() if not df_r.empty else 0.0
    val_h=df_r["Valuacion"].sum() if not df_r.empty else 0.0
    inv_ab=df_r["Inversion"].sum() if not df_r.empty else 0.0
    rp_ab=pnl_nr/inv_ab*100 if inv_ab else 0.0
    pnl_t=pnl_r+pnl_nr; rp_t=pnl_t/cap*100 if cap else 0.0
    hrow(ws1,3,[("Concepto",38),("Valor",18),("Concepto",38),("Valor",18)])
    kpis=[("Capital total invertido",cap,"Total ventas realizadas",vent),
          ("P&L realizado",pnl_r,"Dividendos / rentas cobrados",divs),
          ("Rendimiento s/ capital",rp/100,"Especies ganadoras/perdedoras",
           "{}/{}".format(int((df_h["PnL ARS"]>0).sum()),int((df_h["PnL ARS"]<0).sum())))]
    if not df_r.empty:
        kpis+=[("Inversión posiciones abiertas",inv_ab,"Valuación a mercado hoy",val_h),
               ("P&L no realizado (papel)",pnl_nr,"Rendimiento s/ abierto",rp_ab/100)]
    kpis.append(("P&L TOTAL COMBINADO",pnl_t,"Rendimiento total s/ capital",rp_t/100))
    for i,(k1,v1,k2,v2) in enumerate(kpis):
        r=4+i; is_last=(i==len(kpis)-1)
        bg=CK["BG"] if is_last else(CK["WHT"] if i%2==0 else CK["BG"])
        tcell(ws1,r,1,k1,bold=is_last,bg=bg)
        if isinstance(v1,float):
            vcell(ws1,r,2,v1,"0.00%" if abs(v1)<1 else "#,##0",bold=is_last,bg=bg)
        else: tcell(ws1,r,2,v1,align="right",bg=bg)
        tcell(ws1,r,3,k2,bold=is_last,bg=bg)
        if isinstance(v2,float):
            vcell(ws1,r,4,v2,"0.00%" if abs(v2)<1 else "#,##0",bold=is_last,bg=bg)
        else: tcell(ws1,r,4,str(v2),align="right",bg=bg)
        if is_last:
            for c in range(1,5): ws1.cell(row=r,column=c).border=bdr_t

    # Sheet 2: GP Realizadas
    ws2=wb.create_sheet("GP Realizadas"); ws2.row_dimensions[1].height=22
    ws2.merge_cells("A1:K1")
    cl=ws2["A1"]; cl.value="GANANCIAS Y PÉRDIDAS REALIZADAS"
    cl.font=Font(name="Calibri",bold=True,size=12,color=CK["WHT"])
    cl.fill=PatternFill("solid",fgColor=CK["INK"])
    cl.alignment=Alignment(horizontal="left",vertical="center")
    cols=[("Ticker",10),("Especie",32),("Cat.",12),("Qty Comprada",13),("Qty Vendida",13),
          ("Saldo",10),("Compras ARS",16),("Ventas ARS",16),("Dividendos ARS",15),
          ("PnL ARS",16),("PnL USD",12)]
    hrow(ws2,2,cols)
    for i,(_,r) in enumerate(df_h.sort_values(["Categoria","PnL ARS"]).iterrows()):
        row=3+i; bg=CK["WHT"] if i%2==0 else CK["BG"]
        tcell(ws2,row,1,r["Ticker"],bold=True,bg=bg)
        tcell(ws2,row,2,r["Especie"],bg=bg)
        tcell(ws2,row,3,r["Categoria"],bg=bg)
        for c,col in enumerate(["Qty Comprada","Qty Vendida","Saldo"],4):
            ws2.cell(row=row,column=c,value=r[col]).number_format="#,##0"
            ws2.cell(row=row,column=c).font=Font(name="Calibri",size=9,color=CK["INK"])
            ws2.cell(row=row,column=c).fill=PatternFill("solid",fgColor=bg)
            ws2.cell(row=row,column=c).border=bdr
            ws2.cell(row=row,column=c).alignment=Alignment(horizontal="right")
        for c,col in enumerate(["Compras ARS","Ventas ARS","Rentas/Amort ARS","Dividendos ARS","PnL ARS","PnL USD"],7):
            vcell(ws2,row,c,r[col],"#,##0" if "USD" not in col else "#,##0.00",bg=bg)
    tr=3+len(df_h)
    for c in range(1,12):
        ws2.cell(row=tr,column=c).fill=PatternFill("solid",fgColor=CK["INK"])
        ws2.cell(row=tr,column=c).font=Font(name="Calibri",bold=True,size=9,color=CK["WHT"])
        ws2.cell(row=tr,column=c).border=bdr_t
        ws2.cell(row=tr,column=c).alignment=Alignment(horizontal="right")
    ws2.cell(row=tr,column=1,value="TOTAL").alignment=Alignment(horizontal="left")
    for c,col in enumerate(["Compras ARS","Ventas ARS","Rentas/Amort ARS","Dividendos ARS","PnL ARS","PnL USD"],7):
        ws2.cell(row=tr,column=c,value=df_h[col].sum())
        ws2.cell(row=tr,column=c).number_format="#,##0" if "USD" not in col else "#,##0.00"
    ws2.freeze_panes="A3"; ws2.auto_filter.ref=ws2.dimensions

    # Sheet 3: Tenencia
    if not df_r.empty:
        ws3=wb.create_sheet("Tenencia Actual"); ws3.row_dimensions[1].height=22
        ws3.merge_cells("A1:J1")
        cl=ws3["A1"]; cl.value="TENENCIA ACTUAL — POSICIÓN ABIERTA"
        cl.font=Font(name="Calibri",bold=True,size=12,color=CK["WHT"])
        cl.fill=PatternFill("solid",fgColor=CK["INK"])
        cl.alignment=Alignment(horizontal="left",vertical="center")
        cols=[("Ticker",10),("Especie",32),("Cat.",12),("Cantidad",12),("PPP",14),
              ("Inversión",16),("Precio Actual",14),("Valuación",16),("Diferencia",14),("Rend %",10)]
        hrow(ws3,2,cols)
        for i,(_,r) in enumerate(df_r.sort_values("Diferencia").iterrows()):
            row=3+i; bg=CK["WHT"] if i%2==0 else CK["BG"]
            tcell(ws3,row,1,r["Ticker"],bold=True,bg=bg)
            tcell(ws3,row,2,r["Especie"],bg=bg); tcell(ws3,row,3,r["Categoria"],bg=bg)
            for c,col,fmt in [(4,"Cantidad","#,##0"),(5,"PPP","#,##0.000"),
                              (6,"Inversion","#,##0"),(7,"Precio Actual","#,##0.000"),
                              (8,"Valuacion","#,##0"),(9,"Diferencia","#,##0")]:
                cl=ws3.cell(row=row,column=c,value=r[col])
                is_pnl=c in{8,9}
                cl.font=Font(name="Calibri",size=9,
                             color=CK["GRN"] if is_pnl and r[col]>0 else(CK["RED"] if is_pnl and r[col]<0 else CK["INK"]))
                cl.number_format=fmt; cl.fill=PatternFill("solid",fgColor=bg)
                cl.border=bdr; cl.alignment=Alignment(horizontal="right")
            cl=ws3.cell(row=row,column=10,value=r["Rend %"]/100)
            cl.number_format="0.00%"
            cl.font=Font(name="Calibri",size=9,color=CK["GRN"] if r["Rend %"]>0 else CK["RED"])
            cl.fill=PatternFill("solid",fgColor=bg); cl.border=bdr
            cl.alignment=Alignment(horizontal="right")
        tr=3+len(df_r)
        for c in range(1,11):
            ws3.cell(row=tr,column=c).fill=PatternFill("solid",fgColor=CK["INK"])
            ws3.cell(row=tr,column=c).font=Font(name="Calibri",bold=True,size=9,color=CK["WHT"])
            ws3.cell(row=tr,column=c).border=bdr_t
            ws3.cell(row=tr,column=c).alignment=Alignment(horizontal="right")
        ws3.cell(row=tr,column=1,value="TOTAL").alignment=Alignment(horizontal="left")
        for c,v,fmt in [(6,inv_ab,"#,##0"),(8,val_h,"#,##0"),(9,pnl_nr,"#,##0"),(10,rp_ab/100,"0.00%")]:
            ws3.cell(row=tr,column=c,value=v).number_format=fmt
        ws3.freeze_panes="A3"; ws3.auto_filter.ref=ws3.dimensions

    # Sheet 4: Movimientos
    ws4=wb.create_sheet("Movimientos"); ws4.row_dimensions[1].height=22
    ws4.merge_cells("A1:I1")
    cl=ws4["A1"]; cl.value="DETALLE DE MOVIMIENTOS"
    cl.font=Font(name="Calibri",bold=True,size=12,color=CK["WHT"])
    cl.fill=PatternFill("solid",fgColor=CK["INK"])
    cl.alignment=Alignment(horizontal="left",vertical="center")
    cols=[("Ticker",10),("Especie",30),("Comprobante",14),("Número",10),
          ("F.Concertación",14),("F.Liquidación",14),("Cantidad",12),
          ("Precio",14),("Neto ARS",16),("Moneda",12),("Neto USD",14)]
    hrow(ws4,2,cols)
    for i,(_,r) in enumerate(df_movs.iterrows()):
        row=3+i; bg=CK["WHT"] if i%2==0 else CK["BG"]
        tcell(ws4,row,1,r["Ticker"],bold=True,bg=bg)
        for c2,col in enumerate(["Especie","Comprobante","Numero","Fecha Concertacion","Fecha Liquidacion"],2):
            tcell(ws4,row,c2,r[col],bg=bg)
        for c2,col,fmt in [(7,"Cantidad","#,##0.##"),(8,"Precio","#,##0.000"),
                           (9,"Neto ARS","#,##0"),(11,"Neto USD","#,##0.00")]:
            vcell(ws4,row,c2,r[col],fmt,bg=bg)
        tcell(ws4,row,10,r["Moneda"],bg=bg)
    ws4.freeze_panes="A3"; ws4.auto_filter.ref=ws4.dimensions

    out=BytesIO(); wb.save(out); out.seek(0); return out.getvalue()

# ── Render ────────────────────────────────────────────────────────────────────
def render():
    _write_config()
    st.set_page_config(page_title="NEIX · Portfolio", page_icon="📊",
                       layout="wide", initial_sidebar_state="collapsed")
    st.markdown(CSS, unsafe_allow_html=True)

    # Topbar
    st.markdown(
        '<div class="topbar"><div class="t-left">' + LOGO +
        '<div><div class="t-name">Rendimiento de Portafolio</div>'
        '<div class="t-sub">P&amp;L realizado &middot; tenencia actual &middot; movimientos &middot; PDF &amp; Excel</div>'
        '</div></div>'
        '<div class="t-right">Acciones &middot; CEDEARs &middot; Bonos<br/>LECAPs &middot; Fondos &middot; Cauciones</div>'
        '</div>', unsafe_allow_html=True)

    # Upload
    c1, c2 = st.columns(2, gap="large")
    with c1:
        st.markdown('<div class="u-label">Histórico por Especie</div>', unsafe_allow_html=True)
        up_h = st.file_uploader("h", type=["xls","xlsx"], label_visibility="collapsed", key="uh")
        st.markdown('<div class="u-hint">Ganancias y pérdidas realizadas &middot; movimientos</div>', unsafe_allow_html=True)
    with c2:
        st.markdown('<div class="u-label">Resultados / Tenencia Actual <span style="font-weight:400;text-transform:none;letter-spacing:0">(opcional)</span></div>', unsafe_allow_html=True)
        up_r = st.file_uploader("r", type=["xls","xlsx"], label_visibility="collapsed", key="ur")
        st.markdown('<div class="u-hint">Posición abierta valuada a precio actual</div>', unsafe_allow_html=True)

    if not up_h:
        st.markdown('<div class="empty"><div class="empty-t">Subí el Histórico por Especie para comenzar</div>'
                    '<div class="empty-s">El archivo de Resultados es opcional</div></div>', unsafe_allow_html=True)
        return

    try:
        df_h, df_movs, meta = parse_historico(up_h.getvalue())
    except Exception as e:
        logger.exception("No se pudo leer el Histórico")
        st.error("No se pudo leer el Histórico: " + str(e)); return

    df_r = pd.DataFrame()
    if up_r:
        try: df_r = parse_resultados(up_r.getvalue())
        except Exception as e:
            logger.exception("No se pudo leer Resultados")
            st.warning("No se pudo leer Resultados: " + str(e))

    if df_h.empty:
        st.warning("El archivo no contiene datos."); return

    # Meta
    chips = ('<span class="chip">&#128100; ' + meta.get("usuario","—") + '</span>'
             '<span class="chip">Comitente ' + meta.get("comitente","—") + '</span>'
             '<span class="chip">' + meta.get("fecha_desde","—") + ' &#8594; ' + meta.get("fecha_hasta","hoy") + '</span>'
             '<span class="chip">' + str(len(df_h)) + ' especies</span>'
             '<span class="chip">' + str(len(df_movs)) + ' movimientos</span>')
    if not df_r.empty:
        chips += '<span class="chip">&#128994; ' + str(len(df_r)) + ' posiciones abiertas</span>'
    st.markdown('<div class="meta">' + chips + '</div>', unsafe_allow_html=True)

    # Números
    pnl_r  = df_h["PnL ARS"].sum(); cap = df_h["Compras ARS"].sum()
    divs   = df_h["Dividendos ARS"].sum()
    rta    = df_h["Rentas/Amort ARS"].sum()
    carry  = divs + rta
    rp = pnl_r/cap*100 if cap else 0.0
    pnl_nr = df_r["Diferencia"].sum()  if not df_r.empty else 0.0
    val_h  = df_r["Valuacion"].sum()   if not df_r.empty else 0.0
    inv_ab = df_r["Inversion"].sum()   if not df_r.empty else 0.0
    rp_ab  = pnl_nr/inv_ab*100         if inv_ab else 0.0
    pnl_t  = pnl_r + pnl_nr; rp_t = pnl_t/cap*100 if cap else 0.0

    # Bridge
    def bc(border_cls, val_cls, label, val_str, sub=""):
        return ('<div class="bc ' + border_cls + '">'
                '<div class="bc-l">' + label + '</div>'
                '<div class="bc-v ' + val_cls + '">' + val_str + '</div>'
                + ('<div class="bc-s">' + sub + '</div>' if sub else '') + '</div>')

    no_r = bc("red" if pnl_nr<0 else "grn", _cls(pnl_nr),
              "P&amp;L no realizado", _ars(pnl_nr), _pct(rp_ab)+" s/abierto") if not df_r.empty \
           else bc("ink","c-ink","P&amp;L no realizado","—","sin datos")
    val_b = bc("blu","c-blu","Valuación hoy", _ars(val_h), str(len(df_r))+" pos.") if not df_r.empty \
            else bc("ink","c-ink","Valuación hoy","—","sin datos")

    st.markdown(
        '<div class="bridge">'
        + bc("ink","c-ink","Capital invertido",_ars(cap),str(len(df_h))+" especies")
        + bc("red" if pnl_r<0 else "grn",_cls(pnl_r),"P&amp;L realizado",_ars(pnl_r),_pct(rp)+" s/capital")
        + bc("grn" if carry>=0 else "red",_cls(carry),"Dividendos / rentas",_ars(carry),"carry cobrado")
        + no_r + val_b
        + bc("red" if pnl_t<0 else "grn",_cls(pnl_t),"P&amp;L total",_ars(pnl_t),_pct(rp_t)+" combinado")
        + '</div>', unsafe_allow_html=True)

    # Tabs — sin gráficos
    tabs = st.tabs(["Resumen", "G/P Realizadas", "Tenencia actual", "Movimientos"])

    # ── Resumen ──
    with tabs[0]:
        col_l, col_r = st.columns(2, gap="large")
        with col_l:
            st.markdown('<div class="slbl">Cuadro de resultados</div>', unsafe_allow_html=True)
            rows = [("Capital total invertido",_ars(cap),INK,False),
                    ("Total ventas realizadas",_ars(df_h["Ventas ARS"].sum()),INK,False),
                    ("Dividendos cobrados (acciones/CEDEARs)",_ars(divs),GRN,False),
                    ("Rentas / amort. cobradas (LECAPs/bonos)",_ars(rta),GRN,False),
                    ("P&amp;L realizado",_ars(pnl_r),_col(pnl_r),True)]
            if not df_r.empty:
                rows += [("Inversión posiciones abiertas",_ars(inv_ab),INK,False),
                         ("Valuación a mercado hoy",_ars(val_h),BLU,False),
                         ("P&amp;L no realizado (papel)",_ars(pnl_nr),_col(pnl_nr),True)]
            rows += [("P&amp;L TOTAL COMBINADO",_ars(pnl_t),_col(pnl_t),True),
                     ("Rendimiento total s/ capital",_pct(rp_t),_col(rp_t),True)]
            trs = "".join(
                '<tr' + (' class="bold"' if bold else '') + '>'
                '<td style="color:' + SUB + '">' + k + '</td>'
                '<td class="v" style="color:' + c + '">' + v + '</td></tr>'
                for k,v,c,bold in rows)
            st.markdown('<div class="card"><table class="rtbl">' + trs + '</table></div>', unsafe_allow_html=True)

        with col_r:
            st.markdown('<div class="slbl">Por categoría</div>', unsafe_allow_html=True)
            cat = df_h.groupby("Categoria").agg(
                N=("Ticker","count"), Capital=("Compras ARS","sum"),
                PnL=("PnL ARS","sum"), Divs=("Dividendos ARS","sum"), Rta=("Rentas/Amort ARS","sum")
            ).reset_index().sort_values("PnL")
            hdr = ('<tr><th>Categoría</th><th class="r">Especies</th>'
                   '<th class="r">Capital</th><th class="r">Rentas/Amort</th><th class="r">Dividendos</th><th class="r">P&amp;L ARS</th></tr>')
            trs = "".join(
                '<tr><td><strong>' + str(r["Categoria"]) + '</strong></td>'
                '<td class="r sm">' + str(int(r["N"])) + '</td>'
                '<td class="r">' + _ars(r["Capital"]) + '</td>'
                '<td class="r" style="color:' + GRN + '">' + _ars(r["Divs"] + r.get("Rta",0)) + '</td>'
                '<td class="r">' + _bg(r["PnL"]) + '</td></tr>'
                for _,r in cat.iterrows())
            trs += ('<tr class="tot"><td>TOTAL</td><td class="r sm">' + str(len(df_h)) + '</td>'
                    '<td class="r">' + _ars(cap) + '</td>'
                    '<td class="r" style="color:' + GRN + '">' + _ars(carry) + '</td>'
                    '<td class="r">' + _bg(pnl_r) + '</td></tr>')
            st.markdown('<div class="card"><table class="etbl">' + hdr + trs + '</table></div>', unsafe_allow_html=True)

    # ── G/P Realizadas ──
    with tabs[1]:
        fa, fb, fc = st.columns([3,1,1], gap="medium")
        with fa:
            cats = sorted(df_h["Categoria"].unique())
            sel = st.multiselect("Categoría", cats, default=cats, key="gpc")
        with fb: solo_neg = st.toggle("Solo perdedoras", key="gpn")
        with fc: solo_pos = st.toggle("Solo ganadoras",  key="gpp")
        dh = df_h[df_h["Categoria"].isin(sel)]
        if solo_neg: dh = dh[dh["PnL ARS"]<0]
        if solo_pos: dh = dh[dh["PnL ARS"]>0]

        hdr = ('<tr><th>Ticker</th><th>Especie</th><th>Cat.</th>'
               '<th class="r">Compradas</th><th class="r">Vendidas</th><th class="r">Saldo</th>'
               '<th class="r">Compras ARS</th><th class="r">Ventas ARS</th>'
               '<th class="r">Rentas/Amort</th><th class="r">Dividendos</th><th class="r">P&amp;L ARS</th></tr>')
        trs = "".join(
            '<tr><td><strong>' + str(r["Ticker"]) + '</strong></td>'
            '<td class="sm">' + str(r["Especie"])[:32] + '</td>'
            '<td class="sm">' + str(r["Categoria"]) + '</td>'
            '<td class="r">' + "{:,.0f}".format(r["Qty Comprada"]) + '</td>'
            '<td class="r">' + "{:,.0f}".format(r["Qty Vendida"]) + '</td>'
            '<td class="r">' + "{:,.0f}".format(r["Saldo"]) + '</td>'
            '<td class="r">' + _ars(r["Compras ARS"]) + '</td>'
            '<td class="r">' + _ars(r["Ventas ARS"]) + '</td>'
            '<td class="r" style="color:' + GRN + '">' + _ars(r.get("Rentas/Amort ARS",0)) + '</td>'
            '<td class="r" style="color:' + GRN + '">' + _ars(r["Dividendos ARS"]) + '</td>'
            '<td class="r">' + _bg(r["PnL ARS"]) + '</td></tr>'
            for _,r in dh.sort_values("PnL ARS").iterrows())
        trs += ('<tr class="tot"><td colspan="6">TOTAL ' + str(len(dh)) + ' especies</td>'
                '<td class="r">' + _ars(dh["Compras ARS"].sum()) + '</td>'
                '<td class="r">' + _ars(dh["Ventas ARS"].sum()) + '</td>'
                '<td class="r" style="color:' + GRN + '">' + _ars(dh["Rentas/Amort ARS"].sum()) + '</td>'
                '<td class="r" style="color:' + GRN + '">' + _ars(dh["Dividendos ARS"].sum()) + '</td>'
                '<td class="r">' + _bg(dh["PnL ARS"].sum()) + '</td></tr>')
        st.markdown('<div class="card" style="overflow-x:auto"><table class="etbl">' + hdr + trs + '</table></div>', unsafe_allow_html=True)

    # ── Tenencia actual ──
    with tabs[2]:
        if df_r.empty:
            st.markdown('<div class="empty"><div class="empty-t">Subí el archivo de Resultados</div>'
                        '<div class="empty-s">Para ver la posición abierta valorizada a precio actual</div></div>', unsafe_allow_html=True)
        else:
            hdr = ('<tr><th>Ticker</th><th>Especie</th><th>Cat.</th>'
                   '<th class="r">Cantidad</th><th class="r">PPP</th>'
                   '<th class="r">Inversión</th><th class="r">Precio Actual</th>'
                   '<th class="r">Valuación</th><th class="r">Diferencia</th><th class="r">Rend %</th></tr>')
            trs = "".join(
                '<tr><td><strong>' + str(r["Ticker"]) + '</strong></td>'
                '<td class="sm">' + str(r["Especie"])[:28] + '</td>'
                '<td class="sm">' + str(r["Categoria"]) + '</td>'
                '<td class="r">' + "{:,.0f}".format(r["Cantidad"]) + '</td>'
                '<td class="r">' + _ars(r["PPP"],2) + '</td>'
                '<td class="r">' + _ars(r["Inversion"]) + '</td>'
                '<td class="r">' + _ars(r["Precio Actual"],2) + '</td>'
                '<td class="r" style="color:' + BLU + '">' + _ars(r["Valuacion"]) + '</td>'
                '<td class="r">' + _bg(r["Diferencia"]) + '</td>'
                '<td class="r" style="color:' + _col(r["Rend %"]) + '">' + _pct(r["Rend %"]) + '</td></tr>'
                for _,r in df_r.sort_values("Diferencia").iterrows())
            trs += ('<tr class="tot"><td colspan="5">TOTAL</td>'
                    '<td class="r">' + _ars(inv_ab) + '</td><td></td>'
                    '<td class="r" style="color:' + BLU + '">' + _ars(val_h) + '</td>'
                    '<td class="r">' + _bg(pnl_nr) + '</td>'
                    '<td class="r" style="color:' + _col(rp_ab) + '">' + _pct(rp_ab) + '</td></tr>')
            st.markdown('<div class="card" style="overflow-x:auto"><table class="etbl">' + hdr + trs + '</table></div>', unsafe_allow_html=True)

    # ── Movimientos ──
    with tabs[3]:
        fa, fb = st.columns([2,1], gap="medium")
        with fa: tf = st.selectbox("Especie", ["Todas"] + sorted(df_h["Ticker"].unique()), key="mvt")
        with fb: cpbt_f = st.selectbox("Tipo", ["Todos"] + sorted(df_movs["Comprobante"].unique()), key="mvc")
        mv = df_movs.copy()
        if tf != "Todas": mv = mv[mv["Ticker"]==tf]
        if cpbt_f != "Todos": mv = mv[mv["Comprobante"]==cpbt_f]

        hdr = ('<tr><th>Ticker</th><th>Comprobante</th><th>Número</th>'
               '<th>F. Concertación</th><th>F. Liquidación</th>'
               '<th class="r">Cantidad</th><th class="r">Precio</th>'
               '<th class="r">Neto ARS</th><th>Moneda</th><th class="r">Neto USD</th></tr>')
        trs = "".join(
            '<tr><td><strong>' + str(r["Ticker"]) + '</strong></td>'
            '<td>' + str(r["Comprobante"]) + '</td>'
            '<td class="sm">' + str(r["Numero"]) + '</td>'
            '<td class="sm">' + str(r["Fecha Concertacion"]) + '</td>'
            '<td class="sm">' + str(r["Fecha Liquidacion"]) + '</td>'
            '<td class="r">' + ("{:,.2f}".format(r["Cantidad"]) if r["Cantidad"]!=0 else "—") + '</td>'
            '<td class="r">' + ("{:,.3f}".format(r["Precio"]) if r["Precio"]!=0 else "—") + '</td>'
            '<td class="r" style="color:' + _col(r["Neto ARS"]) + '">' + _ars(r["Neto ARS"]) + '</td>'
            '<td class="sm">' + str(r["Moneda"]) + '</td>'
            '<td class="r" style="color:' + _col(r["Neto USD"]) + '">' + (_ars(r["Neto USD"],2) if r["Neto USD"]!=0 else "—") + '</td></tr>'
            for _,r in mv.iterrows())
        st.markdown('<div class="card" style="overflow-x:auto"><table class="etbl">' + hdr + trs + '</table></div>', unsafe_allow_html=True)
        st.caption(str(len(mv)) + " movimientos")

    # Exportar
    st.markdown("<hr style='border:none;border-top:1px solid #E2E8F0;margin:1.5rem 0 1rem'/>", unsafe_allow_html=True)
    e1, e2, _ = st.columns([0.18, 0.18, 0.64], gap="medium")
    with e1:
        try:
            pdf_b = build_pdf(df_h, df_movs, df_r, meta)
            st.download_button("↓  PDF", data=pdf_b,
                               file_name="neix_portafolio_" + meta.get("comitente","") + ".pdf",
                               mime="application/pdf", key="dl_pdf")
        except Exception as ex:
            logger.exception("Error generando PDF")
            st.error("Error PDF: " + str(ex))
    with e2:
        try:
            xl_b = build_excel(df_h, df_movs, df_r, meta)
            st.download_button("↓  Excel", data=xl_b,
                               file_name="neix_portafolio_" + meta.get("comitente","") + ".xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_xl")
        except Exception as ex:
            logger.exception("Error generando Excel")
            st.error("Error Excel: " + str(ex))

render()
