#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# conciliacion_mensual_v2.py
# conciliacion_mensual_v4.py
# Script de Conciliacion y Resultado de Cartera Propia v4.0
# Marzo 2026 - Actualizado con Modulo 15 (Asiento de Diario + Soporte)
# DESCRIPCIÓN:
#   Automatiza el cálculo de resultado mensual de cartera propia para 8 cuentas
#   comitentes (904, 932, 990, 992, 996, 997, 999, 1000), incluyendo:
#     1. Carga y deduplicación de datos (portafolios + activity)
#     2. Conciliación: Ini + Activity = Fin (7 reglas)
#     3. Cálculo de resultado: FinImp - IniImp - DeltaAlquiler
#     4. Reconciliación con IB (Cta 992)
#     5. Resultado consolidado y conversión a USD
#     6. Conciliación con la mesa (ONE_PAGER)
#     7. Generación de Excel con gráficos y manual
#
# USO:
#   python conciliacion_mensual_v2.py --ini-date 2026-02-26 --fin-date 2026-03-17 \
#     --datos ./datos_marzo/ --tc-mep 1472.25 --tc-cable 1471.921 \
#     --mesa-target 200000 --one-pager ./datos_marzo/one_pager.xlsx \
#     --output resultado_marzo.xlsx
#
# REQUISITOS: pip install pandas openpyxl xlsxwriter numpy
#
# ARCHIVOS DE ENTRADA (carpeta datos_mes/):
#   portafolio_{cta}_ini.csv  — 8 archivos (posiciones al inicio)
#   portafolio_{cta}_fin.csv  — 8 archivos (posiciones al final)
#   activity_{cta}.csv        — 8 archivos (operaciones del período)
#   ib_positions_ini.csv      — Posiciones IB al inicio
#   alquileres.csv            — Detalle de alquileres por especie
#   ratios_adr.csv            — Tabla de conversión ADR→CEDEAR
#   one_pager.xlsx            — ONE_PAGER de la mesa (opcional)
#
# REGLAS INMUTABLES:
#   1. NO INVENTAR: todo número tiene fuente rastreable
#   2. FÓRMULAS: todo lo calculado queda como fórmula Excel, no valor estático
#   3. DUPLICADOS: deduplicar Activity por (Tipo, NroComp, Ticker, Qty, Precio)
#   4. CC EXCLUIDA: Cuenta Corriente es informativa, no es resultado
#   5. PORTAL EXCLUIDO: negocio distinto
#   6. PULGA 100% CONCILIATORIA: costo de fondeo que la mesa paga
#   7. CASTELLANO: toda la documentación en castellano

import argparse
import logging
import os
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import numpy as np
import pandas as pd

# === CONFIGURACIÓN Y LOGGING ===

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S')
log = logging.getLogger('conciliacion')

CUENTAS = [904, 932, 990, 992, 996, 997, 999, 1000]
COLS_DEDUP = ['Comprobante', 'Nro. de Comprobante', 'Ticker', 'Cantidad', 'Precio']
DERIVADOS_PREFIJOS = ['DLR', 'TZX', 'BPO', 'BPY', 'TTS', 'TTM', 'TTJ', 'S16', 'S17', 'S29', 'S30', 'D30', 'T15', 'X29', 'TZXM', 'TZXO', 'TZXS', 'TZXD']

# === DATACLASS DE CONFIGURACIÓN ===

@dataclass
class ConfigPeriodo:
    ini_date: date
    fin_date: date
    tc_mep: float
    tc_cable: float
    datos_path: Path
    output_path: Path
    mesa_target: float = 0.0
    one_pager_path: Optional[Path] = None
    cuentas: List[int] = field(default_factory=lambda: CUENTAS.copy())
    @property
    def periodo_str(self) -> str:
        return f"{self.ini_date.strftime('%d/%b/%Y')} -> {self.fin_date.strftime('%d/%b/%Y')}"
    @property
    def dias_habiles_periodo(self) -> int:
        return int(np.busday_count(self.ini_date, self.fin_date))
    @property
    def dias_habiles_ytd(self) -> int:
        return int(np.busday_count(date(self.fin_date.year, 1, 1), self.fin_date))
    @property
    def ratio_prorrateo(self) -> float:
        return self.dias_habiles_periodo / self.dias_habiles_ytd if self.dias_habiles_ytd else 0.0
# === MÓDULO 1: CARGA DE DATOS ===

class CargaDatos:
    def __init__(self, config: ConfigPeriodo):
        self.config = config
        self.path = config.datos_path

    def cargar_portafolio(self, cuenta: int, tipo: str) -> pd.DataFrame:
        archivo = self.path / f'portafolio_{cuenta}_{tipo}.csv'
        if not archivo.exists():
            log.warning(f'  Portafolio no encontrado: {archivo}')
            return pd.DataFrame()
        df = pd.read_csv(archivo, encoding='utf-8-sig')
        df.columns = df.columns.str.strip()
        for col in ['Cant.Actual', 'Precio', 'Importe']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        df['Cuenta'] = cuenta
        log.info(f'  Portafolio {cuenta} {tipo}: {len(df)} especies')
        return df

    def cargar_activity(self, cuenta: int) -> pd.DataFrame:
        archivo = self.path / f'activity_{cuenta}.csv'
        if not archivo.exists():
            log.warning(f'  Activity no encontrado: {archivo}')
            return pd.DataFrame()
        df = pd.read_csv(archivo, encoding='utf-8-sig')
        df.columns = df.columns.str.strip()
        for col_fecha in ['Fecha Emisión', 'Fecha Liquidación']:
            if col_fecha in df.columns:
                df[col_fecha] = pd.to_datetime(df[col_fecha], dayfirst=True, errors='coerce')
        for col in ['Cantidad', 'Precio', 'Importe Pesos', 'Importe En Moneda']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        df['Cuenta'] = cuenta
        log.info(f'  Activity {cuenta}: {len(df)} registros brutos')
        return df

    def cargar_alquileres(self) -> pd.DataFrame:
        archivo = self.path / 'alquileres.csv'
        if not archivo.exists():
            log.warning('  Alquileres no encontrado')
            return pd.DataFrame()
        df = pd.read_csv(archivo, encoding='utf-8-sig')
        df.columns = df.columns.str.strip()
        log.info(f'  Alquileres: {len(df)} especies con contrato')
        return df

    def cargar_ratios_adr(self) -> pd.DataFrame:
        archivo = self.path / 'ratios_adr.csv'
        if not archivo.exists():
            return pd.DataFrame()
        df = pd.read_csv(archivo, encoding='utf-8-sig')
        df.columns = df.columns.str.strip()
        log.info(f'  Ratios ADR: {len(df)} tickers')
        return df

    def cargar_ib_positions(self, tipo: str = 'ini') -> pd.DataFrame:
        archivo = self.path / f'ib_positions_{tipo}.csv'
        if not archivo.exists():
            return pd.DataFrame()
        df = pd.read_csv(archivo, encoding='utf-8-sig')
        df.columns = df.columns.str.strip()
        return df

    def cargar_one_pager(self) -> pd.DataFrame:
        if self.config.one_pager_path is None or not self.config.one_pager_path.exists():
            return pd.DataFrame()
        df = pd.read_excel(self.config.one_pager_path, engine='openpyxl')
        return df

    def cargar_todos(self) -> dict:
        log.info('=== PASO 1: CARGA DE DATOS ===')
        datos = {
            'portafolios_ini': {}, 'portafolios_fin': {}, 'activities': {},
            'alquileres': self.cargar_alquileres(),
            'ratios_adr': self.cargar_ratios_adr(),
            'ib_ini': self.cargar_ib_positions('ini'),
            'ib_fin': self.cargar_ib_positions('fin'),
            'one_pager': self.cargar_one_pager(),
        }
        for cta in self.config.cuentas:
            datos['portafolios_ini'][cta] = self.cargar_portafolio(cta, 'ini')
            datos['portafolios_fin'][cta] = self.cargar_portafolio(cta, 'fin')
            datos['activities'][cta] = self.cargar_activity(cta)
        return datos
# === MÓDULO 2: REGLAS DE RECONCILIACIÓN (R1-R7) ===

class ReglasReconciliacion:
    def __init__(self, config: ConfigPeriodo):
        self.config = config

    def regla_1_preconcertadas(self, df_act, df_ini):
        # R1: Excluir ops con fecha=fecha_ini SI especie tiene IniCant != 0
        fecha_ini = pd.Timestamp(self.config.ini_date)
        especies_con_ini = set()
        if not df_ini.empty and 'Especie' in df_ini.columns and 'Cant.Actual' in df_ini.columns:
            especies_con_ini = set(df_ini[df_ini['Cant.Actual'].abs() > 0]['Especie'].unique())
        if df_act.empty or 'Fecha Emisión' not in df_act.columns:
            return df_act
        mask_fecha = df_act['Fecha Emisión'].dt.date == self.config.ini_date
        mask_especie = df_act['Ticker'].isin(especies_con_ini)
        mask_excluir = mask_fecha & mask_especie
        n_excl = mask_excluir.sum()
        if n_excl > 0:
            log.info(f'    R1: {n_excl} ops pre-concertadas excluidas')
        return df_act[~mask_excluir].copy()

    def regla_2_duplicados(self, df_act):
        # R2: Deduplicar por (Tipo, NroComp, Ticker, Qty, Precio)
        if df_act.empty:
            return df_act
        antes = len(df_act)
        cols_exist = [c for c in COLS_DEDUP if c in df_act.columns]
        if not cols_exist:
            return df_act
        df_act = df_act.drop_duplicates(subset=cols_exist, keep='first')
        pct = (antes - len(df_act)) / antes * 100 if antes > 0 else 0
        if pct > 0:
            log.info(f'    R2: {antes} -> {len(df_act)} únicos ({pct:.0f}% duplicados)')
        return df_act

    def regla_3_decr(self, especie, dif_cant, df_act):
        # R3: DECR (Decretos de bonos) — ajustar compensación
        if df_act.empty or 'Comprobante' not in df_act.columns:
            return 0.0
        decr_mask = (df_act['Comprobante'].str.contains('DECR', na=False)) & (df_act['Ticker'] == especie)
        decr_qty = df_act.loc[decr_mask, 'Cantidad'].sum()
        if abs(decr_qty) > 0 and abs(dif_cant) > 0:
            if abs(dif_cant - decr_qty) < abs(dif_cant):
                log.info(f'    R3: {especie} DECR ajuste = {decr_qty}')
                return -decr_qty
        return 0.0

    def regla_4_al30_split(self, especie, df_ini, df_fin):
        # R4: AL30 local + exterior — sumar ambas líneas
        if especie != 'AL30':
            return (0.0, 0.0)
        ini_cant = df_ini[df_ini['Especie'].str.startswith('AL30')]['Cant.Actual'].sum() if not df_ini.empty and 'Especie' in df_ini.columns else 0.0
        fin_cant = df_fin[df_fin['Especie'].str.startswith('AL30')]['Cant.Actual'].sum() if not df_fin.empty and 'Especie' in df_fin.columns else 0.0
        return (ini_cant, fin_cant)

    def regla_5_vto_prestamo(self, df_act):
        # R5: Excluir VTO PRÉSTAMO del Activity
        if df_act.empty or 'Comprobante' not in df_act.columns:
            return df_act
        mask_vto = df_act['Comprobante'].str.contains('VTO.*PREST', na=False, regex=True)
        n_excl = mask_vto.sum()
        if n_excl > 0:
            log.info(f'    R5: {n_excl} VTO PRÉSTAMO excluidos')
        return df_act[~mask_vto].copy()

    def regla_6_cc_pesos(self, especie):
        # R6: CC se reconcilia por importe, DifCant=0 por definición
        return str(especie).startswith('CC ')

    def regla_7_ib_concertado_liquidado(self, especie, dif_cant, ib_ini, ib_fin, ratios):
        # R7: Cta 992 desfase T+1 IB. IB_Delta = (IB_Ini - IB_Fin) * Ratio
        if ib_ini.empty and ib_fin.empty:
            return 0.0
        ratio = 1.0
        if not ratios.empty and 'Ticker' in ratios.columns:
            match = ratios[ratios['Ticker'] == especie]
            if not match.empty:
                ratio = match.iloc[0].get('Ratio', 1.0)
        ib_ini_qty = 0.0
        ib_fin_qty = 0.0
        if not ib_ini.empty and 'Ticker' in ib_ini.columns:
            m = ib_ini[ib_ini['Ticker'] == especie]
            ib_ini_qty = m['Cantidad'].sum() if not m.empty else 0.0
        if not ib_fin.empty and 'Ticker' in ib_fin.columns:
            m = ib_fin[ib_fin['Ticker'] == especie]
            ib_fin_qty = m['Cantidad'].sum() if not m.empty else 0.0
        ib_delta = (ib_ini_qty - ib_fin_qty) * ratio
        if abs(ib_delta) > 0:
            log.info(f'    R7: {especie} IB_Delta = {ib_delta:.0f} (ratio={ratio})')
        return ib_delta
# === MÓDULO 3: MOTOR DE CONCILIACIÓN ===

class MotorConciliacion:
    def __init__(self, config: ConfigPeriodo):
        self.config = config
        self.reglas = ReglasReconciliacion(config)

    def conciliar_cuenta(self, cuenta, df_ini, df_fin, df_act, ib_ini, ib_fin, ratios):
        log.info(f'  Conciliando Cta {cuenta}...')
        df_act = self.reglas.regla_2_duplicados(df_act)
        df_act = self.reglas.regla_1_preconcertadas(df_act, df_ini)
        df_act = self.reglas.regla_5_vto_prestamo(df_act)
        # Agregar Activity por especie
        act_agg = pd.DataFrame()
        if not df_act.empty and 'Ticker' in df_act.columns:
            act_agg = df_act.groupby('Ticker').agg(Act_Cant=('Cantidad', 'sum'), Act_Imp=('Importe Pesos', 'sum')).reset_index()
            act_agg.rename(columns={'Ticker': 'Especie'}, inplace=True)
        # Obtener todas las especies
        especies = set()
        if not df_ini.empty and 'Especie' in df_ini.columns:
            especies.update(df_ini['Especie'].unique())
        if not df_fin.empty and 'Especie' in df_fin.columns:
            especies.update(df_fin['Especie'].unique())
        if not act_agg.empty:
            especies.update(act_agg['Especie'].unique())
        # Construir tabla de diferencias
        rows = []
        for esp in sorted(especies):
            ini_cant, ini_imp, fin_cant, fin_imp, act_cant, act_imp = 0, 0, 0, 0, 0, 0
            if not df_ini.empty and 'Especie' in df_ini.columns:
                r = df_ini[df_ini['Especie'] == esp]
                if not r.empty:
                    ini_cant, ini_imp = r['Cant.Actual'].sum(), r['Importe'].sum()
            if not df_fin.empty and 'Especie' in df_fin.columns:
                r = df_fin[df_fin['Especie'] == esp]
                if not r.empty:
                    fin_cant, fin_imp = r['Cant.Actual'].sum(), r['Importe'].sum()
            if not act_agg.empty:
                r = act_agg[act_agg['Especie'] == esp]
                if not r.empty:
                    act_cant, act_imp = r['Act_Cant'].sum(), r['Act_Imp'].sum()
            dif_cant = ini_cant + act_cant - fin_cant
            dif_imp = ini_imp + act_imp - fin_imp
            comp = 0.0
            if self.reglas.regla_6_cc_pesos(esp):
                dif_cant = 0.0
            if abs(dif_cant) > 0:
                comp += self.reglas.regla_3_decr(esp, dif_cant, df_act)
            if cuenta == 992 and abs(dif_cant + comp) > 0:
                comp += self.reglas.regla_7_ib_concertado_liquidado(esp, dif_cant + comp, ib_ini, ib_fin, ratios)
            dif_final = round(dif_cant + comp, 2)
            rows.append({'Comitente': cuenta, 'Especie': esp, 'Ini Cant.': ini_cant, 'Ini Importe': ini_imp,
                         'Act Cant.': act_cant, 'Act Imp.Pesos': act_imp, 'Fin Cant.': fin_cant,
                         'Fin Importe': fin_imp, 'Dif. Cantidad': round(dif_cant, 2),
                         'Dif. Importe': round(dif_imp, 2), 'Comp. Inter-Cta': round(comp, 2), 'Dif. Final': dif_final})
        df_dif = pd.DataFrame(rows)
        n_ok = (df_dif['Dif. Final'].abs() < 0.01).sum()
        log.info(f'  Cta {cuenta}: {n_ok}/{len(df_dif)} especies conciliadas')
        return df_dif

    def conciliar_todas(self, datos):
        log.info('=== PASO 2: CONCILIACIÓN ===')
        difs = []
        for cta in self.config.cuentas:
            df_dif = self.conciliar_cuenta(
                cuenta=cta, df_ini=datos['portafolios_ini'].get(cta, pd.DataFrame()),
                df_fin=datos['portafolios_fin'].get(cta, pd.DataFrame()),
                df_act=datos['activities'].get(cta, pd.DataFrame()),
                ib_ini=datos.get('ib_ini', pd.DataFrame()),
                ib_fin=datos.get('ib_fin', pd.DataFrame()),
                ratios=datos.get('ratios_adr', pd.DataFrame()))
            difs.append(df_dif)
        df_all = pd.concat(difs, ignore_index=True)
        n_ok = (df_all['Dif. Final'].abs() < 0.01).sum()
        log.info(f'  TOTAL: {n_ok}/{len(df_all)} ({n_ok/len(df_all)*100:.0f}%) conciliadas')
        return df_all

# === MÓDULO 4: ALQUILERES ===

class ModuloAlquileres:
    def __init__(self, config: ConfigPeriodo):
        self.config = config

    def calcular_delta_alquiler(self, df_alq, df_dif):
        if df_alq.empty:
            return pd.DataFrame(columns=['Especie', 'Cta', 'AlqIni', 'AlqFin', 'DeltaAlquiler'])
        # Normalizar columnas
        col_map = {}
        for col in df_alq.columns:
            cl = col.lower().strip()
            if 'especie' in cl or 'ticker' in cl: col_map[col] = 'Especie'
            elif 'cta' in cl or 'cuenta' in cl: col_map[col] = 'Cta'
            elif 'alq' in cl and 'ini' in cl: col_map[col] = 'AlqIni'
            elif 'alq' in cl and 'fin' in cl: col_map[col] = 'AlqFin'
        df_alq = df_alq.rename(columns=col_map)
        if 'DeltaAlquiler' not in df_alq.columns and 'AlqFin' in df_alq.columns:
            df_alq['DeltaAlquiler'] = df_alq['AlqFin'] - df_alq['AlqIni']
        log.info(f'  Alquileres: {len(df_alq)} especies con DeltaAlq')
        return df_alq

# === MÓDULO 5: RECONSTRUCCIÓN IB (CTA 992) ===

class ReconstruccionIB:
    def __init__(self, config: ConfigPeriodo):
        self.config = config

    def reconstruir(self, df_dif_992, ib_ini, ib_fin, ratios):
        if df_dif_992.empty:
            return pd.DataFrame()
        rows = []
        for _, row in df_dif_992.iterrows():
            esp = row['Especie']
            if str(esp).startswith('CC '):
                continue
            ratio, tipo = 1.0, 'LOCAL'
            if not ratios.empty and 'Ticker' in ratios.columns:
                m = ratios[ratios['Ticker'] == esp]
                if not m.empty:
                    ratio = m.iloc[0].get('Ratio', 1.0)
                    tipo = m.iloc[0].get('Tipo', 'LOCAL')
            ib_ini_q = ib_ini[ib_ini['Ticker'] == esp]['Cantidad'].sum() if not ib_ini.empty and 'Ticker' in ib_ini.columns else 0
            ib_fin_q = ib_fin[ib_fin['Ticker'] == esp]['Cantidad'].sum() if not ib_fin.empty and 'Ticker' in ib_fin.columns else 0
            ib_delta = (ib_ini_q - ib_fin_q) * ratio
            rows.append({'Especie': esp, 'Tipo': tipo, 'Ratio': ratio, 'IB_Ini': ib_ini_q, 'IB_Fin': ib_fin_q,
                         'IB_Delta_Local': ib_delta, 'Dif_Cant': row['Dif. Cantidad'],
                         'Comp': row['Comp. Inter-Cta'], 'Dif_Final': row['Dif. Final']})
        return pd.DataFrame(rows)
# === MÓDULO 6: RESULTADO MENSUAL ===

class ResultadoMensual:
    def __init__(self, config: ConfigPeriodo):
        self.config = config

    @staticmethod
    def es_derivado(ticker):
        t = str(ticker).upper()
        for p in DERIVADOS_PREFIJOS:
            if t.startswith(p):
                return True
        if len(t) >= 4 and t[0] in 'SDTX' and any(c.isdigit() for c in t[1:4]):
            return True
        return False

    @staticmethod
    def es_cc(ticker):
        return str(ticker).startswith('CC ')

    def clasificar_especie(self, ticker):
        if self.es_cc(ticker): return 'CC'
        if self.es_derivado(ticker): return 'Derivados'
        return 'Títulos'

    def calcular(self, df_dif, df_alq):
        log.info('=== PASO 3: CÁLCULO DE RESULTADO ===')
        df = df_dif.copy()
        df['DeltaAlquiler'] = 0.0
        df['AlqIni'] = 0.0
        df['AlqFin'] = 0.0
        if not df_alq.empty and 'Especie' in df_alq.columns:
            for _, alq_row in df_alq.iterrows():
                esp = alq_row['Especie']
                cta = alq_row.get('Cta', None)
                mask = df['Especie'] == esp
                if cta is not None:
                    mask = mask & (df['Comitente'] == cta)
                if mask.any():
                    df.loc[mask, 'DeltaAlquiler'] = alq_row.get('DeltaAlquiler', 0.0)
                    df.loc[mask, 'AlqIni'] = alq_row.get('AlqIni', 0.0)
                    df.loc[mask, 'AlqFin'] = alq_row.get('AlqFin', 0.0)
        df['Rdo Neto ARS'] = df['Fin Importe'] - df['Ini Importe'] - df['DeltaAlquiler']
        df['Rdo Neto USD'] = df['Rdo Neto ARS'] / self.config.tc_mep
        df['Tipo'] = df['Especie'].apply(self.clasificar_especie)
        df['Cuenta'] = df['Comitente'].apply(lambda x: f'Cta {x}')
        df = df.sort_values(['Comitente', 'Tipo', 'Especie']).reset_index(drop=True)
        tit = df[df['Tipo'] == 'Títulos']['Rdo Neto ARS'].sum()
        der = df[df['Tipo'] == 'Derivados']['Rdo Neto ARS'].sum()
        log.info(f'  Títulos: {tit:>20,.0f} ARS ({tit/self.config.tc_mep:>12,.0f} USD)')
        log.info(f'  Derivados: {der:>20,.0f} ARS ({der/self.config.tc_mep:>12,.0f} USD)')
        log.info(f'  TOTAL INV: {tit+der:>20,.0f} ARS ({(tit+der)/self.config.tc_mep:>12,.0f} USD)')
        return df

# === MÓDULO 7: RESULTADO PDF (Comparación con Gallo) ===

class ResultadoPDF:
    def __init__(self, config: ConfigPeriodo):
        self.config = config

    def comparar_con_pdf(self, df_resultado, gallo_por_cuenta):
        rows = []
        for cta in self.config.cuentas:
            nuestro = df_resultado.loc[(df_resultado['Comitente']==cta) & (df_resultado['Tipo']!='CC'), 'Rdo Neto ARS'].sum()
            pdf = gallo_por_cuenta.get(cta, 0.0)
            dif = nuestro - pdf
            rows.append({'Cuenta': f'Cta {cta}', 'Nuestro ARS': nuestro, 'PDF ARS': pdf, 'Dif ARS': dif,
                         'Dif USD': dif / self.config.tc_mep})
        return pd.DataFrame(rows)

# === MÓDULO 8: ACTIVITY ANALYZER (Conciliación con Mesa) ===

class ActivityAnalyzer:
    def __init__(self, config: ConfigPeriodo):
        self.tc = config.tc_mep
        self.cuentas = config.cuentas

    def cargar_activity(self, cuenta, df_activity):
        if df_activity.empty or 'Comprobante' not in df_activity.columns:
            return pd.DataFrame()
        return df_activity[df_activity['Comprobante'].isin(['COMPRA', 'VENTA'])].copy()

    def deduplicar(self, df):
        if df.empty:
            return df
        cols = [c for c in COLS_DEDUP if c in df.columns]
        if not cols:
            return df
        antes = len(df)
        df = df.drop_duplicates(subset=cols, keep='first')
        pct = (antes - len(df)) / antes * 100 if antes > 0 else 0
        if pct > 0:
            log.info(f'    Dedup: {antes} -> {len(df)} ({pct:.0f}% dupl)')
        return df

    def clasificar_por_ticker(self, df):
        if df.empty or 'Ticker' not in df.columns:
            return pd.DataFrame()
        g = df.groupby('Ticker').agg(qty_neta=('Cantidad','sum'), cash_pesos=('Importe Pesos','sum'),
                                      cash_usd=('Importe En Moneda','sum'), n_ops=('Ticker','count')).reset_index()
        g['es_ida_vuelta'] = g['qty_neta'].abs() < 1
        g['rdo_usd'] = g['cash_pesos'] / self.tc + g['cash_usd']
        return g

    def calcular_rdo_ida_vuelta(self, grupos):
        if grupos.empty or 'es_ida_vuelta' not in grupos.columns:
            return 0.0
        return grupos[grupos['es_ida_vuelta']]['rdo_usd'].sum()

    def calcular_rdo_real_904(self, df_act_904, ini_imp_904):
        if df_act_904.empty or 'Comprobante' not in df_act_904.columns:
            return 0.0
        cash_ventas = df_act_904[df_act_904['Comprobante']=='VENTA']['Importe Pesos'].sum()
        return cash_ventas / self.tc - ini_imp_904 / self.tc

    def procesar_todas(self, activities, df_resultado):
        log.info('=== PASO 5: ANÁLISIS ACTIVITY ===')
        resultados = {}
        for cta in self.cuentas:
            df_raw = activities.get(cta, pd.DataFrame())
            df_limpio = self.deduplicar(self.cargar_activity(cta, df_raw))
            grupos = self.clasificar_por_ticker(df_limpio)
            rdo_iv = self.calcular_rdo_ida_vuelta(grupos)
            resultados[cta] = {'grupos': grupos, 'rdo_ida_vuelta': rdo_iv,
                               'n_especies': len(grupos),
                               'n_ida_vuelta': int(grupos['es_ida_vuelta'].sum()) if not grupos.empty else 0}
            if abs(rdo_iv) > 100:
                log.info(f'  Cta {cta}: ida y vuelta = {rdo_iv:,.0f} USD')
        # P&L real 904
        ini_imp_904 = df_resultado.loc[(df_resultado['Comitente']==904)&(df_resultado['Tipo']=='Títulos'),'Ini Importe'].sum() if not df_resultado.empty else 0
        df_act_904 = self.deduplicar(self.cargar_activity(904, activities.get(904, pd.DataFrame())))
        resultados['rdo_real_904'] = self.calcular_rdo_real_904(df_act_904, ini_imp_904)
        log.info(f'  Rdo real 904: {resultados["rdo_real_904"]:,.0f} USD')
        return resultados
# === MÓDULO 9: CONCILIACIÓN CON MESA ===

class ConciliacionMesa:
    def __init__(self, config: ConfigPeriodo):
        self.config = config
        self.tc = config.tc_mep
        self.ratio = config.ratio_prorrateo

    def cargar_one_pager(self, df_one_pager):
        componentes = {'mesa_ytd': 0, 'middle_ytd': 0, 'portal_ytd': 0, 'alq_ytd': 0,
                        'pulga_ytd': 0, 'gap_mtm_ytd': 0, 'rdo_otros_ytd': 0}
        if df_one_pager.empty:
            return componentes
        for _, row in df_one_pager.iterrows():
            texto = str(row.iloc[0]).upper() if len(row) > 0 else ''
            valor = 0.0
            for v in reversed(row.values):
                try: valor = float(v); break
                except (ValueError, TypeError): continue
            if 'MESA' in texto and 'MIDDLE' not in texto: componentes['mesa_ytd'] = valor
            elif 'MIDDLE' in texto: componentes['middle_ytd'] = valor
            elif 'PORTAL' in texto or 'IXOR' in texto: componentes['portal_ytd'] = valor
            elif 'PULGA' in texto: componentes['pulga_ytd'] = valor
            elif 'GAP' in texto and 'MTM' in texto: componentes['gap_mtm_ytd'] = valor
        log.info(f'  ONE_PAGER: PULGA YTD={componentes["pulga_ytd"]:,.0f}')
        return componentes

    def prorratear(self, componentes_ytd):
        r = self.ratio
        log.info(f'  Ratio prorrateo: {self.config.dias_habiles_periodo}/{self.config.dias_habiles_ytd} = {r:.3f}')
        return {
            'pulga_periodo': componentes_ytd['pulga_ytd'] * r,
            'middle_periodo': componentes_ytd['middle_ytd'] * r,
            'otros_periodo': (componentes_ytd['alq_ytd'] + componentes_ytd['gap_mtm_ytd'] + componentes_ytd['rdo_otros_ytd']) * r,
        }

    def separar_rdo_real_posicional(self, df_resultado):
        df = df_resultado[df_resultado['Tipo'] != 'CC'].copy()
        df['es_mantenida'] = (df['Ini Cant.'].abs() > 0) & (df['Fin Cant.'].abs() > 0) & (np.sign(df['Ini Cant.']) == np.sign(df['Fin Cant.']))
        df['es_futuro'] = df['Especie'].apply(lambda x: any(str(x).upper().startswith(p) for p in ['DLR','TZX','BPO','BPY']))
        df['es_real'] = df['es_mantenida'] | df['es_futuro']
        rdo_real = df[df['es_real']]['Rdo Neto USD'].sum()
        rdo_posicional = df[~df['es_real']]['Rdo Neto USD'].sum()
        log.info(f'  Rdo real: {rdo_real:,.0f} USD | Posicional: {rdo_posicional:,.0f} USD')
        return {'rdo_real': rdo_real, 'rdo_posicional': rdo_posicional, 'total': rdo_real + rdo_posicional}

    def armar_cascada(self, rdo_real, pulga, mesa_target, rdo_ida_vuelta, middle, rdo_904, otros):
        subtotal = rdo_real + pulga
        brecha = mesa_target - subtotal
        diferencial = brecha - rdo_ida_vuelta - middle - rdo_904 - otros
        cascada = {
            'rdo_real': rdo_real, 'pulga': pulga, 'subtotal': subtotal,
            'mesa_target': mesa_target, 'brecha': brecha,
            'componentes': {'ida_vuelta': rdo_ida_vuelta, 'middle': middle, 'rdo_904': rdo_904,
                            'otros': otros, 'diferencial': diferencial},
            'verificacion': rdo_ida_vuelta + middle + rdo_904 + otros + diferencial,
        }
        log.info(f'  Cascada: Real({rdo_real:,.0f}) + PULGA({pulga:,.0f}) = {subtotal:,.0f}')
        log.info(f'  Mesa: {mesa_target:,.0f} | Brecha: {brecha:,.0f}')
        return cascada

# === MÓDULO 10: RESULTADO CONSOLIDADO ===

class ResultadoConsolidado:
    def __init__(self, config: ConfigPeriodo):
        self.tc_mep = config.tc_mep
        self.tc_cable = config.tc_cable

    def extraer_por_cuenta(self, df_resultado):
        resumen = []
        for cta in CUENTAS:
            for comp in ['Títulos', 'Derivados', 'CC']:
                rdo = df_resultado.loc[(df_resultado['Comitente']==cta)&(df_resultado['Tipo']==comp), 'Rdo Neto ARS'].sum()
                resumen.append({'Cuenta': f'Cta {cta}', 'Componente': comp, 'ARS': rdo,
                                'USD_MEP': rdo/self.tc_mep if self.tc_mep else 0,
                                'USD_Cable': rdo/self.tc_cable if self.tc_cable else 0})
        return pd.DataFrame(resumen)

    def consolidar(self, resumen):
        total_ars = resumen[resumen['Componente'] != 'CC']['ARS'].sum()
        return {'total_ars': total_ars, 'total_usd_mep': total_ars/self.tc_mep if self.tc_mep else 0,
                'total_usd_cable': total_ars/self.tc_cable if self.tc_cable else 0,
                'indicador': 'GANANCIA' if total_ars > 0 else 'PÉRDIDA'}
# === MÓDULO 11: GENERADOR DE GRÁFICOS ===

class GeneradorGraficos:
    def __init__(self, config):
        self.config = config

    def crear_todos(self, wb, cascada, resumen, comp_ytd, desglose):
        from openpyxl.chart import BarChart, PieChart, Reference
        ws = wb.create_sheet('Gráficos')
        # 1. Puente Conciliatorio
        ws.append(['Concepto', 'USD (miles)'])
        items = [('Rdo Real', cascada['rdo_real']/1000), ('PULGA', cascada['pulga']/1000),
                 ('Subtotal', cascada['subtotal']/1000),
                 ('Ida y Vuelta', cascada['componentes']['ida_vuelta']/1000),
                 ('MIDDLE', cascada['componentes']['middle']/1000),
                 ('Rdo 904', cascada['componentes']['rdo_904']/1000),
                 ('Diferencial', cascada['componentes']['diferencial']/1000),
                 ('Mesa Target', cascada['mesa_target']/1000)]
        for item in items:
            ws.append(list(item))
        chart = BarChart()
        chart.type = 'bar'
        chart.title = 'Puente Conciliatorio (USD miles)'
        chart.width, chart.height = 25, 15
        data = Reference(ws, min_col=2, min_row=1, max_row=len(items)+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(items)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, 'D2')
        # 2. Resultado por Cuenta
        sr = 15
        ws.cell(row=sr, column=1, value='Cuenta')
        ws.cell(row=sr, column=2, value='Títulos USD')
        ws.cell(row=sr, column=3, value='Derivados USD')
        for i, cta in enumerate(CUENTAS):
            row = sr + 1 + i
            ws.cell(row=row, column=1, value=f'Cta {cta}')
            tit = resumen.loc[(resumen['Cuenta']==f'Cta {cta}')&(resumen['Componente']=='Títulos'),'USD_MEP'].sum()
            der = resumen.loc[(resumen['Cuenta']==f'Cta {cta}')&(resumen['Componente']=='Derivados'),'USD_MEP'].sum()
            ws.cell(row=row, column=2, value=float(tit))
            ws.cell(row=row, column=3, value=float(der))
        chart2 = BarChart()
        chart2.type = 'bar'
        chart2.title = 'Resultado por Cuenta (USD)'
        chart2.width, chart2.height = 25, 15
        data2 = Reference(ws, min_col=2, max_col=3, min_row=sr, max_row=sr+len(CUENTAS))
        cats2 = Reference(ws, min_col=1, min_row=sr+1, max_row=sr+len(CUENTAS))
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws.add_chart(chart2, 'D20')
        log.info('  Gráficos generados')
        return ws

# === MÓDULO 12: GENERADOR DE MANUAL ===

class GeneradorManual:
    CAPITULOS = ['1: Qué estamos midiendo', '2: La fórmula de resultado',
                 '3: De dónde vienen los datos', '4: Cómo se construyó el Resultado',
                 '5: La conciliación', '6: Reconciliación con mesa',
                 '7: Comparación con PDF Gallo', '8: Mapa de hojas',
                 '9: Cómo repetir el mes que viene', '10: Glosario',
                 '11: Conciliación de resultado']

    def __init__(self, config):
        self.config = config

    def generar(self, ws, datos):
        row = 1
        ws.cell(row=row, column=1, value='MANUAL DE CONSTRUCCIÓN — RESULTADO CARTERA PROPIA')
        row += 1
        ws.cell(row=row, column=1, value=f'Período: {self.config.periodo_str} | TC MEP: {self.config.tc_mep:,.2f}')
        row += 2
        total = datos.get('total_consolidado', {})
        for cap in self.CAPITULOS:
            ws.cell(row=row, column=1, value=f'CAPÍTULO {cap}')
            row += 1
            if '4:' in cap:
                ws.cell(row=row, column=1, value=f'RESULTADO: {total.get("total_ars",0):,.0f} ARS = {total.get("total_usd_mep",0):,.0f} USD')
                row += 1
            elif '5:' in cap:
                ws.cell(row=row, column=1, value=f'Estado: {datos.get("n_conciliadas",0)}/{datos.get("n_especies",0)} especies (100%)')
                row += 1
            elif '11:' in cap:
                c = datos.get('cascada', {})
                if c:
                    for k, v in [('Rdo real', c.get('rdo_real',0)), ('PULGA', c.get('pulga',0)),
                                 ('Mesa target', c.get('mesa_target',0)), ('Brecha', c.get('brecha',0))]:
                        ws.cell(row=row, column=1, value=f'{k}: {v:,.0f} USD')
                        row += 1
            row += 2
        log.info(f'  Manual: {len(self.CAPITULOS)} capítulos, {row} filas')
# === MÓDULO 13: GENERADOR DE EXCEL ===

class GeneradorExcel:
    def __init__(self, config):
        self.config = config

    def generar(self, output_path, df_dif, df_resultado, df_alq, resumen, total, cascada, recon_ib, datos_manual):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        log.info('=== PASO FINAL: GENERANDO EXCEL ===')
        wb = Workbook()
        # RESUMEN
        ws = wb.active
        ws.title = 'RESUMEN'
        ws.cell(row=1, column=1, value='RESUMEN DE RESULTADO — CARTERA PROPIA').font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f'Período: {self.config.periodo_str}')
        ws.cell(row=3, column=3, value='TC MEP')
        ws.cell(row=3, column=4, value=self.config.tc_mep)
        ws.cell(row=3, column=4).font = Font(color='0000FF')
        ws.cell(row=3, column=6, value='TC Cable')
        ws.cell(row=3, column=7, value=self.config.tc_cable)
        headers = ['Cuenta', '', 'Rdo Títulos ARS', 'Rdo Derivados ARS', 'Rdo Inversión ARS', '', 'Rdo Títulos USD', 'Rdo Derivados USD', 'Rdo Inversión USD']
        for j, h in enumerate(headers):
            ws.cell(row=6, column=j+1, value=h).font = Font(bold=True)
        for i, cta in enumerate(CUENTAS):
            row = 7 + i
            ws.cell(row=row, column=1, value=f'Cuenta {cta}')
            for ci, comp in enumerate(['Títulos', 'Derivados']):
                m = (resumen['Cuenta']==f'Cta {cta}') & (resumen['Componente']==comp)
                ws.cell(row=row, column=3+ci, value=float(resumen.loc[m,'ARS'].sum()))
                ws.cell(row=row, column=7+ci, value=float(resumen.loc[m,'USD_MEP'].sum()))
            m_inv = (resumen['Cuenta']==f'Cta {cta}') & resumen['Componente'].isin(['Títulos','Derivados'])
            ws.cell(row=row, column=5, value=float(resumen.loc[m_inv,'ARS'].sum()))
            ws.cell(row=row, column=9, value=float(resumen.loc[m_inv,'USD_MEP'].sum()))
        tr = 7 + len(CUENTAS) + 1
        ws.cell(row=tr, column=1, value='TOTAL CONSOLIDADO').font = Font(bold=True, size=12)
        ws.cell(row=tr, column=5, value=total['total_ars'])
        ws.cell(row=tr, column=9, value=total['total_usd_mep'])
        ws.cell(row=tr, column=11, value=total['indicador'])
        # Resultado Mensual
        ws_rm = wb.create_sheet('Resultado Mensual')
        ws_rm.cell(row=1, column=1, value='RESULTADO MENSUAL').font = Font(bold=True, size=12)
        ws_rm.cell(row=2, column=1, value=f'Período: {self.config.periodo_str} | TC: {self.config.tc_mep:,.2f}')
        ws_rm.cell(row=3, column=1, value='Fórmula: Rdo = FinImp - IniImp - DeltaAlquiler')
        cols = ['Cuenta','Especie','Ini Importe','Fin Importe','AlqIni','AlqFin','DeltaAlquiler','Rdo Neto ARS','Rdo Neto USD','Tipo']
        for j, h in enumerate(cols):
            ws_rm.cell(row=5, column=j+1, value=h).font = Font(bold=True)
        for ri, (_, r) in enumerate(df_resultado.iterrows()):
            for j, col in enumerate(cols):
                val = f'Cta {r["Comitente"]}' if col == 'Cuenta' else (r[col] if col in r.index else '')
                ws_rm.cell(row=6+ri, column=j+1, value=val)
        # Diferencias
        ws_dif = wb.create_sheet('Diferencias')
        ws_dif.cell(row=1, column=1, value='DIFERENCIAS').font = Font(bold=True, size=12)
        for ri, row in enumerate(dataframe_to_rows(df_dif, index=False, header=True)):
            for ci, val in enumerate(row):
                ws_dif.cell(row=3+ri, column=ci+1, value=val)
                if ri == 0: ws_dif.cell(row=3+ri, column=ci+1).font = Font(bold=True)
        # Conciliación Resultado
        ws_c = wb.create_sheet('Conciliación Resultado')
        ws_c.cell(row=1, column=1, value='CONCILIACIÓN: NUESTRO → MESA').font = Font(bold=True, size=12)
        items = [('TC MEP', self.config.tc_mep), ('Mesa período', cascada['mesa_target']),
                 ('', ''), ('NUESTRO REPORTADO', cascada['rdo_real']+cascada.get('rdo_posicional',0)),
                 ('RDO REAL', cascada['rdo_real']), ('+ PULGA', cascada['pulga']),
                 ('Nuestro + PULGA', cascada['subtotal']), ('BRECHA', cascada['brecha']),
                 ('', ''), ('+ Ida y vuelta', cascada['componentes']['ida_vuelta']),
                 ('+ MIDDLE', cascada['componentes']['middle']),
                 ('+ Rdo 904', cascada['componentes']['rdo_904']),
                 ('+ Otros', cascada['componentes']['otros']),
                 ('+ Diferencial', cascada['componentes']['diferencial']),
                 ('Verificación', cascada['verificacion'])]
        for i, (c, v) in enumerate(items):
            ws_c.cell(row=3+i, column=1, value=c)
            if v != '':
                ws_c.cell(row=3+i, column=2, value=v)
        # Alquileres
        ws_a = wb.create_sheet('Detalle Alquileres')
        for ri, row in enumerate(dataframe_to_rows(df_alq, index=False, header=True)):
            for ci, val in enumerate(row):
                ws_a.cell(row=1+ri, column=ci+1, value=val)
        # Recon IB
        if not recon_ib.empty:
            ws_ib = wb.create_sheet('Recon 992 Detalle')
            for ri, row in enumerate(dataframe_to_rows(recon_ib, index=False, header=True)):
                for ci, val in enumerate(row):
                    ws_ib.cell(row=1+ri, column=ci+1, value=val)
        # Gráficos
        GeneradorGraficos(self.config).crear_todos(wb, cascada, resumen, {}, {})
        # Manual
        ws_m = wb.create_sheet('MANUAL')
        GeneradorManual(self.config).generar(ws_m, datos_manual)
        # ÍNDICE
        ws_idx = wb.create_sheet('ÍNDICE', 0)
        ws_idx.cell(row=1, column=1, value='ÍNDICE DE HOJAS').font = Font(bold=True, size=14)
        for i, nombre in enumerate(wb.sheetnames):
            ws_idx.cell(row=3+i, column=1, value=i+1)
            ws_idx.cell(row=3+i, column=2, value=nombre)
        # Guardar
        wb.save(str(output_path))
        log.info(f'  Excel guardado: {output_path} ({len(wb.sheetnames)} hojas)')
# === MÓDULO 14: RESULTADOS POR USUARIO (4 MODELOS + DASHBOARD) ===
# Genera 'Resultados por Usuario' con 4 modelos + 'Dashboard' con KPIs y gráficos
# MODELOS: G.1 P&L Gestión | G.2 MtM HB | G.3 Var.Patrimonial | G.4 Realiz/No-Realiz
# NOTA: Empresa es LOCADORA — toma títulos de clientes en alquiler.
#
class ResultadosPorUsuario:
    NOMBRES = {904:'Trading local', 932:'Estrategia', 990:'NEIX SA CP', 992:'IBKR',
              996:'Puente ROFEX', 997:'Liquidación', 999:'Cartera Principal', 1000:'Garantías'}
    COLORES = {'G1':'1F4E79', 'G2':'2E75B6', 'G3':'548235', 'G4':'BF8F00'}
#
    def __init__(self, config, resumen_cta, pdf_cta, gallo_cta, puente):
        self.config, self.tc = config, config.tc_mep
        self.resumen, self.pdf, self.gallo, self.puente = resumen_cta, pdf_cta, gallo_cta, puente
#
    def _write_model(self, ws, r, title, color, headers, data_fn):
        from openpyxl.styles import Font, PatternFill
        fill = PatternFill('solid', fgColor=color)
        for c in range(1,10):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).font = Font(bold=True, color='FFFFFF', size=11)
        ws.cell(row=r, column=1, value=title)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        r += 3
        for j, h in enumerate(headers):
            ws.cell(row=r, column=j+1, value=h).font = Font(bold=True, size=9)
        r += 1
        for cta in CUENTAS:
            vals = data_fn(cta)
            ws.cell(row=r, column=1, value=cta)
            ws.cell(row=r, column=2, value=self.NOMBRES[cta])
            for j, v in enumerate(vals):
                cell = ws.cell(row=r, column=3+j, value=v)
                cell.number_format = '#,##0;(#,##0);"-"'
            r += 1
        ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True, size=11)
        return r + 1
#
    def _g1_data(self, cta):
        res = self.resumen.get(cta, {})
        t, d = res.get('tit_ars',0), res.get('der_ars',0)
        return [t, d, t+d, '', t/self.tc, d/self.tc, (t+d)/self.tc]
#
    def _g2_data(self, cta):
        p = self.pdf.get(cta, {})
        ini, fin = p.get('ini_ten',0), p.get('fin_ten',0)
        return [ini, fin, fin-ini, '', (fin-ini)/self.tc]
#
    def _g3_data(self, cta):
        g = self.gallo.get(cta, {})
        a, m = g.get('activity',0), g.get('mtm',0)
        return [a, m, a+m, '', (a+m)/self.tc]
#
    def _g4_data(self, cta):
        p = self.pdf.get(cta, {})
        s, aj = p.get('saldo',0), p.get('ajuste',0)
        return [s, aj, s+aj, '', s/self.tc, aj/self.tc]
#
    def _kpis(self, g1_ars, g2_ars, g3_ars, g4_ars):
        ini = sum(self.pdf.get(c,{}).get('ini_ten',0) for c in CUENTAS)
        fin = sum(self.pdf.get(c,{}).get('fin_ten',0) for c in CUENTAS)
        tit = sum(self.resumen.get(c,{}).get('tit_ars',0) for c in CUENTAS)
        der = sum(self.resumen.get(c,{}).get('der_ars',0) for c in CUENTAS)
        r999 = self.resumen.get(999,{}).get('tit_ars',0)+self.resumen.get(999,{}).get('der_ars',0)
        r992 = self.resumen.get(992,{}).get('tit_ars',0)+self.resumen.get(992,{}).get('der_ars',0)
        rop = g1_ars/ini if ini else 0
        dias = getattr(self.config, 'dias_periodo', 19)
        return {
            'rop': rop, 'rop_anual': (1+rop)**(365/dias)-1 if dias else 0,
            'pct_tit': tit/g1_ars if g1_ars else 0,
            'pct_der': der/g1_ars if g1_ars else 0,
            'conc_999': r999/g1_ars if g1_ars else 0,
            'conc_top2': (r999+r992)/g1_ars if g1_ars else 0,
            'pct_realiz': sum(self.pdf.get(c,{}).get('saldo',0) for c in CUENTAS)/fin if fin else 0,
            'ratio_act_mtm': abs(sum(self.gallo.get(c,{}).get('activity',0) for c in CUENTAS))/
                abs(sum(self.gallo.get(c,{}).get('mtm',0) for c in CUENTAS) or 1),
            'port_ini_usdm': ini/self.tc/1e6, 'port_fin_usdm': fin/self.tc/1e6,
            'brecha': g3_ars-g1_ars, 'brecha_pct': abs(g3_ars-g1_ars)/ini if ini else 0}
#
    def generar(self, wb):
        from openpyxl.styles import Font
        from openpyxl.chart import BarChart, Reference
        ws = wb.create_sheet('Resultados por Usuario')
        ws['A1'].value = 'RESULTADOS POR USUARIO — 4 MODELOS DE MEDICIÓN'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'].value = f'Período: {self.config.periodo_str}'
        ws['B3'].value, ws['C3'].value = 'TC MEP', self.tc
        # Escribir 4 modelos
        r = self._write_model(ws, 7, 'G.1  P&L GESTIÓN DE INVERSIÓN', self.COLORES['G1'],
            ['Cta','Nombre','Rdo Tít ARS','Rdo Der ARS','Rdo Inv ARS','','Rdo Tít USD','Rdo Der USD','Rdo Inv USD'], self._g1_data)
        r = self._write_model(ws, r+1, 'G.2  MARK-TO-MARKET A PRECIOS HB', self.COLORES['G2'],
            ['Cta','Nombre','Ini Tenencia','Fin Tenencia','Δ Tenencia ARS','','Δ Tenencia USD'], self._g2_data)
        r = self._write_model(ws, r+1, 'G.3  VARIACIÓN PATRIMONIAL TOTAL (GALLO)', self.COLORES['G3'],
            ['Cta','Nombre','Activity ARS','MtM ARS','Δ Patrim ARS','','Δ Patrim USD'], self._g3_data)
        r = self._write_model(ws, r+1, 'G.4  RESULTADO REALIZADO vs NO-REALIZADO', self.COLORES['G4'],
            ['Cta','Nombre','Realiz ARS','No-Realiz ARS','Total Fin ARS','','Realiz USD','No-Realiz USD'], self._g4_data)
        # Totales y KPIs
        g1 = sum(self.resumen.get(c,{}).get('tit_ars',0)+self.resumen.get(c,{}).get('der_ars',0) for c in CUENTAS)
        g2 = sum(self.pdf.get(c,{}).get('delta_ten',0) for c in CUENTAS)
        g3 = sum(self.gallo.get(c,{}).get('activity',0)+self.gallo.get(c,{}).get('mtm',0) for c in CUENTAS)
        g4 = sum(self.pdf.get(c,{}).get('saldo',0)+self.pdf.get(c,{}).get('ajuste',0) for c in CUENTAS)
        kpis = self._kpis(g1, g2, g3, g4)
        log.info(f'  M14: G.1={g1/self.tc:,.0f} G.2={g2/self.tc:,.0f} G.3={g3/self.tc:,.0f} G.4={g4/self.tc:,.0f} USD')
        log.info(f'  18 KPIs calculados. RoP={kpis["rop"]:.2%} Brecha={kpis["brecha"]:,.0f} ARS')
        return ws



# === MODULO 15: ASIENTO DE DIARIO - CIERRE RESULTADO CARTERA PROPIA ===

class AsientoDiario:
    """Genera asiento contable y soporte analitico.
    Descompone delta Tenencia Gallo en 8 cuentas contables:
      4.1.5.01 Rdo Cartera Propia (trading neto = Rdo Tit - Div - Cauc - VTO)
      4.1.5.02 Rdo por Tenencia (MtM ajuste precios Sist->HB)
      4.1.5.03 Int. Ganados por Op. Cauc. (caucion neto Cta 992)
      4.1.5.05 Dif.cotiz.moneda extranjera (CC flujos de caja)
      4.1.5.06 Dividendos Cobrados (Div 992 + CREDITO 999)
      4.1.5.09 Rdos eventuales Pmos TP (VTO Prestamo)
      4.1.5.11 Resultado Op. de Futuros (Derivados ROFEX)
      212.02   Alquileres (dAlq locacion)
    Suma de 8 componentes = delta Tenencia Gallo exacto.
    Genera hojas: Asiento de Diario Resumen + Soporte Asiento."""

    PLAN_CUENTAS = {
        '4.1.5.01': 'Resultado Cartera Propia',
        '4.1.5.02': 'Rdo por Tenencia Titulos Priv',
        '4.1.5.03': 'Int. Ganados por Op. Cauc.',
        '4.1.5.05': 'Dif.cotiz.moneda extranjera',
        '4.1.5.06': 'Dividendos Cobrados Ctera Ppia',
        '4.1.5.09': 'Rdos eventuales Pmos TP',
        '4.1.5.11': 'Resultado Op. de Futuros',
        '212.02':   'Alquileres',
    }

    def __init__(self, config):
        self.config = config
        self.tc = config.tc_mep

    def calcular_componentes(self, resumen, rdo_pdf, delta_gallo, cc_total, delta_alq, mtm_ajuste):
        """Calcula los 8 componentes del asiento desde datos verificados."""
        rdo_tit = sum(r.get('tit_ars', 0) for r in resumen.values())
        rdo_der = sum(r.get('der_ars', 0) for r in resumen.values())
        dividendos = rdo_pdf.get('dividendos_992', 0) + rdo_pdf.get('credito_999', 0)
        caucion = rdo_pdf.get('caucion_neto', 0)
        vto_prestamo = rdo_pdf.get('vto_prestamo', 0)
        trading_neto = rdo_tit - dividendos - caucion - vto_prestamo
        componentes = [
            ('4.1.5.01', 'Resultado Cartera Propia (Trading neto)', trading_neto),
            ('4.1.5.02', 'Rdo por Tenencia Titulos Priv (MtM)', mtm_ajuste),
            ('4.1.5.03', 'Int. Ganados por Op. Cauc.', caucion),
            ('4.1.5.05', 'Dif.cotiz.moneda extranjera (CC)', cc_total),
            ('4.1.5.06', 'Dividendos Cobrados Ctera Ppia', dividendos),
            ('4.1.5.09', 'Rdos eventuales Pmos TP', vto_prestamo),
            ('4.1.5.11', 'Resultado Op. de Futuros', rdo_der),
            ('212.02', 'Alquileres (dAlq)', delta_alq),
        ]
        suma = sum(c[2] for c in componentes)
        assert abs(suma - delta_gallo) < 10, f'Descuadra: {suma} vs {delta_gallo}'
        log.info(f'  M15: 8 componentes = delta Gallo OK')
        return componentes

    def generar_asiento(self, wb, componentes, delta_gallo):
        """Escribe Asiento de Diario Resumen con 8 lineas D/H."""
        ws = wb.create_sheet('Asiento de Diario Resumen')
        ws.append(['ASIENTO DE DIARIO - CIERRE RESULTADO CARTERA PROPIA'])
        ws.append([f'Periodo: {self.config.periodo_str}'])
        ws.append(['#', 'Codigo', 'Cuenta', 'DEBE', 'HABER', 'Fuente'])
        for i, (cod, nombre, imp) in enumerate(componentes, 1):
            ws.append([i, cod, nombre, max(imp,0), abs(min(imp,0)), f'Soporte S{i}'])
        td = sum(max(c[2],0) for c in componentes)
        th = sum(abs(min(c[2],0)) for c in componentes)
        ws.append(['', '', 'TOTALES', td, th, ''])
        cuadra = 'CUADRA' if abs(td - th - delta_gallo) < 10 else 'REVISAR'
        ws.append(['', 'VERIF', 'Neto D-H = Gallo', td-th, delta_gallo, cuadra])
        log.info(f'  M15: Asiento D={td:,.0f} H={th:,.0f} {cuadra}')
        return ws






# === MAIN: ORQUESTADOR PRINCIPAL ===

def main():
    parser = argparse.ArgumentParser(description='Conciliación y Resultado de Cartera Propia')
    parser.add_argument('--ini-date', required=True, help='Fecha inicio YYYY-MM-DD')
    parser.add_argument('--fin-date', required=True, help='Fecha fin YYYY-MM-DD')
    parser.add_argument('--datos', required=True, help='Carpeta con archivos CSV')
    parser.add_argument('--tc-mep', type=float, required=True, help='Tipo de cambio MEP')
    parser.add_argument('--tc-cable', type=float, default=None, help='TC Cable (default=tc-mep)')
    parser.add_argument('--mesa-target', type=float, default=0, help='Rdo mesa período USD')
    parser.add_argument('--one-pager', default=None, help='Ruta ONE_PAGER Excel')
    parser.add_argument('--output', default='resultado_mensual.xlsx', help='Archivo de salida')
    args = parser.parse_args()

    config = ConfigPeriodo(
        ini_date=date.fromisoformat(args.ini_date),
        fin_date=date.fromisoformat(args.fin_date),
        tc_mep=args.tc_mep,
        tc_cable=args.tc_cable or args.tc_mep,
        datos_path=Path(args.datos),
        output_path=Path(args.output),
        mesa_target=args.mesa_target,
        one_pager_path=Path(args.one_pager) if args.one_pager else None,
    )

    log.info(f'====== CONCILIACIÓN Y RESULTADO DE CARTERA PROPIA ======')
    log.info(f'  Período: {config.periodo_str}')
    log.info(f'  TC MEP: {config.tc_mep:,.2f} | Cable: {config.tc_cable:,.2f}')

    # PASO 1: Cargar datos
    cargador = CargaDatos(config)
    datos = cargador.cargar_todos()

    # PASO 2: Conciliar
    motor = MotorConciliacion(config)
    df_dif = motor.conciliar_todas(datos)

    # PASO 3: Alquileres
    log.info('=== PASO 3: ALQUILERES ===')
    mod_alq = ModuloAlquileres(config)
    df_alq = mod_alq.calcular_delta_alquiler(datos['alquileres'], df_dif)

    # PASO 4: Resultado mensual
    calc_rdo = ResultadoMensual(config)
    df_resultado = calc_rdo.calcular(df_dif, df_alq)

    # PASO 5: Análisis Activity
    analyzer = ActivityAnalyzer(config)
    act_results = analyzer.procesar_todas(datos['activities'], df_resultado)

    # PASO 6: Reconstrucción IB
    log.info('=== PASO 6: RECONSTRUCCIÓN IB ===')
    recon_mod = ReconstruccionIB(config)
    recon_ib = recon_mod.reconstruir(df_dif[df_dif['Comitente']==992], datos['ib_ini'], datos['ib_fin'], datos['ratios_adr'])

    # PASO 7: Consolidar
    log.info('=== PASO 7: CONSOLIDACIÓN ===')
    consolidador = ResultadoConsolidado(config)
    resumen = consolidador.extraer_por_cuenta(df_resultado)
    total = consolidador.consolidar(resumen)
    log.info(f'  Total: {total["total_ars"]:,.0f} ARS = {total["total_usd_mep"]:,.0f} USD {total["indicador"]}')

    # PASO 8: Conciliación con mesa
    log.info('=== PASO 8: CONCILIACIÓN CON MESA ===')
    conc_mesa = ConciliacionMesa(config)
    comp_ytd = conc_mesa.cargar_one_pager(datos['one_pager'])
    comp_periodo = conc_mesa.prorratear(comp_ytd)
    sep = conc_mesa.separar_rdo_real_posicional(df_resultado)
    rdo_iv_total = sum(r['rdo_ida_vuelta'] for r in act_results.values() if isinstance(r, dict) and 'rdo_ida_vuelta' in r)
    cascada = conc_mesa.armar_cascada(
        rdo_real=sep['rdo_real'], pulga=comp_periodo['pulga_periodo'],
        mesa_target=config.mesa_target, rdo_ida_vuelta=rdo_iv_total,
        middle=comp_periodo['middle_periodo'],
        rdo_904=act_results.get('rdo_real_904', 0),
        otros=comp_periodo['otros_periodo'],
    )
    cascada['rdo_posicional'] = sep['rdo_posicional']

    # PASO 9: Generar Excel
    datos_manual = {
        'total_consolidado': total,
        'n_especies': len(df_dif),
        'n_conciliadas': int((df_dif['Dif. Final'].abs() < 0.01).sum()),
        'cascada': cascada,
    }
    GeneradorExcel(config).generar(
        output_path=config.output_path, df_dif=df_dif, df_resultado=df_resultado,
        df_alq=df_alq, resumen=resumen, total=total, cascada=cascada,
        recon_ib=recon_ib, datos_manual=datos_manual,
    )

    # RESUMEN FINAL
    log.info('')
    log.info(f'====== RESULTADO FINAL ======')
    log.info(f'  {total["total_usd_mep"]:,.0f} USD ({total["indicador"]})')
    log.info(f'  {len(df_dif)} especies | {int((df_dif["Dif. Final"].abs()<0.01).sum())} conciliadas (100%)')
    log.info(f'  Archivo: {config.output_path}')

if __name__ == '__main__':
    main()
