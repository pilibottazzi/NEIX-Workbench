from __future__ import annotations

import io
import re
import unicodedata
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st

warnings.filterwarnings("ignore")


# =============================================================================
# UI / ESTILO NEIX
# =============================================================================

NEIX_RED = "#ff3b30"
TEXT = "#111827"
MUTED = "#6b7280"
BORDER = "rgba(17,24,39,0.10)"
CARD_BG = "rgba(255,255,255,0.96)"
BG_SOFT = "#f6f7f9"
OK_BG = "rgba(34,197,94,0.10)"
OK_TXT = "#15803d"
WARN_BG = "rgba(245,158,11,0.12)"
WARN_TXT = "#b45309"
BAD_BG = "rgba(255,59,48,0.10)"
BAD_TXT = "#b91c1c"


def inject_css() -> None:
    st.markdown(
        f"""
        <style>
          .block-container {{
            max-width: 1480px;
            padding-top: 1.1rem;
            padding-bottom: 2rem;
          }}

          html, body, [class*="css"] {{
            color: {TEXT};
          }}

          .main {{
            background: {BG_SOFT};
          }}

          .neix-shell {{
            background: linear-gradient(180deg, #ffffff 0%, #fbfbfc 100%);
            border: 1px solid {BORDER};
            border-radius: 24px;
            padding: 1.25rem 1.25rem 1.05rem 1.25rem;
            box-shadow: 0 12px 32px rgba(17,24,39,0.05);
            margin-bottom: 1rem;
          }}

          .neix-title {{
            font-size: 1.7rem;
            font-weight: 750;
            letter-spacing: -0.03em;
            margin-bottom: 0.15rem;
          }}

          .neix-sub {{
            color: {MUTED};
            font-size: 0.96rem;
            margin: 0;
          }}

          .neix-section {{
            font-size: 1.02rem;
            font-weight: 700;
            margin: 0.15rem 0 0.75rem 0;
          }}

          .neix-card {{
            background: {CARD_BG};
            border: 1px solid {BORDER};
            border-radius: 18px;
            padding: 0.95rem 1rem 0.9rem 1rem;
            box-shadow: 0 8px 24px rgba(17,24,39,0.04);
            height: 100%;
          }}

          .neix-kpi-label {{
            font-size: 0.78rem;
            color: {MUTED};
            text-transform: uppercase;
            letter-spacing: 0.06em;
            margin-bottom: 0.35rem;
          }}

          .neix-kpi-value {{
            font-size: 1.6rem;
            font-weight: 760;
            color: {TEXT};
            line-height: 1.1;
          }}

          .neix-kpi-sub {{
            margin-top: 0.35rem;
            font-size: 0.82rem;
            color: {MUTED};
          }}

          .mini-card {{
            background: white;
            border: 1px solid {BORDER};
            border-radius: 16px;
            padding: 0.85rem 0.9rem;
            box-shadow: 0 6px 18px rgba(17,24,39,0.03);
            margin-bottom: 0.65rem;
          }}

          .mini-title {{
            font-size: 0.82rem;
            color: {MUTED};
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-bottom: 0.25rem;
          }}

          .mini-value {{
            font-size: 1.1rem;
            font-weight: 700;
            color: {TEXT};
          }}

          .status-pill {{
            display: inline-block;
            padding: 0.18rem 0.55rem;
            border-radius: 999px;
            font-size: 0.77rem;
            font-weight: 700;
            letter-spacing: 0.01em;
          }}

          .status-ok {{
            background: {OK_BG};
            color: {OK_TXT};
          }}

          .status-warn {{
            background: {WARN_BG};
            color: {WARN_TXT};
          }}

          .status-bad {{
            background: {BAD_BG};
            color: {BAD_TXT};
          }}

          .audit-box {{
            background: #fff;
            border: 1px solid {BORDER};
            border-radius: 18px;
            padding: 0.95rem 1rem;
            margin-bottom: 0.75rem;
          }}

          .audit-title {{
            font-size: 0.96rem;
            font-weight: 700;
            margin-bottom: 0.25rem;
          }}

          .audit-sub {{
            color: {MUTED};
            font-size: 0.88rem;
            margin: 0;
          }}

          div[data-testid="stDownloadButton"] > button,
          div[data-testid="stButton"] > button {{
            width: 100%;
            border-radius: 12px;
            border: 1px solid transparent;
          }}

          div[data-testid="stDownloadButton"] > button {{
            background: {NEIX_RED} !important;
            color: white !important;
          }}

          div[data-testid="stButton"] > button {{
            background: white;
            color: {TEXT};
            border: 1px solid {BORDER};
          }}

          .stTabs [data-baseweb="tab-list"] {{
            gap: 0.35rem;
          }}

          .stTabs [data-baseweb="tab"] {{
            height: 40px;
            border-radius: 10px;
            padding-left: 0.9rem;
            padding-right: 0.9rem;
          }}

          .small-note {{
            color: {MUTED};
            font-size: 0.86rem;
          }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def card_html(label: str, value: str, sub: str = "") -> str:
    return f"""
    <div class="neix-card">
      <div class="neix-kpi-label">{label}</div>
      <div class="neix-kpi-value">{value}</div>
      <div class="neix-kpi-sub">{sub}</div>
    </div>
    """


def semaforo_html(label: str, status: str, sub: str) -> str:
    css_class = {
        "OK": "status-ok",
        "WARN": "status-warn",
        "BAD": "status-bad",
    }.get(status, "status-warn")

    return f"""
    <div class="mini-card">
      <div class="mini-title">{label}</div>
      <div class="mini-value"><span class="status-pill {css_class}">{status}</span></div>
      <div class="neix-kpi-sub">{sub}</div>
    </div>
    """


# =============================================================================
# CONFIGURACIÓN
# =============================================================================

CUENTAS_DEFAULT = [904, 932, 990, 992, 996, 997, 999, 1000]


@dataclass
class PeriodoConciliacion:
    fecha_ini: str
    fecha_fin: str
    cuentas: List[int]

    pares_comp: Dict[int, int] = field(default_factory=lambda: {
        904: 992,
        992: 904,
        997: 999,
        999: 997,
    })

    cuenta_ib: int = 992

    ratios_adr: Dict[str, float] = field(default_factory=lambda: {
        "AAPL": 20, "AMZN": 150, "ADBE": 44, "AMD": 10, "AAL": 3,
        "ABEV": 0.333, "ARKK": 15, "BABA": 8, "BIDU": 4, "COIN": 5,
        "CSCO": 6, "DIS": 5, "GOOG": 12, "INTC": 10, "JNJ": 2,
        "KO": 5, "META": 6, "MSFT": 10, "NFLX": 5, "NKE": 5,
        "NVDA": 24, "PFE": 10, "PYPL": 5, "QCOM": 5, "SBUX": 5,
        "SQ": 5, "TSLA": 15, "V": 2, "WMT": 5, "XOM": 5,
        "84801": 1, "912797SK4": 1,
    })


# =============================================================================
# HELPERS
# =============================================================================

def normalize_text(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [
        re.sub(r"[^a-zA-Z0-9]+", "_", normalize_text(c)).strip("_")
        for c in out.columns
    ]
    return out


def to_numeric_safe(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    s = s.str.replace(r"[Uu][$]", "", regex=True)
    s = s.str.replace("$", "", regex=False)
    s = s.str.replace(r"\((.*?)\)", r"-\1", regex=True)

    def parse_value(x: str) -> float:
        x = str(x).strip()
        x = re.sub(r"[^0-9,.\-]", "", x)
        if x == "" or x == "-":
            return 0.0

        if "," in x and "." in x:
            if x.rfind(",") > x.rfind("."):
                x = x.replace(".", "").replace(",", ".")
            else:
                x = x.replace(",", "")
        else:
            if x.count(",") == 1 and x.count(".") == 0:
                x = x.replace(",", ".")
            elif x.count(".") > 1 and x.count(",") == 0:
                x = x.replace(".", "")
            elif x.count(",") > 1 and x.count(".") == 0:
                x = x.replace(",", "")

        try:
            return float(x)
        except Exception:
            return 0.0

    return s.apply(parse_value)


def to_datetime_safe(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=True)


def format_int(x) -> str:
    try:
        return f"{int(round(float(x))):,}".replace(",", ".")
    except Exception:
        return "-"


def format_num(x, decimals: int = 2) -> str:
    try:
        return f"{float(x):,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "-"


def format_pct(x) -> str:
    try:
        return f"{float(x):.1f}%"
    except Exception:
        return "-"


def read_any_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()

    if name.endswith(".csv"):
        try:
            return pd.read_csv(uploaded_file)
        except Exception:
            uploaded_file.seek(0)
            try:
                return pd.read_csv(uploaded_file, sep=";")
            except Exception:
                uploaded_file.seek(0)
                return pd.read_csv(uploaded_file, sep="|")

    if name.endswith((".xls", ".xlsx", ".xlsm")):
        return pd.read_excel(uploaded_file)

    raise ValueError(f"Formato no soportado: {uploaded_file.name}")


def guess_file_role(filename: str) -> Tuple[Optional[int], Optional[str]]:
    """
    Devuelve:
    - cuenta detectada
    - rol detectado: posiciones / activity / portafolio_ini / portafolio_fin / desconocido
    """
    name = normalize_text(filename)
    cuenta = None

    for c in CUENTAS_DEFAULT:
        if re.search(rf"(^|[^0-9]){c}([^0-9]|$)", name):
            cuenta = c
            break

    role = None

    if "activity" in name or "movimiento" in name or "movimientos" in name:
        role = "activity"
    elif ("portafolio" in name or "portfolio" in name) and ("inicial" in name or "inicio" in name or "ini" in name):
        role = "portafolio_ini"
    elif ("portafolio" in name or "portfolio" in name) and ("final" in name or "cierre" in name or "fin" in name):
        role = "portafolio_fin"
    elif "posiciones" in name or "consolidado" in name or "conciliacion" in name or "base" in name:
        role = "posiciones"
    elif "inicial" in name and "activity" not in name:
        role = "portafolio_ini"
    elif "final" in name and "activity" not in name:
        role = "portafolio_fin"

    return cuenta, role


# =============================================================================
# PREPARACIÓN DE DATASETS
# =============================================================================

def preparar_posiciones(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    alerts: List[str] = []
    df = normalize_columns(df)

    mapping = {
        "especie": ["especie", "ticker", "activo", "simbolo", "security"],
        "ini_cant": ["ini_cant", "cantidad_inicial", "ini", "inicial", "saldo_inicial"],
        "act_cant": ["act_cant", "cantidad_activity", "activity", "movimiento_cantidad", "activity_cant"],
        "fin_cant": ["fin_cant", "cantidad_final", "fin", "final", "saldo_final"],
        "ini_imp": ["ini_imp", "importe_inicial", "monto_inicial"],
        "act_imp": ["act_imp", "importe_activity", "monto_activity"],
        "fin_imp": ["fin_imp", "importe_final", "monto_final"],
        "comp": ["comp", "compensacion", "compensaciones"],
        "ini_ib": ["ini_ib"],
        "fin_ib": ["fin_ib"],
        "ini_local": ["ini_local"],
        "fin_local": ["fin_local"],
    }

    out = pd.DataFrame()
    for target, options in mapping.items():
        found = next((c for c in options if c in df.columns), None)
        if found:
            out[target] = df[found]

    if "especie" not in out.columns:
        raise ValueError("El archivo de posiciones debe contener especie/ticker/activo.")

    numeric_cols = [
        "ini_cant", "act_cant", "fin_cant",
        "ini_imp", "act_imp", "fin_imp",
        "comp", "ini_ib", "fin_ib", "ini_local", "fin_local"
    ]
    for c in numeric_cols:
        if c not in out.columns:
            out[c] = 0
            alerts.append(f"Se asumió columna faltante en posiciones: {c}=0")
        out[c] = to_numeric_safe(out[c])

    out["especie"] = out["especie"].astype(str).str.strip()
    out = out[out["especie"].astype(str).str.strip() != ""].copy()

    if out["especie"].duplicated().any():
        alerts.append("Se detectaron especies duplicadas en posiciones; se consolidaron por suma.")
        agg_map = {c: "sum" for c in numeric_cols}
        out = out.groupby("especie", dropna=False, as_index=False).agg(agg_map)

    return out.reset_index(drop=True), alerts


def preparar_activity(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    alerts: List[str] = []
    df = normalize_columns(df)

    mapping = {
        "ticker": ["ticker", "especie", "activo", "simbolo", "security"],
        "cantidad": ["cantidad", "cant", "quantity", "qty"],
        "tipo": ["tipo", "concepto", "movimiento", "side"],
        "nro_comprobante": ["nro_comprobante", "comprobante", "numero_comprobante", "id"],
        "fecha_emision": ["fecha_emision", "fecha", "trade_date", "fecha_operacion"],
        "descripcion": ["descripcion", "detalle"],
    }

    out = pd.DataFrame()
    for target, options in mapping.items():
        found = next((c for c in options if c in df.columns), None)
        if found:
            out[target] = df[found]

    for req in ["ticker", "cantidad"]:
        if req not in out.columns:
            raise ValueError(f"Falta columna obligatoria en Activity: {req}")

    if "tipo" not in out.columns:
        out["tipo"] = ""
        alerts.append("Activity sin columna tipo; se asumió vacío.")

    if "nro_comprobante" not in out.columns:
        out["nro_comprobante"] = ""
        alerts.append("Activity sin número de comprobante; deduplicación más limitada.")

    if "fecha_emision" not in out.columns:
        out["fecha_emision"] = pd.NaT
        alerts.append("Activity sin fecha; no se podrá analizar pre-concertadas por fecha.")
    else:
        out["fecha_emision"] = to_datetime_safe(out["fecha_emision"])

    if "descripcion" not in out.columns:
        out["descripcion"] = ""

    out["ticker"] = out["ticker"].astype(str).str.strip()
    out["tipo"] = out["tipo"].astype(str).str.strip()
    out["cantidad"] = to_numeric_safe(out["cantidad"])
    out = out[out["ticker"].astype(str).str.strip() != ""].copy()

    return out.reset_index(drop=True), alerts


def preparar_portafolio(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    alerts: List[str] = []
    df = normalize_columns(df)

    mapping = {
        "descripcion": ["descripcion", "especie", "detalle", "ticker", "security"],
        "cantidad": ["cantidad", "cant", "quantity", "qty"],
    }

    out = pd.DataFrame()
    for target, options in mapping.items():
        found = next((c for c in options if c in df.columns), None)
        if found:
            out[target] = df[found]

    if "descripcion" not in out.columns:
        out["descripcion"] = ""
        alerts.append("Portafolio sin descripción; algunas reglas especiales no aplicarán.")

    if "cantidad" not in out.columns:
        out["cantidad"] = 0
        alerts.append("Portafolio sin cantidad; se asumió 0.")

    out["descripcion"] = out["descripcion"].astype(str).str.strip()
    out["cantidad"] = to_numeric_safe(out["cantidad"])
    out = out[out["descripcion"].astype(str).str.strip() != ""].copy()

    return out.reset_index(drop=True), alerts


# =============================================================================
# REGLAS
# =============================================================================

class ReglasReconciliacion:

    @staticmethod
    def regla_preconcertadas(df_activity: pd.DataFrame, fecha_ini: str, tiene_ini: bool) -> pd.DataFrame:
        if df_activity.empty or "fecha_emision" not in df_activity.columns or not tiene_ini:
            return pd.DataFrame()
        return df_activity[df_activity["fecha_emision"] == pd.to_datetime(fecha_ini)].copy()

    @staticmethod
    def regla_duplicados(df_activity: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
        cols_dedup = ["nro_comprobante", "tipo", "ticker", "cantidad"]
        cols_existentes = [c for c in cols_dedup if c in df_activity.columns]
        if not cols_existentes:
            return df_activity.copy(), 0
        before = len(df_activity)
        after_df = df_activity.drop_duplicates(subset=cols_existentes, keep="first").copy()
        removed = before - len(after_df)
        return after_df, removed

    @staticmethod
    def regla_vto_prestamo(df_activity: pd.DataFrame) -> Tuple[pd.DataFrame, float, int]:
        if "tipo" not in df_activity.columns:
            return df_activity.copy(), 0.0, 0

        mask = df_activity["tipo"].astype(str).str.upper() == "VTO PRESTAMO"
        excluidos = df_activity.loc[mask].copy()
        total = excluidos["cantidad"].sum() if "cantidad" in excluidos.columns else 0.0
        return df_activity.loc[~mask].copy(), total, len(excluidos)

    @staticmethod
    def regla_decr_desfase(df_activity: pd.DataFrame, ini: float, fin: float, ticker: str) -> Dict:
        if "tipo" not in df_activity.columns or "cantidad" not in df_activity.columns:
            return {
                "ticker": ticker,
                "decr_activity": 0,
                "decr_posicion": 0,
                "exceso": 0,
                "dif_calculada": ini - fin,
                "match": False,
                "ajuste_fin": 0,
            }

        decr_mask = df_activity["tipo"].astype(str).str.upper() == "DECR"
        decr_cant = df_activity.loc[decr_mask, "cantidad"].sum()
        trading_cant = df_activity.loc[~decr_mask, "cantidad"].sum()

        pos_change = fin - ini
        decr_en_posicion = pos_change - trading_cant
        exceso = abs(decr_cant) - abs(decr_en_posicion)
        dif = ini + df_activity["cantidad"].sum() - fin

        return {
            "ticker": ticker,
            "decr_activity": decr_cant,
            "decr_posicion": decr_en_posicion,
            "exceso": exceso,
            "dif_calculada": dif,
            "match": abs(abs(exceso) - abs(dif)) < 1,
            "ajuste_fin": -exceso,
        }

    @staticmethod
    def regla_al30_local_exterior(df_portafolio: pd.DataFrame, ticker: str = "AL30") -> Dict:
        if df_portafolio is None or df_portafolio.empty:
            return {
                "tiene_local": False,
                "tiene_exterior": False,
                "cant_local": 0.0,
                "cant_exterior": 0.0,
                "cant_total": 0.0,
            }

        if "descripcion" not in df_portafolio.columns or "cantidad" not in df_portafolio.columns:
            return {
                "tiene_local": False,
                "tiene_exterior": False,
                "cant_local": 0.0,
                "cant_exterior": 0.0,
                "cant_total": 0.0,
            }

        al30_rows = df_portafolio[
            df_portafolio["descripcion"].astype(str).str.contains(ticker, case=False, na=False)
        ].copy()

        local = al30_rows[~al30_rows["descripcion"].astype(str).str.contains("EUR", case=False, na=False)]
        exterior = al30_rows[al30_rows["descripcion"].astype(str).str.contains("EUR", case=False, na=False)]

        cant_local = float(local["cantidad"].sum()) if not local.empty else 0.0
        cant_exterior = float(exterior["cantidad"].sum()) if not exterior.empty else 0.0

        return {
            "tiene_local": not local.empty,
            "tiene_exterior": not exterior.empty,
            "cant_local": cant_local,
            "cant_exterior": cant_exterior,
            "cant_total": cant_local + cant_exterior,
        }

    @staticmethod
    def regla_cc_pesos(especie: str) -> bool:
        esp = str(especie).strip().upper()
        return esp in {"CC PESOS", "CC DOLAR LOCAL", "CC U$S EXTERIOR", "CC USD EXTERIOR"}

    @staticmethod
    def regla_comp_solo_si_dif(dif_cant: float, comp: float) -> float:
        return 0 if abs(dif_cant) < 1 else comp


class ConversionADR:
    @staticmethod
    def obtener_ratio(especie: str, ratios_conocidos: Dict[str, float]) -> float:
        return ratios_conocidos.get(str(especie).strip().upper(), 1)

    @staticmethod
    def calcular_total_992(local_value: float, ib_value: float, ratio: float) -> float:
        return local_value + (ib_value * ratio)


# =============================================================================
# MOTOR PRINCIPAL
# =============================================================================

class ConciliadorMensual:
    def __init__(self, periodo: PeriodoConciliacion):
        self.periodo = periodo
        self.reglas = ReglasReconciliacion()

    def conciliar_especie(
        self,
        cuenta: int,
        especie: str,
        ini_cant: float,
        act_cant: float,
        fin_cant: float,
        ini_imp: float = 0,
        act_imp: float = 0,
        fin_imp: float = 0,
        comp: float = 0,
        ratio: float = 1,
        df_activity_especie: Optional[pd.DataFrame] = None,
        df_portafolio_ini: Optional[pd.DataFrame] = None,
        df_portafolio_fin: Optional[pd.DataFrame] = None,
    ) -> Dict:
        ajustes: List[str] = []
        auditoria_flags: List[str] = []
        especie_upper = str(especie).strip().upper()

        ini_original = float(ini_cant)
        act_original = float(act_cant)
        fin_original = float(fin_cant)
        comp_original = float(comp)

        ini_ajust = float(ini_cant)
        act_ajust = float(act_cant)
        fin_ajust = float(fin_cant)
        comp_ajust = float(comp)

        # CC
        if self.reglas.regla_cc_pesos(especie_upper):
            dif_importe = ini_imp + act_imp - fin_imp
            return {
                "cuenta": cuenta,
                "especie": especie,
                "ini": ini_original,
                "act": act_original,
                "fin": fin_original,
                "comp": comp_original,
                "ratio": ratio,
                "ini_imp": ini_imp,
                "act_imp": act_imp,
                "fin_imp": fin_imp,
                "ini_ajust": ini_original,
                "act_ajust": act_original,
                "fin_ajust": fin_original,
                "comp_ajust": 0.0,
                "dif_cant_original": ini_original + act_original - fin_original,
                "dif_final_original": ini_original + act_original - fin_original + comp_original,
                "dif_cant": 0.0,
                "dif_final": 0.0,
                "dif_importe": dif_importe,
                "status": "CERRADA (importe)",
                "ajustes": "CC conciliada por importe",
                "flags": "CC_IMPORTE",
                "reglas_aplicadas": 1,
            }

        # AL30 local + exterior
        if especie_upper == "AL30" and df_portafolio_ini is not None and df_portafolio_fin is not None:
            al30_ini = self.reglas.regla_al30_local_exterior(df_portafolio_ini)
            al30_fin = self.reglas.regla_al30_local_exterior(df_portafolio_fin)

            if al30_ini["tiene_local"] and al30_ini["tiene_exterior"]:
                ini_ajust = al30_ini["cant_total"]
                fin_ajust = al30_fin["cant_total"]
                ajustes.append("AL30 local + exterior sumado")
                auditoria_flags.append("AL30_DOBLE_LINEA")

        # Activity por especie
        if df_activity_especie is not None and not df_activity_especie.empty:
            # preconcertadas
            pre = self.reglas.regla_preconcertadas(
                df_activity_especie,
                self.periodo.fecha_ini,
                tiene_ini=abs(ini_original) > 0
            )
            if not pre.empty:
                pre_total = pre["cantidad"].sum()
                act_ajust -= pre_total
                ajustes.append(f"Pre-concertadas excluidas ({format_num(pre_total)})")
                auditoria_flags.append("PRECONCERTADA")

            # DECR
            if {"tipo", "cantidad"}.issubset(df_activity_especie.columns):
                decr_ops = df_activity_especie[df_activity_especie["tipo"].astype(str).str.upper() == "DECR"]
                if not decr_ops.empty:
                    analisis_decr = self.reglas.regla_decr_desfase(df_activity_especie, ini_ajust, fin_ajust, especie)
                    if analisis_decr["match"] and abs(analisis_decr["exceso"]) > 0:
                        fin_ajust += analisis_decr["ajuste_fin"]
                        ajustes.append(f"DECR ajustado en Fin ({format_num(analisis_decr['ajuste_fin'])})")
                        auditoria_flags.append("DECR_DESFASE")

        dif_cant_original = ini_original + act_original - fin_original
        dif_sin_comp = ini_ajust + act_ajust - fin_ajust
        comp_ajust = self.reglas.regla_comp_solo_si_dif(dif_sin_comp, comp_original)

        if comp_original != 0 and comp_ajust == 0:
            ajustes.append("Compensación anulada por DifCant=0")
            auditoria_flags.append("COMP_IGNORADA")

        dif_final_original = dif_cant_original + comp_original
        dif_cant = ini_ajust + act_ajust - fin_ajust
        dif_final = dif_cant + comp_ajust

        status = "CERRADA" if abs(dif_final) < 1 else "PENDIENTE"

        if status == "PENDIENTE":
            auditoria_flags.append("PENDIENTE")

        if abs(dif_final) >= 1 and abs(dif_final) < 100:
            auditoria_flags.append("DESVIO_BAJO")
        elif abs(dif_final) >= 100 and abs(dif_final) < 10000:
            auditoria_flags.append("DESVIO_MEDIO")
        elif abs(dif_final) >= 10000:
            auditoria_flags.append("DESVIO_ALTO")

        return {
            "cuenta": cuenta,
            "especie": especie,
            "ini": ini_original,
            "act": act_original,
            "fin": fin_original,
            "comp": comp_original,
            "ratio": ratio,
            "ini_imp": ini_imp,
            "act_imp": act_imp,
            "fin_imp": fin_imp,
            "ini_ajust": ini_ajust,
            "act_ajust": act_ajust,
            "fin_ajust": fin_ajust,
            "comp_ajust": comp_ajust,
            "dif_cant_original": dif_cant_original,
            "dif_final_original": dif_final_original,
            "dif_cant": dif_cant,
            "dif_final": dif_final,
            "dif_importe": ini_imp + act_imp - fin_imp,
            "status": status,
            "ajustes": " | ".join(ajustes) if ajustes else "",
            "flags": " | ".join(auditoria_flags) if auditoria_flags else "",
            "reglas_aplicadas": len(ajustes),
        }

    def conciliar_cuenta(
        self,
        cuenta: int,
        df_posiciones: pd.DataFrame,
        df_activity: pd.DataFrame,
        df_portafolio_ini: Optional[pd.DataFrame] = None,
        df_portafolio_fin: Optional[pd.DataFrame] = None,
    ) -> Tuple[pd.DataFrame, Dict, pd.DataFrame]:
        conversor = ConversionADR()
        auditoria_eventos: List[Dict] = []

        activity_original_rows = len(df_activity)
        df_activity_dedup, dup_removed = self.reglas.regla_duplicados(df_activity)
        df_activity_clean, vto_total, vto_count = self.reglas.regla_vto_prestamo(df_activity_dedup)

        if dup_removed > 0:
            auditoria_eventos.append({
                "cuenta": cuenta,
                "tipo_evento": "ACTIVITY_DUPLICADOS",
                "detalle": f"Se eliminaron {dup_removed} filas duplicadas de Activity",
                "impacto": "WARN",
            })

        if vto_count > 0:
            auditoria_eventos.append({
                "cuenta": cuenta,
                "tipo_evento": "VTO_PRESTAMO",
                "detalle": f"Se excluyeron {vto_count} operaciones VTO PRESTAMO por {format_num(vto_total)}",
                "impacto": "WARN",
            })

        resultados: List[Dict] = []

        for _, row in df_posiciones.iterrows():
            especie = row["especie"]

            act_especie = (
                df_activity_clean[df_activity_clean["ticker"].astype(str).str.upper() == str(especie).upper()].copy()
                if "ticker" in df_activity_clean.columns
                else pd.DataFrame()
            )

            ini_cant = float(row.get("ini_cant", 0) or 0)
            fin_cant = float(row.get("fin_cant", 0) or 0)

            ratio = 1.0
            if cuenta == self.periodo.cuenta_ib:
                ini_ib = float(row.get("ini_ib", 0) or 0)
                fin_ib = float(row.get("fin_ib", 0) or 0)
                ini_local = float(row.get("ini_local", ini_cant) or 0)
                fin_local = float(row.get("fin_local", fin_cant) or 0)
                ratio = float(conversor.obtener_ratio(especie, self.periodo.ratios_adr))
                ini_cant = float(conversor.calcular_total_992(ini_local, ini_ib, ratio))
                fin_cant = float(conversor.calcular_total_992(fin_local, fin_ib, ratio))

            result = self.conciliar_especie(
                cuenta=cuenta,
                especie=especie,
                ini_cant=ini_cant,
                act_cant=float(row.get("act_cant", 0) or 0),
                fin_cant=fin_cant,
                ini_imp=float(row.get("ini_imp", 0) or 0),
                act_imp=float(row.get("act_imp", 0) or 0),
                fin_imp=float(row.get("fin_imp", 0) or 0),
                comp=float(row.get("comp", 0) or 0),
                ratio=ratio,
                df_activity_especie=act_especie if not act_especie.empty else None,
                df_portafolio_ini=df_portafolio_ini,
                df_portafolio_fin=df_portafolio_fin,
            )
            resultados.append(result)

        df_result = pd.DataFrame(resultados)

        # Activity sin match
        especies_pos = set(df_posiciones["especie"].astype(str).str.upper().tolist())
        activity_sin_match = df_activity_clean[
            ~df_activity_clean["ticker"].astype(str).str.upper().isin(especies_pos)
        ].copy()

        if not activity_sin_match.empty:
            grouped = (
                activity_sin_match.groupby("ticker", dropna=False)["cantidad"]
                .sum()
                .reset_index()
                .rename(columns={"ticker": "especie", "cantidad": "cantidad_activity"})
            )
            for _, r in grouped.iterrows():
                auditoria_eventos.append({
                    "cuenta": cuenta,
                    "tipo_evento": "ACTIVITY_SIN_MATCH",
                    "detalle": f"Activity sin especie en posiciones: {r['especie']} ({format_num(r['cantidad_activity'])})",
                    "impacto": "BAD",
                })

        # Especies de posiciones sin movimientos
        sin_mov = df_result[
            (df_result["act"].abs() < 1) &
            (df_result["status"] == "PENDIENTE")
        ].copy()
        for _, r in sin_mov.iterrows():
            auditoria_eventos.append({
                "cuenta": cuenta,
                "tipo_evento": "PENDIENTE_SIN_ACTIVITY",
                "detalle": f"{r['especie']} pendiente sin activity explícita. Dif={format_num(r['dif_final'])}",
                "impacto": "BAD",
            })

        resumen = {
            "cuenta": cuenta,
            "activity_rows_original": activity_original_rows,
            "activity_rows_clean": len(df_activity_clean),
            "duplicados_eliminados": dup_removed,
            "vto_prestamo_excluidos": vto_count,
            "total_especies": len(df_result),
            "cerradas": int(df_result["status"].astype(str).str.contains("CERRADA").sum()),
            "pendientes": int((df_result["status"] == "PENDIENTE").sum()),
            "pct_cierre": (
                float(df_result["status"].astype(str).str.contains("CERRADA").mean() * 100)
                if len(df_result) > 0 else 0
            ),
            "dif_total_abs": float(df_result["dif_final"].abs().sum()),
            "dif_total_neto": float(df_result["dif_final"].sum()),
        }

        df_auditoria = pd.DataFrame(auditoria_eventos)
        return df_result, resumen, df_auditoria

    def generar_reporte(self, resultados: List[pd.DataFrame]) -> pd.DataFrame:
        if not resultados:
            return pd.DataFrame()
        return pd.concat(resultados, ignore_index=True)


# =============================================================================
# EXPORT EXCEL
# =============================================================================

def auto_adjust_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                value_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, value_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def style_header(ws) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    fill = PatternFill("solid", fgColor="FF3B30")
    font = Font(color="FFFFFF", bold=True)
    border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def build_excel_bytes(
    df_all: pd.DataFrame,
    df_resumen_cuentas: pd.DataFrame,
    df_pendientes: pd.DataFrame,
    df_top_pendientes: pd.DataFrame,
    df_auditoria: pd.DataFrame,
    df_reglas: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Orden recomendado
        df_resumen_cuentas.to_excel(writer, sheet_name="Resumen", index=False)
        df_top_pendientes.to_excel(writer, sheet_name="Top Pendientes", index=False)
        df_pendientes.to_excel(writer, sheet_name="Pendientes", index=False)
        df_all.to_excel(writer, sheet_name="Detalle Consolidado", index=False)
        df_auditoria.to_excel(writer, sheet_name="Auditoria", index=False)
        df_reglas.to_excel(writer, sheet_name="Reglas Aplicadas", index=False)

        wb = writer.book
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            style_header(ws)
            auto_adjust_columns(ws)
            ws.freeze_panes = "A2"

    output.seek(0)
    return output.getvalue()


# =============================================================================
# TABLAS DERIVADAS
# =============================================================================

def build_resumen_cuentas(df_all: pd.DataFrame) -> pd.DataFrame:
    if df_all.empty:
        return pd.DataFrame()

    resumen = (
        df_all.groupby("cuenta", dropna=False)
        .agg(
            especies=("especie", "count"),
            cerradas=("status", lambda s: (s.astype(str).str.contains("CERRADA")).sum()),
            pendientes=("status", lambda s: (s == "PENDIENTE").sum()),
            reglas_aplicadas=("reglas_aplicadas", "sum"),
            dif_total_abs=("dif_final", lambda s: s.abs().sum()),
            dif_total_neto=("dif_final", "sum"),
        )
        .reset_index()
    )

    resumen["pct_cierre"] = np.where(
        resumen["especies"] > 0,
        resumen["cerradas"] / resumen["especies"] * 100,
        0,
    )

    def classify(row) -> str:
        if row["pendientes"] == 0:
            return "OK"
        if row["pct_cierre"] >= 85:
            return "WARN"
        return "BAD"

    resumen["semaforo"] = resumen.apply(classify, axis=1)
    resumen = resumen.sort_values(["semaforo", "cuenta"], ascending=[True, True]).reset_index(drop=True)
    return resumen


def build_top_pendientes(df_all: pd.DataFrame, top_n: int = 15) -> pd.DataFrame:
    if df_all.empty:
        return pd.DataFrame()

    top = df_all[df_all["status"] == "PENDIENTE"].copy()
    if top.empty:
        return top

    top["abs_dif"] = top["dif_final"].abs()
    top = top.sort_values(["abs_dif", "cuenta", "especie"], ascending=[False, True, True]).head(top_n)
    cols = [
        "cuenta", "especie", "ini", "act", "fin", "comp",
        "ini_ajust", "act_ajust", "fin_ajust", "comp_ajust",
        "dif_final", "ajustes", "flags"
    ]
    return top[cols].reset_index(drop=True)


def build_reglas_aplicadas(df_all: pd.DataFrame) -> pd.DataFrame:
    if df_all.empty:
        return pd.DataFrame(columns=["regla", "cantidad"])

    counter: Dict[str, int] = {}

    for val in df_all["flags"].fillna("").astype(str):
        parts = [p.strip() for p in val.split("|") if p.strip()]
        for p in parts:
            counter[p] = counter.get(p, 0) + 1

    df = pd.DataFrame(
        [{"regla": k, "cantidad": v} for k, v in counter.items()]
    )
    if df.empty:
        return pd.DataFrame(columns=["regla", "cantidad"])
    return df.sort_values("cantidad", ascending=False).reset_index(drop=True)


# =============================================================================
# RENDER
# =============================================================================

def render_header() -> None:
    st.markdown(
        """
        <div class="neix-shell">
          <div class="neix-title">Conciliación mensual de cartera propia — Premium</div>
          <p class="neix-sub">
            Módulo ejecutivo para conciliación operativa de cartera: control end-to-end, auditoría, top pendientes y export consolidado.
          </p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_main_kpis(df_all: pd.DataFrame) -> None:
    total = len(df_all)
    cerradas = int(df_all["status"].astype(str).str.contains("CERRADA").sum()) if total else 0
    pendientes = int((df_all["status"] == "PENDIENTE").sum()) if total else 0
    pct = (cerradas / total * 100) if total else 0
    dif_abs = float(df_all["dif_final"].abs().sum()) if total else 0
    reglas = int(df_all["reglas_aplicadas"].sum()) if total else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        st.markdown(card_html("Especies analizadas", format_int(total)), unsafe_allow_html=True)
    with c2:
        st.markdown(card_html("Cerradas", format_int(cerradas), format_pct(pct)), unsafe_allow_html=True)
    with c3:
        st.markdown(card_html("Pendientes", format_int(pendientes)), unsafe_allow_html=True)
    with c4:
        st.markdown(card_html("Dif. absoluta acumulada", format_num(dif_abs, 0)), unsafe_allow_html=True)
    with c5:
        st.markdown(card_html("Reglas aplicadas", format_int(reglas)), unsafe_allow_html=True)


def render_semaforo(df_resumen: pd.DataFrame) -> None:
    st.markdown('<div class="neix-section">Semáforo por cuenta</div>', unsafe_allow_html=True)
    cols = st.columns(min(4, max(len(df_resumen), 1)))

    for i, (_, row) in enumerate(df_resumen.iterrows()):
        with cols[i % len(cols)]:
            sub = (
                f"{int(row['cerradas'])}/{int(row['especies'])} cerradas · "
                f"{int(row['pendientes'])} pendientes · "
                f"Dif abs {format_num(row['dif_total_abs'], 0)}"
            )
            st.markdown(
                semaforo_html(f"Cta {int(row['cuenta'])}", row["semaforo"], sub),
                unsafe_allow_html=True,
            )


def render_management_summary(df_resumen: pd.DataFrame, df_top: pd.DataFrame) -> None:
    st.markdown('<div class="neix-section">Resumen ejecutivo</div>', unsafe_allow_html=True)

    total_cuentas = len(df_resumen)
    ok = int((df_resumen["semaforo"] == "OK").sum()) if total_cuentas else 0
    warn = int((df_resumen["semaforo"] == "WARN").sum()) if total_cuentas else 0
    bad = int((df_resumen["semaforo"] == "BAD").sum()) if total_cuentas else 0

    texto = (
        f"Se procesaron {total_cuentas} cuentas. "
        f"{ok} quedaron en estado OK, {warn} en observación y {bad} con desvíos relevantes. "
    )

    if not df_top.empty:
        top_row = df_top.iloc[0]
        texto += (
            f"El principal pendiente corresponde a la cuenta {int(top_row['cuenta'])}, "
            f"especie {top_row['especie']}, con una diferencia de {format_num(top_row['dif_final'], 0)}."
        )
    else:
        texto += "No se detectaron pendientes relevantes."

    st.markdown(
        f"""
        <div class="audit-box">
          <div class="audit-title">Lectura management</div>
          <p class="audit-sub">{texto}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_top_pendientes(df_top: pd.DataFrame) -> None:
    st.markdown('<div class="neix-section">Top pendientes</div>', unsafe_allow_html=True)

    if df_top.empty:
        st.success("No hay pendientes para mostrar.")
        return

    st.dataframe(
        df_top.style.format({
            "ini": "{:,.2f}",
            "act": "{:,.2f}",
            "fin": "{:,.2f}",
            "comp": "{:,.2f}",
            "ini_ajust": "{:,.2f}",
            "act_ajust": "{:,.2f}",
            "fin_ajust": "{:,.2f}",
            "comp_ajust": "{:,.2f}",
            "dif_final": "{:,.2f}",
        }),
        use_container_width=True,
        hide_index=True,
    )


def render_auditoria(df_auditoria: pd.DataFrame) -> None:
    st.markdown('<div class="neix-section">Auditoría</div>', unsafe_allow_html=True)

    if df_auditoria.empty:
        st.info("No se registraron eventos de auditoría.")
        return

    c1, c2 = st.columns([1, 2])

    with c1:
        resumen = (
            df_auditoria.groupby(["impacto", "tipo_evento"], dropna=False)
            .size()
            .reset_index(name="cantidad")
            .sort_values(["impacto", "cantidad"], ascending=[True, False])
        )
        st.dataframe(resumen, use_container_width=True, hide_index=True)

    with c2:
        st.dataframe(df_auditoria, use_container_width=True, hide_index=True)


def render_carga_status(cuenta: int, payload: Dict) -> None:
    ok_pos = payload.get("posiciones") is not None
    ok_act = payload.get("activity") is not None
    ok_ini = payload.get("portafolio_ini") is not None
    ok_fin = payload.get("portafolio_fin") is not None

    def pill(ok: bool, txt_ok: str, txt_bad: str) -> str:
        cls = "status-ok" if ok else "status-bad"
        txt = txt_ok if ok else txt_bad
        return f"<span class='status-pill {cls}'>{txt}</span>"

    st.markdown(
        f"""
        <div style="margin-top:0.55rem;">
          {pill(ok_pos, "Posiciones OK", "Falta posiciones")}
          &nbsp;
          {pill(ok_act, "Activity OK", "Falta activity")}
          &nbsp;
          {pill(ok_ini, "Portafolio ini OK", "Ini opcional")}
          &nbsp;
          {pill(ok_fin, "Portafolio fin OK", "Fin opcional")}
        </div>
        """,
        unsafe_allow_html=True,
    )


# =============================================================================
# APP
# =============================================================================

def main() -> None:
    st.set_page_config(
        page_title="NEIX • Conciliación Premium",
        page_icon="📘",
        layout="wide",
    )
    inject_css()
    render_header()

    with st.sidebar:
        st.markdown("### Parámetros")
        fecha_ini = st.date_input("Fecha inicio", value=date.today().replace(day=1))
        fecha_fin = st.date_input("Fecha fin", value=date.today())
        cuentas_sel = st.multiselect(
            "Cuentas a procesar",
            options=CUENTAS_DEFAULT,
            default=CUENTAS_DEFAULT,
        )

        st.markdown("---")
        auto_naming = st.toggle("Lectura automática por naming", value=True)
        st.caption(
            "Si está activa, intentará asignar archivos por cuenta y tipo usando el nombre del archivo."
        )

        st.markdown("---")
        st.markdown("### Archivos masivos")
        bulk_files = st.file_uploader(
            "Subir múltiples archivos",
            type=["xlsx", "xls", "xlsm", "csv"],
            accept_multiple_files=True,
        )

        st.markdown("---")
        st.markdown("### Nota")
        st.caption(
            "Obligatorios por cuenta: Posiciones + Activity. "
            "Portafolio inicial/final son opcionales pero mejoran reglas especiales."
        )

    periodo = PeriodoConciliacion(
        fecha_ini=str(fecha_ini),
        fecha_fin=str(fecha_fin),
        cuentas=cuentas_sel,
    )
    conciliador = ConciliadorMensual(periodo)

    st.markdown('<div class="neix-section">Carga por cuenta</div>', unsafe_allow_html=True)

    cuentas_data: Dict[int, Dict[str, Optional[pd.DataFrame]]] = {
        c: {
            "posiciones": None,
            "activity": None,
            "portafolio_ini": None,
            "portafolio_fin": None,
            "errores": [],
            "alertas": [],
            "fuentes": [],
        }
        for c in cuentas_sel
    }

    # ---------------------------------
    # AUTO ASIGNACIÓN MASIVA POR NAMING
    # ---------------------------------
    if bulk_files and auto_naming:
        for f in bulk_files:
            cuenta, role = guess_file_role(f.name)
            if cuenta not in cuentas_data or role is None:
                continue

            try:
                df_raw = read_any_file(f)
                if role == "posiciones":
                    df_ready, alerts = preparar_posiciones(df_raw)
                elif role == "activity":
                    df_ready, alerts = preparar_activity(df_raw)
                elif role == "portafolio_ini":
                    df_ready, alerts = preparar_portafolio(df_raw)
                elif role == "portafolio_fin":
                    df_ready, alerts = preparar_portafolio(df_raw)
                else:
                    continue

                cuentas_data[cuenta][role] = df_ready
                cuentas_data[cuenta]["alertas"].extend(alerts)
                cuentas_data[cuenta]["fuentes"].append(f"{role}: {f.name}")

            except Exception as e:
                cuentas_data[cuenta]["errores"].append(f"{role}: {f.name} → {e}")

    tabs = st.tabs([f"Cta {c}" for c in cuentas_sel]) if cuentas_sel else []

    for cuenta, tab in zip(cuentas_sel, tabs):
        with tab:
            c1, c2 = st.columns(2)

            with c1:
                pos_file = st.file_uploader(
                    f"Posiciones consolidadas - Cta {cuenta}",
                    type=["xlsx", "xls", "xlsm", "csv"],
                    key=f"pos_{cuenta}",
                )
                act_file = st.file_uploader(
                    f"Activity - Cta {cuenta}",
                    type=["xlsx", "xls", "xlsm", "csv"],
                    key=f"act_{cuenta}",
                )

            with c2:
                ini_file = st.file_uploader(
                    f"Portafolio inicial - Cta {cuenta} (opcional)",
                    type=["xlsx", "xls", "xlsm", "csv"],
                    key=f"ini_{cuenta}",
                )
                fin_file = st.file_uploader(
                    f"Portafolio final - Cta {cuenta} (opcional)",
                    type=["xlsx", "xls", "xlsm", "csv"],
                    key=f"fin_{cuenta}",
                )

            # Carga manual pisa auto naming si el usuario la usa
            manual_map = [
                ("posiciones", pos_file, preparar_posiciones),
                ("activity", act_file, preparar_activity),
                ("portafolio_ini", ini_file, preparar_portafolio),
                ("portafolio_fin", fin_file, preparar_portafolio),
            ]

            for role, file_obj, prep_fn in manual_map:
                if file_obj is not None:
                    try:
                        df_ready, alerts = prep_fn(read_any_file(file_obj))
                        cuentas_data[cuenta][role] = df_ready
                        cuentas_data[cuenta]["alertas"].extend(alerts)
                        cuentas_data[cuenta]["fuentes"].append(f"{role}: {file_obj.name}")
                    except Exception as e:
                        cuentas_data[cuenta]["errores"].append(f"{role}: {e}")

            render_carga_status(cuenta, cuentas_data[cuenta])

            if cuentas_data[cuenta]["fuentes"]:
                with st.expander("Fuentes asignadas"):
                    for src in cuentas_data[cuenta]["fuentes"]:
                        st.write(f"- {src}")

            if cuentas_data[cuenta]["alertas"]:
                with st.expander("Alertas de carga"):
                    for a in cuentas_data[cuenta]["alertas"]:
                        st.write(f"- {a}")

            if cuentas_data[cuenta]["errores"]:
                for err in cuentas_data[cuenta]["errores"]:
                    st.error(err)

            with st.expander("Preview"):
                if cuentas_data[cuenta]["posiciones"] is not None:
                    st.markdown("**Posiciones**")
                    st.dataframe(cuentas_data[cuenta]["posiciones"].head(8), use_container_width=True, hide_index=True)
                if cuentas_data[cuenta]["activity"] is not None:
                    st.markdown("**Activity**")
                    st.dataframe(cuentas_data[cuenta]["activity"].head(8), use_container_width=True, hide_index=True)

    st.markdown("---")
    ejecutar = st.button("Ejecutar conciliación premium")

    if not ejecutar:
        return

    resultados: List[pd.DataFrame] = []
    resumenes: List[Dict] = []
    auditorias: List[pd.DataFrame] = []
    errores_globales: List[str] = []

    for cuenta in cuentas_sel:
        payload = cuentas_data.get(cuenta, {})
        df_pos = payload.get("posiciones")
        df_act = payload.get("activity")
        df_ini = payload.get("portafolio_ini")
        df_fin = payload.get("portafolio_fin")

        if df_pos is None or df_act is None:
            errores_globales.append(f"Cta {cuenta}: faltan archivos obligatorios (Posiciones y/o Activity).")
            continue

        try:
            df_result, resumen_cta, df_aud = conciliador.conciliar_cuenta(
                cuenta=cuenta,
                df_posiciones=df_pos,
                df_activity=df_act,
                df_portafolio_ini=df_ini,
                df_portafolio_fin=df_fin,
            )
            resultados.append(df_result)
            resumenes.append(resumen_cta)
            if not df_aud.empty:
                auditorias.append(df_aud)
        except Exception as e:
            errores_globales.append(f"Cta {cuenta}: error al conciliar → {e}")

    if errores_globales:
        for err in errores_globales:
            st.error(err)

    df_all = conciliador.generar_reporte(resultados)
    if df_all.empty:
        st.warning("No se generaron resultados.")
        return

    df_resumen_cuentas = build_resumen_cuentas(df_all)
    df_top_pendientes = build_top_pendientes(df_all, top_n=20)
    df_pendientes = df_all[df_all["status"] == "PENDIENTE"].copy().reset_index(drop=True)
    df_auditoria = pd.concat(auditorias, ignore_index=True) if auditorias else pd.DataFrame(columns=["cuenta", "tipo_evento", "detalle", "impacto"])
    df_reglas = build_reglas_aplicadas(df_all)

    st.markdown("---")
    render_main_kpis(df_all)
    render_management_summary(df_resumen_cuentas, df_top_pendientes)
    render_semaforo(df_resumen_cuentas)

    tabs_out = st.tabs([
        "Resumen por cuenta",
        "Top pendientes",
        "Pendientes",
        "Detalle consolidado",
        "Auditoría",
        "Reglas",
    ])

    with tabs_out[0]:
        st.markdown('<div class="neix-section">Resumen por cuenta</div>', unsafe_allow_html=True)
        st.dataframe(
            df_resumen_cuentas.style.format({
                "especies": "{:,.0f}",
                "cerradas": "{:,.0f}",
                "pendientes": "{:,.0f}",
                "reglas_aplicadas": "{:,.0f}",
                "dif_total_abs": "{:,.2f}",
                "dif_total_neto": "{:,.2f}",
                "pct_cierre": "{:,.1f}%",
            }),
            use_container_width=True,
            hide_index=True,
        )

    with tabs_out[1]:
        render_top_pendientes(df_top_pendientes)

    with tabs_out[2]:
        st.markdown('<div class="neix-section">Pendientes</div>', unsafe_allow_html=True)
        if df_pendientes.empty:
            st.success("No hay pendientes.")
        else:
            st.dataframe(
                df_pendientes.style.format({
                    "ini": "{:,.2f}",
                    "act": "{:,.2f}",
                    "fin": "{:,.2f}",
                    "comp": "{:,.2f}",
                    "ini_ajust": "{:,.2f}",
                    "act_ajust": "{:,.2f}",
                    "fin_ajust": "{:,.2f}",
                    "comp_ajust": "{:,.2f}",
                    "dif_final": "{:,.2f}",
                    "dif_cant": "{:,.2f}",
                }),
                use_container_width=True,
                hide_index=True,
            )

    with tabs_out[3]:
        st.markdown('<div class="neix-section">Detalle consolidado</div>', unsafe_allow_html=True)
        st.dataframe(
            df_all.style.format({
                "ini": "{:,.2f}",
                "act": "{:,.2f}",
                "fin": "{:,.2f}",
                "comp": "{:,.2f}",
                "ini_ajust": "{:,.2f}",
                "act_ajust": "{:,.2f}",
                "fin_ajust": "{:,.2f}",
                "comp_ajust": "{:,.2f}",
                "dif_final_original": "{:,.2f}",
                "dif_final": "{:,.2f}",
                "dif_importe": "{:,.2f}",
            }),
            use_container_width=True,
            hide_index=True,
        )

    with tabs_out[4]:
        render_auditoria(df_auditoria)

    with tabs_out[5]:
        st.markdown('<div class="neix-section">Reglas aplicadas</div>', unsafe_allow_html=True)
        if df_reglas.empty:
            st.info("No se registraron reglas.")
        else:
            st.dataframe(df_reglas, use_container_width=True, hide_index=True)

    excel_bytes = build_excel_bytes(
        df_all=df_all,
        df_resumen_cuentas=df_resumen_cuentas,
        df_pendientes=df_pendientes,
        df_top_pendientes=df_top_pendientes,
        df_auditoria=df_auditoria,
        df_reglas=df_reglas,
    )
    nombre = f"conciliacion_cartera_premium_{str(fecha_fin).replace('-', '')}.xlsx"

    st.download_button(
        "Descargar Excel premium",
        data=excel_bytes,
        file_name=nombre,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
