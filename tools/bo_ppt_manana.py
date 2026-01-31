# ppt_manana.py
from __future__ import annotations

import io
import re
import datetime as dt
from dataclasses import dataclass
from typing import Optional, Tuple, Dict

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================
# Estética (minimal NEIX-like, sin depender de /static/css)
# ============================================================
def _inject_css():
    st.markdown(
        """
        <style>
          .neix-wrap{max-width:1200px;margin:0 auto;padding:6px 4px 18px 4px;}
          .neix-topbar{display:flex;justify-content:space-between;align-items:center;margin:10px 0 16px 0;}
          .neix-brand{display:flex;align-items:center;gap:12px;}
          .neix-logo{width:36px;height:36px;border-radius:12px;background:#0b1220;color:white;
                     display:flex;align-items:center;justify-content:center;font-weight:800;}
          .neix-title{margin:0;font-size:26px;font-weight:800;color:#111827;line-height:1.1;}
          .neix-sub{margin:2px 0 0 0;color:#6b7280;font-size:13px;}
          .neix-card{border:1px solid rgba(0,0,0,0.08);border-radius:18px;background:white;
                     padding:16px;box-shadow:0 10px 24px rgba(0,0,0,0.04);}
          .neix-row{display:grid;grid-template-columns:repeat(3, 1fr);gap:12px;}
          @media (max-width: 980px){ .neix-row{grid-template-columns:1fr;} }
          .neix-bad{display:inline-block;padding:8px 10px;border-radius:12px;background:#fff1f2;color:#991b1b;
                    border:1px solid rgba(153,27,27,0.15);font-weight:600;}
          .neix-ok{display:inline-block;padding:8px 10px;border-radius:12px;background:#ecfdf5;color:#065f46;
                   border:1px solid rgba(6,95,70,0.15);font-weight:600;}
          .neix-warn{display:inline-block;padding:8px 10px;border-radius:12px;background:#fffbeb;color:#92400e;
                     border:1px solid rgba(146,64,14,0.15);font-weight:600;}
          .neix-hr{height:1px;background:rgba(0,0,0,0.06);margin:14px 0;}
        </style>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# Utilidades de parsing / normalización (equivalentes a tu JS)
# ============================================================
def _canon(x) -> str:
    if x is None:
        return ""
    s = str(x)
    s = s.strip()
    # quitar tildes
    s = (
        s.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
        .replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
        .replace("Ñ", "N").replace("ñ", "n")
    )
    s = re.sub(r"\s+", " ", s).strip().upper()
    return s


def _coerce_number(v) -> Optional[float]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    # estilo AR: miles con . y decimales con ,
    s = s.replace(".", "").replace(",", ".")
    s = re.sub(r"\s+", "", s)
    try:
        n = float(s)
        return n
    except Exception:
        return None


def _format_date_short(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if isinstance(v, (dt.date, dt.datetime)):
        d = v.day
        m = v.month
        y = v.year
        return f"{d:02d}/{m:02d}/{y:04d}"
    s = str(v).strip()
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    return s


def _read_excel_first_sheet(file_bytes: bytes) -> pd.DataFrame:
    bio = io.BytesIO(file_bytes)
    # pandas elige engine según formato; para xls puede requerir xlrd instalado.
    # Si xlrd no está, conviene que suban xlsx o xlsb.
    df = pd.read_excel(bio, sheet_name=0, dtype=object)
    return df


def _detect_csv_delim(sample: str) -> str:
    sc = sample.count(";")
    cc = sample.count(",")
    return ";" if sc >= cc else ","


def _read_csv_robust(file_bytes: bytes) -> pd.DataFrame:
    text = file_bytes.decode("utf-8", errors="replace")
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")

    first_useful = ""
    for ln in lines:
        t = (ln or "").strip()
        if not t:
            continue
        if re.match(r"^[;,]+$", t):
            continue
        first_useful = ln
        break

    delim = _detect_csv_delim(first_useful or ";")

    # Filtramos líneas vacías o ";;;;;;"
    cleaned = []
    for ln in lines:
        t = (ln or "").strip()
        if not t:
            continue
        if re.match(r"^[;,]+$", t):
            continue
        cleaned.append(ln)

    cleaned_text = "\n".join(cleaned)
    bio = io.StringIO(cleaned_text)
    df = pd.read_csv(bio, sep=delim, dtype=str, engine="python")
    return df


def _with_header_row2(df: pd.DataFrame) -> pd.DataFrame:
    # tu JS: toma la fila 2 como header (index 1)
    if df.shape[0] < 2:
        return df
    new_header = df.iloc[1].astype(str).tolist()
    body = df.iloc[2:].copy()
    body.columns = new_header
    return body


# ============================================================
# Transformaciones principales
# ============================================================
def transform_moc(df_raw: pd.DataFrame) -> pd.DataFrame:
    df = _with_header_row2(df_raw)
    df.columns = [str(c).strip() for c in df.columns]

    cols = {_canon(c): c for c in df.columns}
    c_codigo = cols.get("CODIGO")
    c_comit  = cols.get("COMITENTE")
    c_comp   = cols.get("COMPRAS")
    c_vent   = cols.get("VENTAS")

    out = pd.DataFrame({
        "Comitente": df[c_comit].apply(_coerce_number) if c_comit else None,
        "Codigo":    df[c_codigo] if c_codigo else "",
        "Compras":   df[c_comp].apply(_coerce_number) if c_comp else 0,
        "Ventas":    df[c_vent].apply(_coerce_number) if c_vent else 0,
    })

    out["MAE"] = 0.0
    out["Neto"] = (out["Compras"].fillna(0) + out["Ventas"].fillna(0) + out["MAE"].fillna(0))
    out["Observaciones"] = ""
    return out


def build_mae_map(df_mae: pd.DataFrame) -> Dict[str, float]:
    # Busca header real que incluya COMITENTE/CODIGO/COMPRA/VENTA
    df = df_mae.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # normalizar columnas
    cols = {_canon(c): c for c in df.columns}
    needed = ["COMITENTE", "CODIGO", "COMPRA", "VENTA"]
    if not all(k in cols for k in needed):
        return {}

    c_com  = cols["COMITENTE"]
    c_cod  = cols["CODIGO"]
    c_comp = cols["COMPRA"]
    c_vent = cols["VENTA"]

    m: Dict[str, float] = {}
    for _, r in df.iterrows():
        com = _coerce_number(r.get(c_com))
        cod = r.get(c_cod)
        if com is None or cod is None or str(cod).strip() == "":
            continue
        key = f"{int(com)}|{_canon(cod)}"
        compra = _coerce_number(r.get(c_comp)) or 0.0
        venta  = _coerce_number(r.get(c_vent)) or 0.0
        m[key] = m.get(key, 0.0) + (compra + venta)
    return m


def apply_mae_to_moc(moc: pd.DataFrame, mae_map: Dict[str, float]) -> pd.DataFrame:
    if moc is None or moc.empty:
        return moc
    moc = moc.copy()
    maes = []
    for _, r in moc.iterrows():
        com = r.get("Comitente")
        cod = r.get("Codigo")
        key = f"{int(com) if pd.notna(com) else ''}|{_canon(cod)}"
        maes.append(mae_map.get(key, 0.0))
    moc["MAE"] = maes
    moc["Neto"] = moc["Compras"].fillna(0) + moc["Ventas"].fillna(0) + moc["MAE"].fillna(0)
    return moc


def transform_neg(df_raw: pd.DataFrame) -> pd.DataFrame:
    df = _with_header_row2(df_raw)
    df.columns = [str(c).strip() for c in df.columns]
    cols = {_canon(c): c for c in df.columns}

    c_codigo = cols.get("CODIGO")
    c_comit  = cols.get("COMITENTE")
    c_cant   = cols.get("CANTIDAD")
    c_neto   = cols.get("NETO")
    c_sneg   = cols.get("SALDO NEGATIVOS")

    out = pd.DataFrame({
        "Comitente": df[c_comit].apply(_coerce_number) if c_comit else None,
        "Codigo":    df[c_codigo].apply(_coerce_number) if c_codigo else None,
        "Neto":      df[c_neto].apply(_coerce_number) if c_neto else None,
        "Cantidad":  df[c_cant].apply(_coerce_number) if c_cant else None,
        "Saldo Negativos": df[c_sneg].apply(_coerce_number) if c_sneg else None,
    })
    return out


def map_byma_comitente(x):
    n = _coerce_number(x)
    if n is None:
        return None
    n = int(n)
    if n in (3, 777777777, 888888888):
        return 999
    return n


def build_byma_map(df_raw: pd.DataFrame) -> Dict[str, float]:
    df = _with_header_row2(df_raw)
    df.columns = [str(c).strip() for c in df.columns]
    cols = {_canon(c): c for c in df.columns}

    c_com = cols.get("COMITENTE")
    c_cod = cols.get("CODIGO TV")
    c_tot = cols.get("TOTAL")

    m: Dict[str, float] = {}
    if not (c_com and c_cod and c_tot):
        return m

    for _, r in df.iterrows():
        com = map_byma_comitente(r.get(c_com))
        cod = _coerce_number(r.get(c_cod))
        tot = _coerce_number(r.get(c_tot)) or 0.0
        if com is None or cod is None:
            continue
        key = f"{int(com)}|{int(cod)}"
        m[key] = m.get(key, 0.0) + tot
    return m


def merge_byma_to_neg(
    neg: pd.DataFrame,
    byma_map: Dict[str, float],
    mae_map: Dict[str, float],
    use_mae_for_otros: bool,
) -> pd.DataFrame:
    neg = neg.copy()
    garantias = []
    otros = []
    total = []
    obs = []

    for _, r in neg.iterrows():
        com = _coerce_number(r.get("Comitente"))
        cod = _coerce_number(r.get("Codigo"))
        saldo = _coerce_number(r.get("Saldo Negativos")) or 0.0

        by = 0.0
        if com is not None and cod is not None:
            by = byma_map.get(f"{int(com)}|{int(cod)}", 0.0)

        ot = 0.0
        if use_mae_for_otros and mae_map and com is not None and cod is not None:
            key_mae = f"{int(com)}|{_canon(int(cod))}"
            # tu JS usa canon(cod) (cod viene numérico), esto lo replicamos
            ot = mae_map.get(key_mae, 0.0)

        tt = saldo + by + ot
        garantias.append(by)
        otros.append(ot)
        total.append(tt)
        obs.append("OK" if tt >= 0 else "REVISAR")

    neg["Garantías BYMA"] = garantias
    neg["OTROS"] = otros
    neg["Total"] = total
    neg["Observación"] = obs
    return neg


def build_neix_map(df_raw: pd.DataFrame) -> Dict[int, Tuple[str, float]]:
    df = _with_header_row2(df_raw)
    df.columns = [str(c).strip() for c in df.columns]
    cols = {_canon(c): c for c in df.columns}

    c_code = cols.get("CSVA CODE")
    c_sym  = cols.get("SYMBOL")
    c_qty  = cols.get("QTY BYMA")
    c_tipo = cols.get("TIPO")
    c_op   = cols.get("OPERACION")

    m: Dict[int, Tuple[str, float]] = {}
    if not c_code:
        return m

    for _, r in df.iterrows():
        code = _coerce_number(r.get(c_code))
        if code is None:
            continue
        code_i = int(code)

        tipo = _canon(r.get(c_tipo)) if c_tipo else ""
        op   = _canon(r.get(c_op))   if c_op   else ""

        allowed = (tipo == "CEDEAR" and op == "ISSUE") or (tipo == "ADR" and op == "CXL")
        qty_src = _coerce_number(r.get(c_qty)) if c_qty else 0.0
        qty_row = abs(qty_src or 0.0) if allowed else 0.0

        prev_sym, prev_qty = m.get(code_i, ("", 0.0))
        sym = prev_sym or (str(r.get(c_sym)) if c_sym else "")
        m[code_i] = (sym, prev_qty + qty_row)

    return m


def apply_neix_to_neg992(neg992: pd.DataFrame, neix_map: Dict[int, Tuple[str, float]]) -> pd.DataFrame:
    neg992 = neg992.copy()
    symbols = []
    qtys = []
    for _, r in neg992.iterrows():
        cod = _coerce_number(r.get("Codigo"))
        if cod is None:
            symbols.append("")
            qtys.append(0.0)
            continue
        sym, qty = neix_map.get(int(cod), ("", 0.0))
        symbols.append(sym)
        qtys.append(abs(qty or 0.0))
    neg992["Symbol"] = symbols
    neg992["Qty BYMA"] = qtys
    # reorden como tu JS
    cols = ["Comitente","Symbol","Codigo","Neto","Cantidad","Saldo Negativos","Qty BYMA","Garantías BYMA","OTROS","Total","Observación"]
    for c in cols:
        if c not in neg992.columns:
            neg992[c] = None
    return neg992[cols]


def build_moc_neto_map_by_code(moc: pd.DataFrame) -> Dict[str, float]:
    # tu JS: por CODIGO suma COMPRAS+VENTAS (sin MAE)
    m: Dict[str, float] = {}
    if moc is None or moc.empty:
        return m
    for _, r in moc.iterrows():
        code_raw = r.get("Codigo")
        code_key = str(code_raw).strip()
        if code_key == "":
            continue
        comp = _coerce_number(r.get("Compras")) or 0.0
        vent = _coerce_number(r.get("Ventas")) or 0.0
        m[code_key] = m.get(code_key, 0.0) + (comp + vent)
    return m


def transform_sliq_with_moc(df_sliq: pd.DataFrame, moc: pd.DataFrame) -> pd.DataFrame:
    if df_sliq is None or df_sliq.empty:
        return df_sliq

    sliq = df_sliq.copy()
    sliq.columns = [str(c).strip() for c in sliq.columns]
    Hc = [_canon(c) for c in sliq.columns]

    # columna CODIGO: primera que contenga 'CODIGO'
    i_code = next((i for i, c in enumerate(Hc) if "CODIGO" in c), None)

    # columna Neto a Liquidar (exacto)
    i_neto = next((i for i, c in enumerate(Hc) if c == "NETO A LIQUIDAR"), None)

    # dropear columnas
    drop = {"FALTANTE","ADELANTADAS","LIBERADAS","FALTAN LIBERAR","PND"}
    keep_cols = [c for c in sliq.columns if _canon(c) not in drop]
    sliq = sliq[keep_cols].copy()

    moc_map = build_moc_neto_map_by_code(moc)

    # Agregar Q NASDAQ
    qvals = []
    seen = set()
    code_col_name = None
    if i_code is not None:
        code_col_name = df_sliq.columns[i_code]

    for _, r in sliq.iterrows():
        code = str(r.get(code_col_name, "")).strip() if code_col_name else ""
        if code:
            seen.add(code)
        qvals.append(moc_map.get(code, 0.0))

    sliq["Q NASDAQ"] = qvals

    # Control = Neto a Liquidar - Q NASDAQ (si existe neto)
    neto_col_name = None
    if i_neto is not None:
        neto_col_name = df_sliq.columns[i_neto]
        # si neto fue dropeado, intentamos recuperarlo si está en keep_cols
        if neto_col_name not in sliq.columns and neto_col_name in df_sliq.columns:
            # si fue dropeado por error, no hacemos nada; pero normalmente no se dropea
            pass

    if neto_col_name and neto_col_name in sliq.columns:
        neto_num = sliq[neto_col_name].apply(_coerce_number).fillna(0.0)
    else:
        neto_num = pd.Series([0.0]*len(sliq))

    sliq["Control"] = neto_num - sliq["Q NASDAQ"].fillna(0.0)
    sliq["Observación"] = sliq["Control"].apply(lambda x: "OK" if (float(x or 0.0) == 0.0) else "REVISAR")

    # Agregar filas faltantes desde MOC (códigos no vistos)
    if code_col_name and code_col_name in sliq.columns:
        missing_rows = []
        for code, q in moc_map.items():
            if code in seen:
                continue
            row = {c: "" for c in sliq.columns}
            row[code_col_name] = code
            row["Q NASDAQ"] = q
            row["Control"] = (0.0 - q)
            row["Observación"] = "REVISAR" if q != 0 else "OK"
            missing_rows.append(row)
        if missing_rows:
            sliq = pd.concat([sliq, pd.DataFrame(missing_rows)], ignore_index=True)

    # ordenar: los OK (Control=0) al final, como tu sort (flag OK=1)
    def _flag_ok(v):
        try:
            return 1 if float(v) == 0.0 else 0
        except Exception:
            return 0

    sliq = sliq.sort_values(by="Control", key=lambda s: s.map(_flag_ok), ascending=True)
    return sliq


def transform_asign(df_raw: pd.DataFrame) -> pd.DataFrame:
    df = df_raw.copy()
    df.columns = [str(c).strip() for c in df.columns]
    cols = {_canon(c): c for c in df.columns}

    required = [
        "Nro. Mercado Corto","Moneda","Cantidad","Monto","Fecha Concertación","Fecha Liquidación","CUIT",
        "Negociación","Participante contraparte","Compra - Venta","Código Especie","Código CV","Código Liquidación"
    ]
    req_canon = [_canon(x) for x in required]
    idx = [cols.get(rc) for rc in req_canon]

    out = pd.DataFrame()
    for name, src in zip(required, idx):
        out[name] = df[src] if src else ""

    out["Observacion"] = ""

    # normalizar num/fechas como en JS
    for c in ["Cantidad","Monto","Código CV","Código Liquidación"]:
        out[c] = out[c].apply(_coerce_number)
    for c in ["Fecha Concertación","Fecha Liquidación"]:
        out[c] = out[c].apply(_format_date_short)

    return out


# ============================================================
# Export Excel con estilos (freeze + formato contable)
# ============================================================
ACCOUNTING_FMT = '_-* #,##0_-;_* -#,##0_-;_-* "-"_-;_-@_-'

def _df_to_excel_bytes(
    moc: pd.DataFrame,
    asign: Optional[pd.DataFrame],
    neg: pd.DataFrame,
    neg992: pd.DataFrame,
    sliq: Optional[pd.DataFrame],
    filename_hint: str,
) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        moc.to_excel(writer, index=False, sheet_name="MOC")
        if asign is not None and not asign.empty:
            asign.to_excel(writer, index=False, sheet_name="ASIGNACIONES_PENDIENTES")
        neg.to_excel(writer, index=False, sheet_name="SALDOS NEGATIVOS")
        neg992.to_excel(writer, index=False, sheet_name="SALDOS NEGATIVOS 992")
        if sliq is not None and not sliq.empty:
            sliq.to_excel(writer, index=False, sheet_name="Control SLIQ")

    output.seek(0)
    wb = load_workbook(output)

    def freeze(ws):
        ws.freeze_panes = "A2"

    def apply_accounting(ws, colnames):
        headers = [cell.value for cell in ws[1]]
        hc = [_canon(h) for h in headers]
        idxs = []
        for name in colnames:
            n = _canon(name)
            if n in hc:
                idxs.append(hc.index(n) + 1)  # 1-based openpyxl
        if not idxs:
            return
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            for c in idxs:
                ws.cell(row=r, column=c).number_format = ACCOUNTING_FMT

    # MOC
    ws = wb["MOC"]
    freeze(ws)
    apply_accounting(ws, ["Compras","Ventas","MAE","Neto"])

    if "ASIGNACIONES_PENDIENTES" in wb.sheetnames:
        ws = wb["ASIGNACIONES_PENDIENTES"]
        freeze(ws)
        apply_accounting(ws, ["Cantidad","Monto","Código CV","Código Liquidación"])

    ws = wb["SALDOS NEGATIVOS"]
    freeze(ws)
    apply_accounting(ws, ["Neto","Cantidad","Saldo Negativos","Garantías BYMA","OTROS","Total"])

    ws = wb["SALDOS NEGATIVOS 992"]
    freeze(ws)
    apply_accounting(ws, ["Neto","Cantidad","Saldo Negativos","Qty BYMA","Garantías BYMA","OTROS","Total"])

    if "Control SLIQ" in wb.sheetnames:
        ws = wb["Control SLIQ"]
        freeze(ws)
        # Control / Neto / Q
        apply_accounting(ws, ["NETO A LIQUIDAR","Q NASDAQ","Control"])

    out2 = io.BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2.read()


def _ventas_excel_bytes(moc: pd.DataFrame) -> Optional[Tuple[bytes, str]]:
    if moc is None or moc.empty:
        return None
    ventas = moc.copy()
    ventas["Ventas"] = ventas["Ventas"].apply(_coerce_number).fillna(0.0)
    ventas = ventas[ventas["Ventas"] != 0][["Comitente","Codigo","Ventas"]]
    if ventas.empty:
        return None
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        ventas.to_excel(writer, index=False, sheet_name="VENTAS")
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb["VENTAS"]
    ws.freeze_panes = "A2"
    # formato contable en Ventas
    headers = [c.value for c in ws[1]]
    hc = [_canon(h) for h in headers]
    if "VENTAS" in hc:
        col = hc.index("VENTAS") + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = ACCOUNTING_FMT
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), "VENTAS"


# ============================================================
# Pipeline state
# ============================================================
@dataclass
class Result:
    moc: pd.DataFrame
    asign: Optional[pd.DataFrame]
    neg: pd.DataFrame
    neg992: pd.DataFrame
    sliq: Optional[pd.DataFrame]
    tags: list[Tuple[str, str]]  # (msg, level ok/warn/bad)


def _tag(msg: str, lvl: str) -> Tuple[str, str]:
    return (msg, lvl)


def process_all(
    file_moc: bytes,
    file_neg: bytes,
    file_mae: Optional[bytes],
    file_asig: Optional[bytes],
    file_lista: Optional[bytes],
    file_neix: Optional[bytes],
    file_sliq: Optional[bytes],
) -> Result:
    tags: list[Tuple[str, str]] = []

    # MOC
    df_moc_raw = _read_excel_first_sheet(file_moc)
    moc = transform_moc(df_moc_raw)
    tags.append(_tag("MOC: procesado.", "ok"))

    # MAE (csv)
    mae_map: Dict[str, float] = {}
    if file_mae:
        try:
            df_mae = _read_csv_robust(file_mae)
            mae_map = build_mae_map(df_mae)
            tags.append(_tag(f"MOC MAE: leído (keys={len(mae_map)}).", "ok"))
        except Exception as e:
            tags.append(_tag(f"Error leyendo MOC MAE: {e}", "warn"))
    else:
        tags.append(_tag("MOC MAE no adjuntado: MAE=0.", "warn"))

    moc = apply_mae_to_moc(moc, mae_map)

    # Saldos Negativos
    df_neg_raw = _read_excel_first_sheet(file_neg)
    neg_all = transform_neg(df_neg_raw)
    tags.append(_tag("Saldos Negativos: procesado (split 992).", "ok"))

    neg992 = neg_all[neg_all["Comitente"].fillna(0).astype(int) == 992].copy()
    neg = neg_all[neg_all["Comitente"].fillna(0).astype(int) != 992].copy()

    # BYMA + MAE->OTROS solo en neg (otros)
    if file_lista:
        try:
            df_lista = _read_excel_first_sheet(file_lista)
            byma_map = build_byma_map(df_lista)

            # En SALDOS NEGATIVOS (OTROS) aplicamos MAE->OTROS si hay match
            neg = merge_byma_to_neg(neg, byma_map, mae_map, use_mae_for_otros=True)

            # En 992 NO aplicamos MAE->OTROS
            neg992 = merge_byma_to_neg(neg992, byma_map, mae_map, use_mae_for_otros=False)

            tags.append(_tag("BYMA: merge aplicado (MAE->OTROS solo en SALDOS NEGATIVOS).", "ok"))
        except Exception as e:
            tags.append(_tag(f"Error leyendo/merging BYMA: {e}", "warn"))
    else:
        tags.append(_tag("No se adjuntó Lista BYMA: sin BYMA/OTROS/Total.", "warn"))
        # aun así armamos columnas mínimas para que exporte prolijo
        for df_ in (neg, neg992):
            for c in ["Garantías BYMA","OTROS","Total","Observación"]:
                if c not in df_.columns:
                    df_[c] = 0.0 if c != "Observación" else ""

    # Emisiones 992
    if file_neix:
        try:
            df_neix = _read_excel_first_sheet(file_neix)
            neix_map = build_neix_map(df_neix)
            # aplicar en 992
            neg992 = apply_neix_to_neg992(neg992, neix_map)
            tags.append(_tag("Emisiones 992: agregado Symbol y Qty BYMA.", "ok"))
        except Exception as e:
            tags.append(_tag(f"Error leyendo/merging Emisiones 992: {e}", "warn"))
            # asegurar columnas
            if "Symbol" not in neg992.columns:
                neg992["Symbol"] = ""
            if "Qty BYMA" not in neg992.columns:
                neg992["Qty BYMA"] = 0.0
    else:
        tags.append(_tag("No se adjuntó Emisiones 992: no se agrega Symbol/Qty BYMA.", "warn"))
        if "Symbol" not in neg992.columns:
            neg992["Symbol"] = ""
        if "Qty BYMA" not in neg992.columns:
            neg992["Qty BYMA"] = 0.0
        # reorden igual
        neg992 = apply_neix_to_neg992(neg992, {})

    # Asignaciones (opcional)
    asign = None
    if file_asig:
        try:
            df_asig_raw = _read_excel_first_sheet(file_asig)
            asign = transform_asign(df_asig_raw)
            tags.append(_tag("Asignaciones: procesado.", "ok"))
        except Exception as e:
            tags.append(_tag(f"Error leyendo Asignaciones: {e}", "warn"))
            asign = None
    else:
        tags.append(_tag("Asignaciones no adjuntado: se exporta sin esa hoja.", "warn"))

    # SLIQ (opcional)
    sliq = None
    if file_sliq:
        try:
            df_sliq = _read_csv_robust(file_sliq)
            sliq = transform_sliq_with_moc(df_sliq, moc)
            tags.append(_tag("Control SLIQ: procesado.", "ok"))
        except Exception as e:
            tags.append(_tag(f"Error leyendo Control SLIQ: {e}", "warn"))
            sliq = None
    else:
        tags.append(_tag("Control SLIQ no adjuntado: se exporta sin esa hoja.", "warn"))

    # Ordenes similares a tu JS:
    # - MOC: por Codigo y luego Comitente
    def _sort_moc(m: pd.DataFrame) -> pd.DataFrame:
        mm = m.copy()
        mm["__cod_sort__"] = mm["Codigo"].apply(lambda x: _coerce_number(x) if _coerce_number(x) is not None else str(x))
        mm["__com_sort__"] = mm["Comitente"].fillna(0)
        mm = mm.sort_values(by=["__cod_sort__", "__com_sort__"], ascending=[True, True])
        return mm.drop(columns=["__cod_sort__", "__com_sort__"])

    moc = _sort_moc(moc)

    # - Neg: por Total asc
    if "Total" in neg.columns:
        neg = neg.sort_values(by=["Total"], ascending=True)

    # - Neg992: por total computado asc
    if all(c in neg992.columns for c in ["Saldo Negativos","Qty BYMA","Garantías BYMA","OTROS"]):
        neg992["_t"] = (
            neg992["Saldo Negativos"].fillna(0)
            + neg992["Qty BYMA"].fillna(0)
            + neg992["Garantías BYMA"].fillna(0)
            + neg992["OTROS"].fillna(0)
        )
        neg992 = neg992.sort_values(by=["_t"], ascending=True).drop(columns=["_t"])

    # Validación mínima
    if moc.empty or neg.empty or neg992.empty:
        tags.append(_tag("Faltan datos para exportar (MOC/Neg/Neg992).", "bad"))
    else:
        tags.append(_tag("Listo para exportar.", "ok"))

    return Result(moc=moc, asign=asign, neg=neg, neg992=neg992, sliq=sliq, tags=tags)


# ============================================================
# UI principal
# ============================================================
def render():
    _inject_css()

    st.markdown(
        """
        <div class="neix-wrap">
          <div class="neix-topbar">
            <div class="neix-brand">
              <div class="neix-logo">N</div>
              <div>
                <div class="neix-title">☀️ MOC MAÑANA</div>
                <div class="neix-sub">Papel de Trabajo Definitivo</div>
              </div>
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="neix-card">', unsafe_allow_html=True)

    st.markdown("### Subí los archivos")
    c1, c2, c3 = st.columns(3)
    with c1:
        up_moc = st.file_uploader("MOC (xlsx)", type=["xlsx"], key="moc")
        up_mae = st.file_uploader("MOC MAE (csv) [opcional]", type=["csv"], key="mae")
        up_asg = st.file_uploader("Asignaciones Pendientes (xls/xlsx/xlsb) [opcional]", type=["xls","xlsx","xlsb"], key="asig")
    with c2:
        up_neg = st.file_uploader("Saldos Negativos (xlsx)", type=["xlsx"], key="neg")
        up_lst = st.file_uploader("Lista de Saldos BYMA (xls/xlsx/xlsb)", type=["xls","xlsx","xlsb"], key="lista")
        up_nei = st.file_uploader("Emisiones 992 (xlsx) [recomendado]", type=["xlsx"], key="neix")
    with c3:
        up_sliq = st.file_uploader("Control SLIQ (csv) [opcional]", type=["csv"], key="sliq")

    st.markdown('<div class="neix-hr"></div>', unsafe_allow_html=True)

    colA, colB, colC = st.columns([1,1,2])
    with colA:
        do_validar = st.button("Validar / Procesar", use_container_width=True)
    with colB:
        st.caption("Después de validar, se habilitan las descargas.")

    if "ppt_res" not in st.session_state:
        st.session_state["ppt_res"] = None

    if do_validar:
        # chequeos mínimos
        if up_moc is None:
            st.error("Falta archivo MOC.")
        elif up_neg is None:
            st.error("Falta archivo Saldos Negativos.")
        else:
            res = process_all(
                file_moc=up_moc.getvalue(),
                file_neg=up_neg.getvalue(),
                file_mae=up_mae.getvalue() if up_mae else None,
                file_asig=up_asg.getvalue() if up_asg else None,
                file_lista=up_lst.getvalue() if up_lst else None,
                file_neix=up_nei.getvalue() if up_nei else None,
                file_sliq=up_sliq.getvalue() if up_sliq else None,
            )
            st.session_state["ppt_res"] = res

    res: Result | None = st.session_state.get("ppt_res")

    # Tags estado
    if res is not None:
        st.markdown("### Estado")
        for msg, lvl in res.tags:
            cls = {"ok":"neix-ok","warn":"neix-warn","bad":"neix-bad"}.get(lvl, "neix-warn")
            st.markdown(f'<span class="{cls}">{msg}</span>', unsafe_allow_html=True)

        st.markdown('<div class="neix-hr"></div>', unsafe_allow_html=True)

        # Preview chico
        with st.expander("Ver preview (MOC / Saldos)", expanded=False):
            st.write("MOC (primeras 30 filas)")
            st.dataframe(res.moc.head(30), use_container_width=True)
            st.write("Saldos Negativos (primeras 30 filas)")
            st.dataframe(res.neg.head(30), use_container_width=True)
            st.write("Saldos Negativos 992 (primeras 30 filas)")
            st.dataframe(res.neg992.head(30), use_container_width=True)
            if res.sliq is not None:
                st.write("Control SLIQ (primeras 30 filas)")
                st.dataframe(res.sliq.head(30), use_container_width=True)

        # Downloads
        hoy = dt.date.today()
        dd = f"{hoy.day:02d}"
        mm = f"{hoy.month:02d}"
        yyyy = f"{hoy.year:04d}"
        nombre_moc = f"MOC {dd}-{mm}-{yyyy} DEFINITIVO.xlsx"
        nombre_ventas = f"VENTAS {dd}-{mm}-{yyyy}.xlsx"

        excel_bytes = _df_to_excel_bytes(
            moc=res.moc,
            asign=res.asign,
            neg=res.neg,
            neg992=res.neg992,
            sliq=res.sliq,
            filename_hint=nombre_moc,
        )

        st.markdown("### Exportación")
        d1, d2, _ = st.columns([1,1,2])
        with d1:
            st.download_button(
                "⬇️ Exportar MOC definitivo",
                data=excel_bytes,
                file_name=nombre_moc,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with d2:
            ventas_pack = _ventas_excel_bytes(res.moc)
            if ventas_pack is None:
                st.download_button(
                    "⬇️ Exportar VENTAS",
                    data=b"",
                    file_name=nombre_ventas,
                    disabled=True,
                    help="No hay ventas distintas de 0 para exportar.",
                    use_container_width=True
                )
            else:
                ventas_bytes, _ = ventas_pack
                st.download_button(
                    "⬇️ Exportar VENTAS",
                    data=ventas_bytes,
                    file_name=nombre_ventas,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    st.markdown("</div>", unsafe_allow_html=True)
